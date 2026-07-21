#Librerias estandar de Python
import json
import logging
import unicodedata
import csv
import io
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

#Herramientas base de Django
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db import transaction
from django.db.models import Q, Sum, Count, Max
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt

from clinica.utils import sincronizar_google_sheet
from .api_auth import api_key_required
from django.urls import reverse

#Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

#Modelos, Formularios y Utilidades de nuestra App
from .models import (
    AccesoDirectoPortal,
    AperturaExpediente,
    AperturaExpedienteGrupal,
    BloqueoAgendaTerapeuta,
    Consultoria,
    Consultorio,
    SolicitudHorasExpositor,
    Division,
    DocumentoPaciente,
    Empresa,
    ExpedienteGrupal,
    Host,
    HostChecklistTask,
    NotaExpedienteGrupal,
    Paciente,
    PacienteTerapeutaAcceso,
    NotaTerapeutaPaciente,
    ReporteSesion,
    Terapeuta,
    Cita,
    ConfiguracionWhatsApp,
    MensajeWhatsApp,
    MensajeWhatsAppEntrante,
    MensajeWhatsAppDemo,
    Horario,
    SolicitudCita,
    SolicitudReagendo,
    Servicio,
    NotificacionTerapeuta,
    obtener_bloqueo_terapeuta_en_fecha,
    ReporteIncidente,
    RecursoPropio,
    RegistroActividad,
)
from .models import CorteSemanal, LineaNomina, BonoExtra
from .models import NotaRecepcion, ReaccionNota, ComentarioNota
from .models import PerfilCatalogo
from .models import Instrumento, PreguntaInstrumento, EnvioInstrumento, RespuestaInstrumento
from .forms import (
    AperturaExpedienteForm,
    AperturaExpedienteGrupalForm,
    BloqueoAgendaTerapeutaForm,
    CitaEmpresaForm,
    PacienteEmpresaForm,
    PacienteForm,
    CitaForm,
    DocumentoPacienteForm,
    ManualPortalForm,
    NotaTerapeutaPacienteForm,
    CheckoutCitaForm,
    ReporteSesionForm,
)
from .services import calcular_nomina_semanal, preview_nomina_semanal, aprobar_corte_semanal, registrar_pago_penalizacion_terapeuta
from .services_instrumentos import (calcular_resultado_instrumento, calcular_raven,
                                     _RAVEN_SERIES, _RAVEN_ESPERADOS,
                                     _SCL90_REF, _SCL90_REF_SUMMARY)
from . import services_whatsapp as wa
#from .utils import sincronizar_google_sheet

def _registrar_actividad(request, accion, categoria, descripcion, terapeuta=None, paciente=None):
    ip = request.META.get('REMOTE_ADDR') or None
    RegistroActividad.objects.create(
        usuario=request.user if request.user.is_authenticated else None,
        accion=accion,
        categoria=categoria,
        descripcion=descripcion,
        terapeuta_afectado=terapeuta,
        paciente_afectado=paciente,
        ip_address=ip,
    )


def quitar_tildes(texto):
    if not texto:
        return ""
    return ''.join(c for c in unicodedata.normalize('NFD', str(texto))
                   if unicodedata.category(c) != 'Mn').lower()


def resolver_paciente_por_nombre(nombre):
    nombre_normalizado = quitar_tildes(nombre).strip()
    if not nombre_normalizado:
        return None

    paciente = Paciente.objects.filter(nombre__iexact=str(nombre).strip()).first()
    if paciente:
        return paciente

    paciente = Paciente.objects.filter(nombre_normalizado=nombre_normalizado).first()
    if paciente:
        return paciente

    tokens = [token for token in nombre_normalizado.split() if token]
    if tokens:
        filtro_normalizado = Q()
        filtro_nombre = Q()
        for token in tokens:
            filtro_normalizado &= Q(nombre_normalizado__icontains=token)
            filtro_nombre &= Q(nombre__icontains=token)
        paciente = Paciente.objects.filter(
            filtro_normalizado | filtro_nombre
        ).order_by('nombre').first()
        if paciente:
            return paciente

    return Paciente.objects.filter(
        Q(nombre_normalizado__icontains=nombre_normalizado) |
        Q(nombre__icontains=str(nombre).strip())
    ).order_by('nombre').first()


# ─── Excel helper ─────────────────────────────────────────────────────────────
_TEAL_FILL  = PatternFill("solid", fgColor="26C6DA")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_BODY_FONT   = Font(name="Calibri", size=10)
_ALT_FILL    = PatternFill("solid", fgColor="F0F4F8")

def _build_excel_response(filename, sheet_title, headers, rows, col_widths=None):
    """
    Genera un HttpResponse con un .xlsx estilizado.
    headers: lista de str
    rows: lista de listas (valores de celda)
    col_widths: lista opcional de anchos de columna
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Fila de encabezado
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _TEAL_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    # Filas de datos
    for row_idx, row in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _BODY_FONT
            if fill:
                cell.fill = fill

    # Anchos de columna
    if col_widths:
        for idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w
    else:
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Freeze panes (encabezado fijo)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
# ──────────────────────────────────────────────────────────────────────────────


SERVICIOS_GRUPALES = {
    'terapia de pareja',
    'terapia familiar',
    'terapia infantil',
}


def es_servicio_grupal(servicio):
    nombre = quitar_tildes(servicio.nombre if servicio else '').strip()
    return any(base in nombre for base in SERVICIOS_GRUPALES)


def _vincular_expediente_grupal(cita):
    """Busca o crea un ExpedienteGrupal para una cita con pacientes_adicionales."""
    pacientes_adicionales = cita.pacientes_adicionales.all()
    if not pacientes_adicionales.exists():
        if cita.expediente_grupal_id:
            cita.expediente_grupal = None
            cita.save(update_fields=['expediente_grupal'])
        return

    todos_ids = frozenset(
        [cita.paciente_id] + list(pacientes_adicionales.values_list('id', flat=True))
    )

    expediente = None
    for eg in ExpedienteGrupal.objects.prefetch_related('pacientes'):
        if frozenset(eg.pacientes.values_list('id', flat=True)) == todos_ids:
            expediente = eg
            break

    if not expediente:
        tipo = ExpedienteGrupal.TIPO_PAREJA if len(todos_ids) == 2 else ExpedienteGrupal.TIPO_FAMILIA
        expediente = ExpedienteGrupal.objects.create(tipo=tipo)
        expediente.pacientes.set(todos_ids)
        expediente.nombre = expediente.generar_nombre()
        expediente.save(update_fields=['nombre'])

    if cita.expediente_grupal_id != expediente.id:
        cita.expediente_grupal = expediente
        cita.save(update_fields=['expediente_grupal'])


def obtener_configuracion_estatus_cita(estatus):
    configuraciones = {
        Cita.ESTATUS_CONFIRMADA: {
            'label': 'Confirmada',
            'color': '#22c55e',
            'bg': '#dcfce7',
            'border': '#16a34a',
            'text': '#14532d',
        },
        Cita.ESTATUS_SIN_CONFIRMAR: {
            'label': 'Sin confirmar',
            'color': '#06b6d4',
            'bg': '#cffafe',
            'border': '#0891b2',
            'text': '#164e63',
        },
        Cita.ESTATUS_REAGENDO: {
            'label': 'Reagendo',
            'color': '#f59e0b',
            'bg': '#fef3c7',
            'border': '#d97706',
            'text': '#78350f',
        },
        Cita.ESTATUS_CANCELO: {
            'label': 'Cancelo',
            'color': '#ef4444',
            'bg': '#fee2e2',
            'border': '#dc2626',
            'text': '#7f1d1d',
        },
        Cita.ESTATUS_SI_ASISTIO: {
            'label': 'Si asistio',
            'color': '#16a34a',
            'bg': '#dcfce7',
            'border': '#15803d',
            'text': '#14532d',
        },
        Cita.ESTATUS_NO_ASISTIO: {
            'label': 'No asistio',
            'color': '#7c3aed',
            'bg': '#ede9fe',
            'border': '#6d28d9',
            'text': '#4c1d95',
        },
        Cita.ESTATUS_INCIDENCIA: {
            'label': 'Incidencia',
            'color': '#64748b',
            'bg': '#e2e8f0',
            'border': '#475569',
            'text': '#1e293b',
        },
    }
    return configuraciones.get(
        estatus,
        {
            'label': estatus or 'Sin estatus',
            'color': '#334155',
            'bg': '#e2e8f0',
            'border': '#334155',
            'text': '#0f172a',
        },
    )

def _sin_reagendar_stats():
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    ayer = hoy - timedelta(days=1)
    fecha_inicio_query = min(inicio_mes, inicio_semana, ayer)

    excluir = set(
        Cita.objects.filter(
            fecha__gte=hoy,
            estatus__in=Cita.ESTATUS_ACTIVOS,
            paciente__isnull=False, # dar de alta a un paciente
        ).values_list('paciente_id', flat=True)
    ) | set(
        SolicitudCita.objects.filter(
            estado='pendiente',
            paciente__isnull=False,
        ).values_list('paciente_id', flat=True)
    )

    citas = list(
        Cita.objects.filter(
            fecha__range=(fecha_inicio_query, hoy),
            estatus__in=[Cita.ESTATUS_SI_ASISTIO, Cita.ESTATUS_NO_ASISTIO],
            paciente__isnull=False,
            paciente__dado_de_alta=False, # agregar a un paciente
        ).values('paciente_id', 'estatus', 'fecha').order_by('paciente_id', '-fecha')
    )

    citas_todo = list(
        Cita.objects.filter(
            fecha__lte=hoy,
            estatus__in=[Cita.ESTATUS_SI_ASISTIO, Cita.ESTATUS_NO_ASISTIO],
            paciente__isnull=False,
            paciente__dado_de_alta=False, # agregar un paciente
        ).values('paciente_id', 'estatus', 'fecha').order_by('paciente_id', '-fecha')
    )

    def calc(subset):
        seen = {}
        for c in subset:
            pid = c['paciente_id']
            if pid not in excluir and pid not in seen:
                seen[pid] = c['estatus']
        si = sum(1 for s in seen.values() if s == Cita.ESTATUS_SI_ASISTIO)
        return {'total': len(seen), 'si': si, 'no': len(seen) - si}

    return {
        'dia': calc([c for c in citas if c['fecha'] == ayer]),
        'semana': calc([c for c in citas if c['fecha'] >= inicio_semana]),
        'mes': calc([c for c in citas if c['fecha'] >= inicio_mes]),
        'todo': calc(citas_todo),
    }


def aviso_privacidad(request):
    return render(request, 'clinica/privacidad.html')


def politicas_atencion(request):
    return render(request, 'clinica/politicas_atencion.html')


def terminos_condiciones(request):
    return render(request, 'clinica/terminos_condiciones.html')


@login_required
def home(request):
    # --- EL SEMAFORO INTELIGENTE ---
    if hasattr(request.user, 'perfil_terapeuta'):
        return redirect('portal_terapeuta')

    if hasattr(request.user, 'perfil_paciente'):
        return redirect('portal_paciente')

    if hasattr(request.user, 'perfil_empresa'):
        return redirect('portal_empresa')

    if hasattr(request.user, 'perfil_host'):
        return redirect('portal_host')

    if hasattr(request.user, 'perfil_consultoria'):
        return redirect('portal_consultoria')

    hoy = timezone.now().date()
    mes_actual = timezone.now().month

    solicitudes_pendientes = SolicitudCita.objects.filter(estado='pendiente').order_by('fecha_creacion')
    solicitudes_expositor_pendientes = SolicitudHorasExpositor.objects.filter(estado='pendiente').select_related('terapeuta').order_by('creado_en')

    # 1. ESTADÍSTICAS
    citas_hoy_count = Cita.objects.filter(fecha=hoy).count()

    pacientes_nuevos = Paciente.objects.filter(
        fecha_registro__month=mes_actual
    ).count()

    sin_reagendar = _sin_reagendar_stats()

    # 2. PRÓXIMAS CITAS
    # CORRECCIÓN 2: Usamos 'estatus' y ordenamos por 'hora' (no hora_inicio)
    proximas_citas = Cita.objects.filter(
        fecha__gte=hoy,
        estatus__in=Cita.ESTATUS_ACTIVOS,
    ).order_by('fecha', 'hora')

    dia_tablero = request.GET.get('dia', 'hoy')
    if dia_tablero == 'manana':
        fecha_tablero = hoy + timedelta(days=1)
    else:
        dia_tablero = 'hoy'
        fecha_tablero = hoy

    citas_tablero = Cita.objects.filter(
        fecha=fecha_tablero,
    ).select_related(
        'division', 'servicio', 'terapeuta', 'consultorio', 'paciente'
    ).prefetch_related(
        'pacientes_adicionales'
    ).order_by('fecha', 'hora')

    reagendos_pendientes = SolicitudReagendo.objects.filter(
        estado='pendiente',
    ).select_related('cita', 'cita__paciente', 'terapeuta').order_by('creado_en')

    bloqueos_vigentes = BloqueoAgendaTerapeuta.objects.filter(
        activo=True,
    ).filter(
        Q(tipo_bloqueo=BloqueoAgendaTerapeuta.TIPO_PERMANENTE) |
        Q(fecha_fin__gte=hoy)
    ).select_related('terapeuta').order_by('fecha_inicio', 'terapeuta__nombre')
    manual_portal = AccesoDirectoPortal.objects.filter(
        clave=AccesoDirectoPortal.CLAVE_MANUAL_PORTAL_MEDICO,
    ).first()

    manual_form = ManualPortalForm(initial={
        'titulo': manual_portal.titulo if manual_portal else 'Manual del sistema',
    })

    from collections import defaultdict
    from django.db.models import Prefetch
    notas_raw = NotaRecepcion.objects.select_related('creado_por').prefetch_related(
        Prefetch('reacciones', queryset=ReaccionNota.objects.select_related('usuario')),
        Prefetch('comentarios', queryset=ComentarioNota.objects.select_related('creado_por')),
    ).all()

    notas_recepcion = []
    for nota in notas_raw:
        grupos = defaultdict(list)
        for r in nota.reacciones.all():
            grupos[r.emoji].append(r.usuario_id)
        notas_recepcion.append({
            'obj': nota,
            'reacciones': [
                {'emoji': e, 'count': len(uids), 'yo': request.user.id in uids}
                for e, uids in grupos.items()
            ],
            'comentarios': list(nota.comentarios.all()),
        })
    total_notas = len(notas_recepcion)

    return render(request, 'clinica/home.html', {
        'citas_hoy': citas_hoy_count,
        'pacientes_nuevos': pacientes_nuevos,
        'sin_reagendar': sin_reagendar,
        'proximas_citas': proximas_citas,
        'citas_tablero': citas_tablero,
        'dia_tablero': dia_tablero,
        'fecha_tablero': fecha_tablero,
        'hoy': hoy,
        'form': CitaForm(),
        'solicitudes_pendientes': solicitudes_pendientes,
        'solicitudes_expositor_pendientes': solicitudes_expositor_pendientes,
        'bloqueos_vigentes': bloqueos_vigentes,
        'reagendos_pendientes': reagendos_pendientes,
        'manual_portal': manual_portal,
        'manual_form': manual_form,
        'notas_recepcion': notas_recepcion,
        'total_notas': total_notas,
    })


@login_required
def portal_host(request):
    if not hasattr(request.user, 'perfil_host'):
        return redirect('home')

    sedes_excluidas = {'zoom', 'externo'}
    sede_labels = dict(Consultorio.SEDE_CHOICES)
    sedes_disponibles = [
        {
            'value': sede,
            'label': sede_labels.get(sede, sede),
        }
        for sede in (
            Consultorio.objects
            .filter(activo=True)
            .exclude(sede__in=sedes_excluidas)
            .exclude(sede__isnull=True)
            .exclude(sede='')
            .values_list('sede', flat=True)
            .distinct()
            .order_by('sede')
        )
    ]
    sedes_validas = {sede['value'] for sede in sedes_disponibles}

    if request.method == 'POST':
        sede_seleccionada = request.POST.get('consultorio_sede', '').strip()
        if sede_seleccionada in sedes_validas:
            request.session['host_consultorio_sede'] = sede_seleccionada
            messages.success(request, f"Consultorio seleccionado: {sede_labels.get(sede_seleccionada, sede_seleccionada)}.")
        else:
            request.session.pop('host_consultorio_sede', None)
            messages.error(request, 'Selecciona un consultorio valido para continuar.')
        return redirect('portal_host')

    sede_actual = request.session.get('host_consultorio_sede')
    if sede_actual not in sedes_validas:
        sede_actual = None
        request.session.pop('host_consultorio_sede', None)

    hoy = timezone.localdate()
    ahora = timezone.localtime().time()
    citas = Cita.objects.none()
    proxima_cita = None

    if sede_actual:
        citas = (
            Cita.objects
            .filter(
                fecha=hoy,
                consultorio__sede=sede_actual,
                estatus__in=Cita.ESTATUS_ACTIVOS,
            )
            .select_related('paciente', 'terapeuta', 'servicio', 'consultorio')
            .prefetch_related('pacientes_adicionales')
            .order_by('hora')
        )
        proxima_cita = citas.filter(hora__gte=ahora).first() or citas.first()

    avatar_classes = ['av-teal', 'av-coral', 'av-slate']
    citas_host = []
    for index, cita in enumerate(citas):
        nombre = cita.pacientes_display_natural() or 'Paciente sin nombre'
        iniciales = ''.join([parte[0] for parte in nombre.split()[:2]]).upper() or 'IN'
        citas_host.append({
            'nombre': nombre,
            'tipo': cita.servicio.nombre if cita.servicio else cita.get_tipo_paciente_display(),
            'terapeuta': cita.terapeuta.nombre if cita.terapeuta else 'Sin terapeuta',
            'hora': cita.hora.strftime('%H:%M'),
            'sala': cita.consultorio.nombre if cita.consultorio else 'Sin consultorio',
            'estatus': cita.get_estatus_display(),
            'iniciales': iniciales,
            'avatar_class': avatar_classes[index % len(avatar_classes)],
        })

    tareas_checklist = (
        HostChecklistTask.objects
        .filter(activo=True)
        .filter(Q(hosts=request.user.perfil_host) | Q(hosts__isnull=True))
        .distinct()
        .order_by('orden', 'id')
    )

    return render(request, 'clinica/portal_host.html', {
        'host': request.user.perfil_host,
        'sedes_disponibles': sedes_disponibles,
        'sede_actual': sede_actual,
        'sede_actual_label': sede_labels.get(sede_actual, '') if sede_actual else '',
        'requiere_consultorio': not bool(sede_actual),
        'citas_host': citas_host,
        'citas_total': len(citas_host),
        'proxima_cita': proxima_cita,
        'tareas_checklist': tareas_checklist,
        'tareas_total': tareas_checklist.count(),
        'hoy': hoy,
    })


@login_required
def checklist_host_config(request):
    es_usuario_portal = (
        hasattr(request.user, 'perfil_terapeuta') or
        hasattr(request.user, 'perfil_paciente') or
        hasattr(request.user, 'perfil_empresa') or
        hasattr(request.user, 'perfil_host')
    )
    if es_usuario_portal and not (request.user.is_superuser or request.user.is_staff):
        return redirect('home')

    hosts = Host.objects.select_related('usuario').order_by('nombre')

    if request.method == 'POST':
        accion = request.POST.get('accion', 'crear')
        tarea_id = request.POST.get('tarea_id')

        if accion in ('crear', 'actualizar'):
            titulo = request.POST.get('titulo', '').strip()
            if not titulo:
                messages.error(request, 'El titulo de la tarea es obligatorio.')
                return redirect('checklist_host_config')

            if accion == 'actualizar':
                tarea = get_object_or_404(HostChecklistTask, id=tarea_id)
            else:
                tarea = HostChecklistTask()

            try:
                orden = int(request.POST.get('orden') or 0)
            except ValueError:
                orden = 0

            tarea.titulo = titulo
            tarea.subtitulo = request.POST.get('subtitulo', '').strip()
            tarea.etiqueta = request.POST.get('etiqueta', '').strip()
            tarea.urgente = request.POST.get('urgente') == 'on'
            tarea.activo = request.POST.get('activo') == 'on'
            tarea.orden = orden
            tarea.save()

            host_ids = request.POST.getlist('hosts')
            tarea.hosts.set(Host.objects.filter(id__in=host_ids))
            messages.success(request, 'Tarea de checklist guardada correctamente.')
            return redirect('checklist_host_config')

        if accion == 'toggle':
            tarea = get_object_or_404(HostChecklistTask, id=tarea_id)
            tarea.activo = not tarea.activo
            tarea.save(update_fields=['activo'])
            messages.success(request, 'Estado de la tarea actualizado.')
            return redirect('checklist_host_config')

        if accion == 'eliminar':
            tarea = get_object_or_404(HostChecklistTask, id=tarea_id)
            tarea.delete()
            messages.success(request, 'Tarea eliminada del checklist.')
            return redirect('checklist_host_config')

    tareas = list(
        HostChecklistTask.objects
        .prefetch_related('hosts')
        .order_by('orden', 'id')
    )
    tareas_config = [
        {
            'obj': tarea,
            'host_ids': list(tarea.hosts.values_list('id', flat=True)),
            'hosts_display': ', '.join(tarea.hosts.values_list('nombre', flat=True)) or 'Todos los hosts',
        }
        for tarea in tareas
    ]
    return render(request, 'clinica/checklist_host_config.html', {
        'tareas_config': tareas_config,
        'hosts': hosts,
    })


@api_key_required
def api_sin_reagendar(request):
    periodo = request.GET.get('periodo', 'mes')
    if periodo not in ('dia', 'semana', 'mes', 'todo'):
        periodo = 'mes'

    hoy = timezone.now().date()

    if periodo == 'dia':
        fecha_inicio = fecha_fin = hoy - timedelta(days=1)
    elif periodo == 'semana':
        fecha_inicio = hoy - timedelta(days=hoy.weekday())
        fecha_fin = hoy
    elif periodo == 'todo':
        fecha_inicio = None
        fecha_fin = hoy
    else:
        fecha_inicio = hoy.replace(day=1)
        fecha_fin = hoy

    excluir = set(
        Cita.objects.filter(
            fecha__gte=hoy,
            estatus__in=Cita.ESTATUS_ACTIVOS,
            paciente__isnull=False,
        ).values_list('paciente_id', flat=True)
    ) | set(
        SolicitudCita.objects.filter(
            estado='pendiente',
            paciente__isnull=False,
        ).values_list('paciente_id', flat=True)
    )

    citas_qs = Cita.objects.filter(
        estatus__in=[Cita.ESTATUS_SI_ASISTIO, Cita.ESTATUS_NO_ASISTIO],
        paciente__isnull=False,
        paciente__dado_de_alta=False, # dar de alta a un paciente
        paciente__estado='activo', # dar suspension a un paciente
    )
    if fecha_inicio is not None:
        citas_qs = citas_qs.filter(fecha__range=(fecha_inicio, fecha_fin))
    else:
        citas_qs = citas_qs.filter(fecha__lte=fecha_fin)
    citas = citas_qs.select_related('paciente', 'terapeuta').order_by('paciente_id', '-fecha')

    seen = set()
    result = []
    for cita in citas:
        pid = cita.paciente_id
        if pid not in excluir and pid not in seen:
            seen.add(pid)
            result.append({
                'paciente_id': cita.paciente.id,
                'nombre': cita.paciente.nombre,
                'estatus': cita.estatus,
                'fecha': cita.fecha.strftime('%d %b %Y'),
                'terapeuta': cita.terapeuta.nombre if cita.terapeuta else '—',
                'agendar_url': reverse('agendar_cita', args=[cita.paciente.id]),
            })

    si = sum(1 for r in result if r['estatus'] == Cita.ESTATUS_SI_ASISTIO)
    no = len(result) - si
    return JsonResponse({'pacientes': result, 'total': len(result), 'si': si, 'no': no})


@login_required
def actualizar_manual_portal(request):
    if not request.user.is_superuser:
        messages.error(request, 'Solo administracion puede actualizar el manual del sistema.')
        return redirect('home')

    if request.method != 'POST':
        return redirect('home')

    form = ManualPortalForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, 'Revisa el archivo del manual antes de guardarlo.')
        return redirect('home')

    archivo = form.cleaned_data['archivo']
    manual, _ = AccesoDirectoPortal.objects.get_or_create(
        clave=AccesoDirectoPortal.CLAVE_MANUAL_PORTAL_MEDICO,
        defaults={'titulo': 'Manual del sistema'},
    )
    manual.titulo = form.cleaned_data.get('titulo') or manual.titulo or 'Manual del sistema'
    manual.nombre_archivo = archivo.name
    manual.tipo_mime = getattr(archivo, 'content_type', '') or 'application/octet-stream'
    manual.contenido = archivo.read()
    manual.activo = True
    manual.save()

    messages.success(request, 'Manual del sistema actualizado correctamente.')
    return redirect('home')
# Asegúrate de tener esto arriba: from django.db.models import Q

@login_required
def lista_pacientes(request):
    query = request.GET.get('q')
    division_id = request.GET.get('division')

    if query:
        q_limpio = quitar_tildes(query)
        pacientes = Paciente.objects.filter(
            Q(nombre_normalizado__icontains=q_limpio) |
            Q(telefono__icontains=query)
        ).order_by('-fecha_registro')
    else:
        pacientes = Paciente.objects.all().order_by('-fecha_registro')

    if division_id:
        pacientes = pacientes.filter(division_id=division_id)

    # select_related evita una query extra por fila para servicio_inicial y
    # division (ambos FK), que la tabla de lista_pacientes.html muestra siempre.
    pacientes = pacientes.select_related('servicio_inicial', 'division')

    divisiones = Division.objects.all().order_by('nombre')
    return render(request, 'clinica/lista_pacientes.html', {
        'pacientes': pacientes,
        'divisiones': divisiones,
        'division_activa': division_id,
    })


@login_required
def asignar_division_paciente(request, paciente_id):
    paciente = get_object_or_404(Paciente, id=paciente_id)
    if request.method == 'POST':
        division_id = request.POST.get('division') or None
        paciente.division_id = division_id
        paciente.save()
        messages.success(request, 'División actualizada correctamente.')
    return redirect('detalle_paciente', paciente_id=paciente.id)
@login_required
def registrar_paciente(request):
    if request.method == 'POST':
        form = PacienteForm(request.POST, request.FILES)
        if form.is_valid():
            paciente = form.save()
            _registrar_actividad(request, 'paciente_registrado', 'paciente',
                f'Expediente de {paciente.nombre} registrado en el sistema.',
                paciente=paciente)
            return redirect('lista_pacientes')
    else:
        form = PacienteForm()
    return render(request, 'clinica/registro_paciente.html', {
        'form': form,
        'cancel_url': reverse('lista_pacientes'),
        'navbar_url': reverse('home'),
        'titulo_form': 'Nuevo Expediente',
        'subtitulo_form': 'Ingresa los datos del paciente',
    })
# En clinica/views.py

# En clinica/views.py (dentro de detalle_paciente)

def _generar_pdf_apertura(apertura):
    """Genera los bytes del PDF de apertura de expediente usando reportlab."""
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=2 * cm, leftMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    blue  = colors.HexColor('#003087')
    gray  = colors.HexColor('#546E7A')
    lblue = colors.HexColor('#90CAF9')

    T_HEADING = ParagraphStyle('heading', parent=styles['Normal'],
                               fontSize=12, fontName='Helvetica-Bold',
                               alignment=TA_CENTER, textColor=blue, spaceAfter=3)
    T_SUB     = ParagraphStyle('sub', parent=styles['Normal'],
                               fontSize=9, alignment=TA_CENTER, spaceAfter=8)
    T_SECTION = ParagraphStyle('section', parent=styles['Normal'],
                               fontSize=8, fontName='Helvetica-Bold',
                               textColor=blue, spaceBefore=8, spaceAfter=3)
    T_LABEL   = ParagraphStyle('label', parent=styles['Normal'],
                               fontSize=7, fontName='Helvetica-Bold', textColor=gray)
    T_VALUE   = ParagraphStyle('value', parent=styles['Normal'],
                               fontSize=9, fontName='Helvetica', spaceAfter=3)
    T_PRIVACY = ParagraphStyle('privacy', parent=styles['Normal'],
                               fontSize=6.5, textColor=gray, leading=9, spaceBefore=6)

    def lbl(text):
        return Paragraph(text.upper(), T_LABEL)

    def val(v):
        return Paragraph(str(v) if v else '—', T_VALUE)

    def grid(rows):
        """Crea una tabla de 4 columnas label/valor, label/valor."""
        tbl = Table(rows, colWidths=['18%', '32%', '18%', '32%'])
        tbl.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
        ]))
        return tbl

    p = apertura.paciente
    elems = []

    # ── Encabezado ──────────────────────────────────────────────────────────
    elems.append(Paragraph('INSTITUTO DE ATENCIÓN INTEGRAL Y DESARROLLO HUMANO A.C.', T_HEADING))
    elems.append(Paragraph('Información Personal del Paciente — Apertura de Expediente', T_SUB))
    elems.append(HRFlowable(width='100%', thickness=1.5, color=blue, spaceAfter=6))

    # Expediente No / fecha
    elems.append(grid([
        [lbl('Expediente No.'), val(apertura.expediente_no or '—'),
         lbl('Fecha de apertura'), val(apertura.creado_en.strftime('%d/%m/%Y'))],
    ]))
    elems.append(Spacer(1, 6))

    # ── Datos Personales ────────────────────────────────────────────────────
    elems.append(Paragraph('Datos Personales', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    elems.append(grid([
        [lbl('Nombre completo'), val(p.nombre),
         lbl('Fecha de Nacimiento'),
         val(p.fecha_nacimiento.strftime('%d/%m/%Y') if p.fecha_nacimiento else '—')],
        [lbl('Ocupación'), val(apertura.ocupacion),
         lbl('Estado Civil'), val(apertura.get_estado_civil_display() if apertura.estado_civil else '—')],
        [lbl('Lugar de Trabajo'), val(apertura.lugar_de_trabajo),
         lbl('Cargo'), val(apertura.cargo)],
        [lbl('Calle'), val(apertura.calle),
         lbl('Núm.'), val(apertura.num_exterior)],
        [lbl('Colonia'), val(apertura.colonia),
         lbl('División'), val(str(apertura.division) if apertura.division else '—')],
        [lbl('Número de Celular'), val(p.telefono or '—'),
         lbl('Religión'), val(apertura.religion or 'No especificada')],
    ]))

    # ── Convivencia e Hijos ─────────────────────────────────────────────────
    elems.append(Paragraph('Convivencia', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    elems.append(grid([
        [lbl('Vive con'), val(apertura.vive_con or '—'),
         lbl('Tiene hijos'), val('Sí' if apertura.tiene_hijos else 'No')],
        [lbl('No. de Hijos'), val(str(apertura.num_hijos) if apertura.num_hijos is not None else '—'),
         lbl(''), val('')],
    ]))
    if apertura.tiene_hijos:
        hijos = [
            (1, apertura.hijo_1), (2, apertura.hijo_2),
            (3, apertura.hijo_3), (4, apertura.hijo_4),
        ]
        h_rows = [[lbl(f'Hijo {i}'), val(h), lbl(''), val('')]
                  for i, h in hijos if h]
        if h_rows:
            elems.append(grid(h_rows))

    # ── Motivo de Consulta ──────────────────────────────────────────────────
    elems.append(Paragraph('Motivo de Consulta', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    elems.append(Paragraph(apertura.motivo_consulta or '—', T_VALUE))

    # ── Emergencia ──────────────────────────────────────────────────────────
    elems.append(Paragraph('Contacto de Emergencia', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    emg = Table(
        [[lbl('En caso de emergencia llamar a'), val(apertura.emergencia_contacto),
          lbl('Teléfono'), val(apertura.emergencia_telefono)]],
        colWidths=['22%', '28%', '14%', '36%'],
    )
    emg.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    elems.append(emg)

    # ── Cómo se enteró ──────────────────────────────────────────────────────
    elems.append(Spacer(1, 6))
    elems.append(Table(
        [[lbl('¿Cómo se enteró de nosotros?'), val(apertura.como_se_entero or '—')]],
        colWidths=['28%', '72%'],
    ))

    # ── Antecedentes Médicos ────────────────────────────────────────────────
    elems.append(Paragraph('Antecedentes Médicos', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    elems.append(grid([
        [lbl('¿Tiene alguna enfermedad?'), val('Sí' if apertura.tiene_enfermedad else 'No'),
         lbl('¿Cuál?'), val(apertura.cual_enfermedad or '—')],
    ]))

    # ── Tratamiento Psiquiátrico ────────────────────────────────────────────
    elems.append(Paragraph('Tratamiento Psiquiátrico', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    elems.append(grid([
        [lbl('¿Está o ha estado en Tx psiquiátrico?'), val('Sí' if apertura.tx_psiquiatrico else 'No'),
         lbl('¿Hace cuánto?'), val(apertura.tx_psiquiatrico_hace_cuanto or '—')],
        [lbl('Medicamento(s)'), val(apertura.tx_psiquiatrico_medicamento or '—'),
         lbl(''), val('')],
    ]))
    if apertura.tx_psiquiatrico and apertura.tx_psiquiatrico_motivo:
        elems.append(Table(
            [[lbl('Motivo del tratamiento'), val(apertura.tx_psiquiatrico_motivo)]],
            colWidths=['22%', '78%'],
        ))

    # ── Terapia Previa ──────────────────────────────────────────────────────
    elems.append(Paragraph('Terapia Previa', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    elems.append(grid([
        [lbl('¿Ha tomado terapia anteriormente?'), val('Sí' if apertura.ha_tomado_terapia else 'No'),
         lbl('¿Hace cuánto?'), val(apertura.terapia_hace_cuanto or '—')],
        [lbl('¿Cuánto duró?'), val(apertura.terapia_duracion or '—'),
         lbl(''), val('')],
    ]))
    if apertura.ha_tomado_terapia and apertura.terapia_motivo:
        elems.append(Table(
            [[lbl('Motivo'), val(apertura.terapia_motivo)]],
            colWidths=['22%', '78%'],
        ))

    # ── Sustancias ──────────────────────────────────────────────────────────
    elems.append(Paragraph('Sustancias', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    elems.append(grid([
        [lbl('Fuma'), val('Sí' if apertura.fuma else 'No'),
         lbl('Consume alcohol'), val('Sí' if apertura.consume_alcohol else 'No')],
        [lbl('Otras sustancias'), val('Sí' if apertura.consume_otras_sustancias else 'No'),
         lbl('¿Cuáles?'), val(apertura.cuales_sustancias or '—')],
    ]))

    # ── Hábitos ─────────────────────────────────────────────────────────────
    elems.append(Paragraph('Hábitos', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    elems.append(grid([
        [lbl('Comidas al día'), val(str(apertura.comidas_al_dia) if apertura.comidas_al_dia is not None else '—'),
         lbl('Horas de sueño al día'), val(str(apertura.horas_sueno) if apertura.horas_sueno is not None else '—')],
        [lbl('Actividad física'), val('Sí' if apertura.actividad_fisica else 'No'),
         lbl('¿Cuál?'), val(apertura.cual_actividad_fisica or '—')],
    ]))

    # ── Riesgo Suicida ──────────────────────────────────────────────────────
    elems.append(Paragraph('Riesgo Suicida', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    elems.append(grid([
        [lbl('¿Ha intentado quitarse la vida?'), val('Sí' if apertura.intento_suicida else 'No'),
         lbl('¿Hace cuánto?'), val(apertura.intento_suicida_hace_cuanto or '—')],
    ]))
    if apertura.intento_suicida:
        rows_suicida = []
        if apertura.intento_suicida_que_hizo:
            rows_suicida.append([lbl('¿Qué hizo?'), val(apertura.intento_suicida_que_hizo)])
        if apertura.intento_suicida_motivo:
            rows_suicida.append([lbl('¿Por qué?'), val(apertura.intento_suicida_motivo)])
        for r in rows_suicida:
            elems.append(Table([r], colWidths=['22%', '78%']))

    # ── Vida Sexual ─────────────────────────────────────────────────────────
    elems.append(Paragraph('Vida Sexual', T_SECTION))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=lblue, spaceAfter=3))
    elems.append(Table(
        [[lbl('¿Tiene vida sexual activa?'), val('Sí' if apertura.vida_sexual_activa else 'No'), lbl(''), val('')]],
        colWidths=['28%', '22%', '18%', '32%'],
    ))

    # ── Aviso de Privacidad ─────────────────────────────────────────────────
    elems.append(Spacer(1, 10))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=gray))
    elems.append(Paragraph(
        '<b>Aviso de Privacidad</b> — De acuerdo con lo establecido en la Constitución Política de los '
        'Estados Unidos Mexicanos, la Ley Federal De Protección de Datos Personales y en el Código Ético '
        'del Psicólogo, la totalidad de la información como de los registros e historias clínicas, están '
        'cubiertas por el secreto profesional del Instituto de Atención Integral y Desarrollo Humano A.C. '
        'y sus colaboradores. WhatsApp: 844 443 99 87 | ANA CECILIA TREVIÑO No. 158 COL. IGNACIO ZARAGOZA',
        T_PRIVACY,
    ))

    doc.build(elems)
    return buffer.getvalue()


def detalle_paciente(request, paciente_id):
    paciente = get_object_or_404(Paciente, id=paciente_id)

    can_subir_documentos = request.user.is_authenticated and request.user.is_superuser
    if request.method == 'POST':
        if not can_subir_documentos:
            messages.error(request, 'Solo un superusuario puede subir documentos al expediente.')
            return redirect('detalle_paciente', paciente_id=paciente.id)

        form_documento = DocumentoPacienteForm(request.POST, request.FILES)
        if form_documento.is_valid():
            archivo = request.FILES['archivo']
            documento = form_documento.save(commit=False)
            documento.paciente = paciente
            documento.subido_por = request.user
            documento.nombre_archivo = archivo.name
            documento.tipo_mime = archivo.content_type
            documento.contenido = archivo.read()
            documento.save()
            messages.success(request, 'Documento agregado correctamente.')
            return redirect('detalle_paciente', paciente_id=paciente.id)
        messages.error(request, 'Revisa el archivo y los datos del documento.')
    else:
        form_documento = DocumentoPacienteForm()

    historial = Cita.objects.filter(
        Q(paciente=paciente) | Q(pacientes_adicionales=paciente)
    ).distinct().order_by('-fecha', '-hora')
    
    # Extraemos los terapeutas unicos que han atendido a este paciente
    terapeutas_previos = set(cita.terapeuta for cita in historial if cita.terapeuta)
    notas_historial = NotaTerapeutaPaciente.objects.filter(
        paciente=paciente
    ).select_related('terapeuta').order_by('-creado_en')
    documentos_historial = DocumentoPaciente.objects.filter(
        paciente=paciente
    ).select_related('terapeuta', 'subido_por').order_by('-creado_en')
    reportes_historial = ReporteSesion.objects.filter(
        paciente=paciente
    ).select_related('terapeuta').order_by('-fecha', '-creado_en')

    apertura = getattr(paciente, 'apertura_expediente_obj', None)

    context = {
        'paciente': paciente,
        'historial': historial,
        'terapeutas_previos': terapeutas_previos,
        'notas_historial': notas_historial,
        'documentos_historial': documentos_historial,
        'reportes_historial': reportes_historial,
        'apertura': apertura,
        'form_documento': form_documento,
        'can_subir_documentos': can_subir_documentos,
        'divisiones': Division.objects.all().order_by('nombre'),
    }
    return render(request, 'clinica/detalle_paciente.html', context)


def _pacientes_ids_terapeuta(terapeuta):
    ids = set(
        Cita.objects.filter(
            terapeuta=terapeuta,
            paciente__isnull=False,
        ).values_list('paciente_id', flat=True)
    )
    ids.update(
        Cita.objects.filter(
            terapeuta=terapeuta,
            pacientes_adicionales__isnull=False,
        ).values_list('pacientes_adicionales__id', flat=True)
    )
    ids.update(
        PacienteTerapeutaAcceso.objects.filter(
            terapeuta=terapeuta,
        ).values_list('paciente_id', flat=True)
    )
    return {pid for pid in ids if pid}


@login_required
def registrar_paciente_terapeuta(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    terapeuta = request.user.perfil_terapeuta
    if request.method == 'POST':
        form = PacienteForm(request.POST, request.FILES)
        if form.is_valid():
            paciente = form.save()
            PacienteTerapeutaAcceso.objects.get_or_create(
                terapeuta=terapeuta,
                paciente=paciente,
                defaults={'creado_por': request.user},
            )
            messages.success(request, 'Paciente agregado a tus expedientes correctamente.')
            return redirect('expediente_terapeuta_detalle', paciente_id=paciente.id)
    else:
        form = PacienteForm()

    return render(request, 'clinica/registro_paciente.html', {
        'form': form,
        'cancel_url': reverse('expedientes_terapeuta'),
        'navbar_url': reverse('portal_terapeuta'),
        'titulo_form': 'Nuevo expediente desde portal medico',
        'subtitulo_form': 'Crea el paciente y quedara vinculado a tus expedientes.',
    })


@login_required
def expedientes_terapeuta(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    terapeuta = request.user.perfil_terapeuta
    paciente_ids = _pacientes_ids_terapeuta(terapeuta)
    pacientes = Paciente.objects.filter(id__in=paciente_ids).prefetch_related('expedientes_grupales').order_by('nombre')

    return render(request, 'clinica/expedientes_terapeuta.html', {
        'terapeuta': terapeuta,
        'pacientes': pacientes,
    })


@login_required
def expediente_terapeuta_detalle(request, paciente_id):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    terapeuta = request.user.perfil_terapeuta
    paciente_ids = _pacientes_ids_terapeuta(terapeuta)
    if paciente_id not in paciente_ids:
        messages.error(request, 'Solo puedes abrir expedientes de pacientes agendados contigo.')
        return redirect('expedientes_terapeuta')

    paciente = get_object_or_404(Paciente.objects.prefetch_related('expedientes_grupales'), id=paciente_id)
    historial = Cita.objects.filter(
        terapeuta=terapeuta
    ).filter(
        Q(paciente=paciente) | Q(pacientes_adicionales=paciente)
    ).distinct().order_by('-fecha', '-hora')
    notas_historial = NotaTerapeutaPaciente.objects.filter(
        terapeuta=terapeuta,
        paciente=paciente,
    ).order_by('-creado_en')
    documentos_historial = DocumentoPaciente.objects.filter(
        paciente=paciente
    ).select_related('terapeuta', 'subido_por').order_by('-creado_en')
    reportes_historial = ReporteSesion.objects.filter(
        terapeuta=terapeuta,
        paciente=paciente,
    ).select_related('cita__expediente_grupal').order_by('-fecha', '-creado_en')
    for reporte in reportes_historial:
        reporte.form_edicion = ReporteSesionForm(instance=reporte, prefix=f'reporte_{reporte.id}')

    apertura = getattr(paciente, 'apertura_expediente_obj', None)

    # --- Auto-calc: número de sesión = reportes ya guardados + 1 ---
    numero_sesion_sugerido = ReporteSesion.objects.filter(
        terapeuta=terapeuta,
        paciente=paciente,
    ).count() + 1

    # --- Auto-calc: hora de la cita más reciente de hoy como hora_inicio ---
    hoy = date.today()
    cita_hoy = historial.filter(fecha=hoy).order_by('hora').last()
    hora_inicio_sugerida = cita_hoy.hora if cita_hoy else None
    cita_sugerida = cita_hoy

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'guardar_apertura':
            form_apertura = AperturaExpedienteForm(request.POST, instance=apertura)
            form_reporte = ReporteSesionForm()
            form = NotaTerapeutaPacienteForm()
            form_documento = DocumentoPacienteForm()
            if form_apertura.is_valid():
                ap = form_apertura.save(commit=False)
                ap.paciente = paciente
                ap.save()

                nombre_completo = (form_apertura.cleaned_data.get('nombre') or '').strip()

                # Los apellidos ya no se capturan por separado en la UI.
                # Conservamos el modelo satisfecho con un valor de respaldo,
                # pero la fuente real del nombre visible/PDF es paciente.nombre.
                ap.apellido_paterno = ap.apellido_paterno or nombre_completo or '-'
                ap.apellido_materno = ap.apellido_materno or ''
                ap.save(update_fields=['apellido_paterno', 'apellido_materno'])

                paciente.nombre = nombre_completo or paciente.nombre
                paciente.fecha_nacimiento = form_apertura.cleaned_data.get('fecha_nacimiento') or paciente.fecha_nacimiento
                paciente.telefono = form_apertura.cleaned_data.get('telefono') or paciente.telefono
                paciente.save(update_fields=['nombre', 'fecha_nacimiento', 'telefono'])

                # Generar PDF y enlazar como DocumentoPaciente
                pdf_bytes = _generar_pdf_apertura(ap)
                nombre_pdf = f'apertura_expediente_{paciente.id}.pdf'
                if ap.documento:
                    doc_obj = ap.documento
                    doc_obj.contenido = pdf_bytes
                    doc_obj.nombre_archivo = nombre_pdf
                    doc_obj.save(update_fields=['contenido', 'nombre_archivo'])
                else:
                    doc_obj = DocumentoPaciente.objects.create(
                        paciente=paciente,
                        terapeuta=terapeuta,
                        subido_por=request.user,
                        tipo_documento='apertura',
                        nombre_archivo=nombre_pdf,
                        tipo_mime='application/pdf',
                        contenido=pdf_bytes,
                        descripcion='Apertura de expediente generada automáticamente.',
                    )
                    ap.documento = doc_obj
                    ap.save(update_fields=['documento'])
                messages.success(request, 'Apertura de expediente guardada y PDF generado correctamente.')
                return redirect('expediente_terapeuta_detalle', paciente_id=paciente.id)
            else:
                messages.error(request, 'Revisa los campos de la apertura.')

        elif accion == 'guardar_reporte':
            form_apertura = AperturaExpedienteForm(instance=apertura)
            form_reporte = ReporteSesionForm(request.POST)
            form = NotaTerapeutaPacienteForm()
            form_documento = DocumentoPacienteForm()
            if form_reporte.is_valid():
                reporte = form_reporte.save(commit=False)
                reporte.paciente = paciente
                reporte.terapeuta = terapeuta
                reporte.fecha = form_reporte.cleaned_data.get('fecha') or hoy
                reporte.numero_sesion = numero_sesion_sugerido
                reporte.hora_inicio = form_reporte.cleaned_data.get('hora_inicio') or hora_inicio_sugerida
                cita_relacionada = historial.filter(fecha=reporte.fecha).order_by('hora').last()
                if reporte.hora_inicio:
                    cita_relacionada = historial.filter(
                        fecha=reporte.fecha,
                        hora=reporte.hora_inicio,
                    ).first() or cita_relacionada
                if cita_relacionada:
                    reporte.cita = cita_relacionada
                reporte.save()
                messages.success(request, f'Reporte de sesión #{reporte.numero_sesion} guardado correctamente.')
                return redirect('expediente_terapeuta_detalle', paciente_id=paciente.id)
            else:
                messages.error(request, 'Revisa los campos del reporte.')

        elif accion == 'editar_reporte':
            reporte_id = request.POST.get('reporte_id')
            reporte = get_object_or_404(
                ReporteSesion,
                id=reporte_id,
                terapeuta=terapeuta,
                paciente=paciente,
            )
            form_apertura = AperturaExpedienteForm(instance=apertura)
            form_reporte = ReporteSesionForm(initial={
                'fecha': hoy,
                'hora_inicio': hora_inicio_sugerida,
            })
            form = NotaTerapeutaPacienteForm()
            form_documento = DocumentoPacienteForm()
            form_editar = ReporteSesionForm(
                request.POST,
                instance=reporte,
                prefix=f'reporte_{reporte.id}',
            )
            if form_editar.is_valid():
                reporte_editado = form_editar.save(commit=False)
                reporte_editado.paciente = paciente
                reporte_editado.terapeuta = terapeuta
                cita_relacionada = historial.filter(fecha=reporte_editado.fecha).order_by('hora').last()
                if reporte_editado.hora_inicio:
                    cita_relacionada = historial.filter(
                        fecha=reporte_editado.fecha,
                        hora=reporte_editado.hora_inicio,
                    ).first() or cita_relacionada
                reporte_editado.cita = cita_relacionada
                reporte_editado.save()
                messages.success(request, f'Reporte de sesión #{reporte_editado.numero_sesion} actualizado correctamente.')
                return redirect('expediente_terapeuta_detalle', paciente_id=paciente.id)

            reporte.form_edicion = form_editar
            messages.error(request, f'Revisa los campos del reporte #{reporte.numero_sesion}.')

        elif accion == 'subir_documento':
            form_apertura = AperturaExpedienteForm(instance=apertura)
            form_reporte = ReporteSesionForm()
            form_documento = DocumentoPacienteForm(request.POST, request.FILES)
            form = NotaTerapeutaPacienteForm()
            if form_documento.is_valid():
                archivo = request.FILES['archivo']
                documento = form_documento.save(commit=False)
                documento.paciente = paciente
                documento.terapeuta = terapeuta
                documento.subido_por = request.user
                documento.nombre_archivo = archivo.name
                documento.tipo_mime = archivo.content_type
                documento.contenido = archivo.read()
                documento.save()
                messages.success(request, 'Documento agregado correctamente.')
                return redirect('expediente_terapeuta_detalle', paciente_id=paciente.id)

        else:
            form_apertura = AperturaExpedienteForm(instance=apertura)
            form_reporte = ReporteSesionForm()
            form = NotaTerapeutaPacienteForm(request.POST)
            form_documento = DocumentoPacienteForm()
            if form.is_valid():
                nota = form.save(commit=False)
                nota.terapeuta = terapeuta
                nota.paciente = paciente
                nota.save()
                messages.success(request, 'Nota agregada correctamente.')
                return redirect('expediente_terapeuta_detalle', paciente_id=paciente.id)
    else:
        form = NotaTerapeutaPacienteForm()
        form_documento = DocumentoPacienteForm()
        form_reporte = ReporteSesionForm(initial={
            'fecha': hoy,
            'hora_inicio': hora_inicio_sugerida,
        })
        form_apertura = AperturaExpedienteForm(instance=apertura)

    historial_completo = (
        Cita.objects
        .filter(Q(paciente=paciente) | Q(pacientes_adicionales=paciente))
        .distinct()
        .select_related('terapeuta', 'servicio', 'consultorio')
        .order_by('-fecha', '-hora')
    )

    instrumentos_disponibles = Instrumento.objects.filter(activo=True).order_by('nombre')
    envios_instrumento = EnvioInstrumento.objects.filter(
        paciente=paciente
    ).select_related('instrumento').order_by('-creado_en')
    for envio in envios_instrumento:
        envio.link_publico = request.build_absolute_uri(reverse('responder_instrumento', args=[envio.token]))

    return render(request, 'clinica/expediente_terapeuta_detalle.html', {
        'terapeuta': terapeuta,
        'paciente': paciente,
        'historial': historial,
        'historial_completo': historial_completo,
        'form_notas': form,
        'notas_historial': notas_historial,
        'form_documento': form_documento,
        'documentos_historial': documentos_historial,
        'form_reporte': form_reporte,
        'reportes_historial': reportes_historial,
        'numero_sesion_sugerido': numero_sesion_sugerido,
        'hora_inicio_sugerida': hora_inicio_sugerida,
        'fecha_hoy': hoy,
        'apertura': apertura,
        'form_apertura': form_apertura,
        'instrumentos_disponibles': instrumentos_disponibles,
        'envios_instrumento': envios_instrumento,
    })

@login_required
def catalogo_instrumentos(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    q = request.GET.get('q', '').strip()
    instrumentos = Instrumento.objects.filter(activo=True).annotate(
        total_preguntas=Count('preguntas')
    ).order_by('nombre')
    if q:
        instrumentos = instrumentos.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))

    return render(request, 'clinica/catalogo_instrumentos.html', {
        'instrumentos': instrumentos,
        'q': q,
        'total_catalogo': Instrumento.objects.filter(activo=True).count(),
    })


@login_required
def generar_envio_instrumento(request, paciente_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)
    if not hasattr(request.user, 'perfil_terapeuta'):
        return JsonResponse({'ok': False, 'error': 'No autorizado.'}, status=403)

    terapeuta = request.user.perfil_terapeuta
    if paciente_id not in _pacientes_ids_terapeuta(terapeuta):
        return JsonResponse({'ok': False, 'error': 'Solo puedes generar enlaces para pacientes agendados contigo.'}, status=403)

    paciente = get_object_or_404(Paciente, id=paciente_id)
    instrumento = get_object_or_404(Instrumento, id=request.POST.get('instrumento_id'), activo=True)

    envio = EnvioInstrumento.objects.create(
        instrumento=instrumento,
        paciente=paciente,
        generado_por=request.user,
    )
    link = request.build_absolute_uri(reverse('responder_instrumento', args=[envio.token]))

    _registrar_actividad(
        request,
        accion=RegistroActividad.ACCION_INSTRUMENTO_ENVIADO,
        categoria=RegistroActividad.CAT_INSTRUMENTO,
        descripcion=f'{terapeuta.nombre} generó un enlace de "{instrumento.nombre}" para {paciente.nombre}.',
        terapeuta=terapeuta,
        paciente=paciente,
    )

    return JsonResponse({
        'ok': True,
        'link': link,
        'envio': {
            'id': envio.id,
            'instrumento': instrumento.nombre,
            'estado': envio.estado,
            'estado_display': envio.get_estado_display(),
            'creado_en': envio.creado_en.strftime('%d/%m/%Y %H:%M'),
        },
    })


@login_required
def resultado_envio_instrumento(request, paciente_id, envio_id):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    terapeuta = request.user.perfil_terapeuta
    if paciente_id not in _pacientes_ids_terapeuta(terapeuta):
        messages.error(request, 'Solo puedes ver instrumentos de pacientes agendados contigo.')
        return redirect('expedientes_terapeuta')

    paciente = get_object_or_404(Paciente, id=paciente_id)
    envio = get_object_or_404(
        EnvioInstrumento.objects.select_related('instrumento', 'paciente'),
        id=envio_id, paciente=paciente,
    )

    if envio.estado != EnvioInstrumento.ESTADO_RESPONDIDO:
        messages.info(request, 'Este instrumento todavía no ha sido respondido por el paciente.')
        return redirect('expediente_terapeuta_detalle', paciente_id=paciente.id)

    respuestas = list(envio.respuestas.select_related('pregunta').order_by('pregunta__orden', 'pregunta__id'))

    detalle_items = []
    if isinstance(envio.resultado_detalle, dict):
        # Para Raven, las claves de discrepancia se muestran en los headers de serie;
        # se excluyen de la grid genérica para evitar duplicación.
        _excluir_raven = {'Patrón de respuesta'} if envio.instrumento.clave == 'raven' else set()
        detalle_items = [
            (k, v) for k, v in envio.resultado_detalle.items()
            if not k.startswith('Disc. Serie') and k not in _excluir_raven
        ]

    evolucion = list(EnvioInstrumento.objects.filter(
        paciente=paciente,
        instrumento=envio.instrumento,
        estado=EnvioInstrumento.ESTADO_RESPONDIDO,
    ).order_by('respondido_en'))

    # Agrupamiento por instrumento para display enriquecido
    clave = envio.instrumento.clave
    raven_series = None
    allport_sec1 = None
    allport_sec2 = None
    scl90_data = None

    if clave == 'raven':
        resp_by_orden = {r.pregunta.orden: r for r in respuestas}
        raven_series = []
        for nombre, start, end in _RAVEN_SERIES:
            items = []
            for orden in range(start, end + 1):
                r = resp_by_orden.get(orden)
                if r:
                    es_correcta = bool(r.pregunta.opciones) and any(
                        opt.get('valor') == r.valor and opt.get('correcta')
                        for opt in r.pregunta.opciones
                    )
                    items.append({'resp': r, 'correcta': es_correcta})
            raven_series.append({
                'nombre': nombre,
                'items': items,
                'correctas': sum(1 for x in items if x['correcta']),
            })

        # Análisis de discrepancia: expected vs actual por serie
        total_correctas = sum(s['correctas'] for s in raven_series)
        esperados_tuple = _RAVEN_ESPERADOS.get(total_correctas)
        for i, serie in enumerate(raven_series):
            esp = esperados_tuple[i] if esperados_tuple else None
            diff = (serie['correctas'] - esp) if esp is not None else None
            serie['esperado'] = esp
            serie['diff'] = diff
            if diff is None:
                serie['disc_nivel'] = 'sin_dato'
            elif abs(diff) >= 3:
                serie['disc_nivel'] = 'significativa'
            elif abs(diff) == 2:
                serie['disc_nivel'] = 'notable'
            else:
                serie['disc_nivel'] = 'normal'

    elif clave == 'scl90':
        det = envio.resultado_detalle or {}
        subescalas = det.get('_scl90_subescalas')
        summary    = det.get('_scl90_summary')
        if subescalas and summary:
            import json
            scl90_data = {
                'subescalas': subescalas,
                'summary': summary,
                'ref': _SCL90_REF,
                'ref_summary': _SCL90_REF_SUMMARY,
                # Series JSON para Chart.js
                'chart_json': json.dumps({
                    'labels': [s['abrev'] for s in subescalas],
                    'paciente':  [s['media'] for s in subescalas],
                    'normal':    _SCL90_REF['normal'],
                    'corte':     [round(s['media_norm'] + s['ds_norm'], 2) for s in subescalas],
                    'externos':  _SCL90_REF['externos'],
                }),
            }

    elif clave == 'allport':
        allport_sec1 = [r for r in respuestas if r.pregunta.orden <= 30]
        sec2_items = [r for r in respuestas if r.pregunta.orden >= 31]
        # Agrupar sección 2 en grupos de 4 sub-ítems (a, b, c, d) por pregunta
        allport_sec2 = []
        for i in range(0, len(sec2_items), 4):
            grupo = sec2_items[i:i + 4]
            if grupo:
                allport_sec2.append({'num': i // 4 + 1, 'items': grupo})

    raven_patron = None
    if clave == 'raven' and isinstance(envio.resultado_detalle, dict):
        raven_patron = envio.resultado_detalle.get('Patrón de respuesta')

    # Para SCL-90: excluir claves privadas (_) del grid genérico
    if clave == 'scl90':
        detalle_items = [(k, v) for k, v in detalle_items if not k.startswith('_')]

    return render(request, 'clinica/resultado_instrumento.html', {
        'terapeuta': terapeuta,
        'paciente': paciente,
        'envio': envio,
        'respuestas': respuestas,
        'detalle_items': detalle_items,
        'evolucion': evolucion,
        'raven_series': raven_series,
        'raven_patron': raven_patron,
        'allport_sec1': allport_sec1,
        'allport_sec2': allport_sec2,
        'scl90_data':   scl90_data,
    })


@login_required
def registrar_resultado_raven(request, paciente_id):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    terapeuta = request.user.perfil_terapeuta
    if paciente_id not in _pacientes_ids_terapeuta(terapeuta):
        messages.error(request, 'Solo puedes registrar resultados de pacientes agendados contigo.')
        return redirect('expedientes_terapeuta')

    paciente = get_object_or_404(Paciente, id=paciente_id)
    instrumento = get_object_or_404(Instrumento, clave='raven', activo=True)

    if request.method == 'POST':
        raw_str = request.POST.get('puntaje_raw', '').strip()
        try:
            puntaje_raw = int(raw_str)
            if not (0 <= puntaje_raw <= 60):
                raise ValueError
        except ValueError:
            messages.error(request, 'El puntaje debe ser un número entero entre 0 y 60.')
            return render(request, 'clinica/registrar_resultado_raven.html', {
                'terapeuta': terapeuta, 'paciente': paciente, 'puntaje_raw': raw_str,
            })

        resultado = calcular_raven(puntaje_raw)
        now = timezone.now()
        envio = EnvioInstrumento.objects.create(
            instrumento=instrumento,
            paciente=paciente,
            generado_por=request.user,
            estado=EnvioInstrumento.ESTADO_RESPONDIDO,
            respondido_en=now,
            puntaje_total=resultado['puntaje_total'],
            interpretacion=resultado['interpretacion'],
            resultado_detalle=resultado['detalle'],
        )

        _registrar_actividad(
            request,
            accion=RegistroActividad.ACCION_INSTRUMENTO_ENVIADO,
            categoria=RegistroActividad.CAT_INSTRUMENTO,
            descripcion=(
                f'{terapeuta.nombre} registró resultado del Raven SPM '
                f'(puntaje {puntaje_raw}) para {paciente.nombre}.'
            ),
            terapeuta=terapeuta,
            paciente=paciente,
        )

        return redirect('resultado_envio_instrumento', paciente_id=paciente.id, envio_id=envio.id)

    return render(request, 'clinica/registrar_resultado_raven.html', {
        'terapeuta': terapeuta,
        'paciente': paciente,
    })


@login_required
def descargar_documento(request, doc_id):
    documento = get_object_or_404(DocumentoPaciente, id=doc_id)
    response = HttpResponse(bytes(documento.contenido), content_type=documento.tipo_mime or 'application/octet-stream')
    response['Content-Disposition'] = f'inline; filename="{documento.nombre_archivo}"'
    return response

@login_required
def agendar_cita(request, paciente_id):
    paciente = get_object_or_404(Paciente, id=paciente_id)
    
    if request.method == 'POST':
        form = CitaForm(request.POST)
        if form.is_valid():
            cita = form.save(commit=False)
            cita.paciente = paciente  # Aquí vinculamos la cita al paciente automáticamente
            cita.save()
            if es_servicio_grupal(cita.servicio):
                pacientes_extra = form.cleaned_data.get('pacientes_extra')
                cita.pacientes_adicionales.set(
                    (pacientes_extra or Paciente.objects.none()).exclude(pk=cita.paciente_id)
                )
            else:
                cita.pacientes_adicionales.clear()
            return redirect('detalle_paciente', paciente_id=paciente.id)
    else:
        # Pre-llenamos el terapeuta por defecto si quieres, o lo dejamos vacío
        form = CitaForm(initial={'costo': 500})
    
    return render(request, 'clinica/agendar_cita.html', {'form': form, 'paciente': paciente})
@login_required
def editar_paciente(request, paciente_id):
    paciente = get_object_or_404(Paciente, id=paciente_id)
    
    if request.method == 'POST':
        # FILES es necesario para recibir archivos (PDFs, fotos)
        form = PacienteForm(request.POST, request.FILES, instance=paciente)
        if form.is_valid():
            form.save()
            _registrar_actividad(request, 'paciente_editado', 'paciente',
                f'Expediente de {paciente.nombre} modificado.',
                paciente=paciente)
            return redirect('detalle_paciente', paciente_id=paciente.id)
    else:
        form = PacienteForm(instance=paciente)
    
    return render(request, 'clinica/editar_paciente.html', {'form': form, 'paciente': paciente})
@login_required
def agendar_cita(request, paciente_id):
    from clinica.models import PenalizacionPaciente
    paciente = get_object_or_404(Paciente, id=paciente_id)

    # Penalización pendiente del paciente (la más reciente no pagada)
    penalizacion_pendiente = (
        PenalizacionPaciente.objects
        .filter(paciente=paciente, pagada=False)
        .select_related('cita_origen__servicio')
        .order_by('-fecha_creacion')
        .first()
    )

    if request.method == 'POST':
        from .forms import verificar_empalme_paciente
        form = CitaForm(request.POST)
        if form.is_valid():
            fecha = form.cleaned_data.get('fecha')
            hora  = form.cleaned_data.get('hora')
            conflicto = verificar_empalme_paciente(paciente, fecha, hora) if fecha and hora else None
            if conflicto:
                terapeuta_str = conflicto.terapeuta.nombre if conflicto.terapeuta else 'sin terapeuta'
                messages.error(
                    request,
                    f"{paciente.nombre} ya tiene una cita el {fecha:%d/%m/%Y} a las {hora:%H:%M} "
                    f"con {terapeuta_str} ({conflicto.get_estatus_display()}). No se puede empalmar."
                )
            else:
                cita = form.save(commit=False)
                cita.paciente = paciente

                if penalizacion_pendiente:
                    costo_actual = cita.costo or 0
                    cita.costo = costo_actual + penalizacion_pendiente.monto

                cita.save()
                _registrar_actividad(request, 'cita_creada', 'cita',
                    f'Cita de {paciente.nombre} con '
                    f'{cita.terapeuta.nombre if cita.terapeuta else "Sin terapeuta"} '
                    f'el {cita.fecha.strftime("%d/%m/%Y")} a las {cita.hora.strftime("%H:%M")} agendada.',
                    terapeuta=cita.terapeuta, paciente=paciente)

                if penalizacion_pendiente:
                    penalizacion_pendiente.pagada = True
                    penalizacion_pendiente.cita_cobro = cita
                    penalizacion_pendiente.save(update_fields=['pagada', 'cita_cobro'])

                if es_servicio_grupal(cita.servicio):
                    pacientes_extra = form.cleaned_data.get('pacientes_extra')
                    cita.pacientes_adicionales.set(
                        (pacientes_extra or Paciente.objects.none()).exclude(pk=cita.paciente_id)
                    )
                else:
                    cita.pacientes_adicionales.clear()
                _vincular_expediente_grupal(cita)

                try:
                    sincronizar_google_sheet(cita)
                    print("✅ Sincronización exitosa")
                except Exception as e:
                    print(f" Error al sincronizar con Google: {e}")

                return redirect('detalle_paciente', paciente_id=paciente.id)
    else:
        form = CitaForm(initial={'costo': 500})

    return render(request, 'clinica/agendar_cita.html', {
        'form': form,
        'paciente': paciente,
        'penalizacion_pendiente': penalizacion_pendiente,
    })

# --- Agrega esto al final de clinica/views.py ---

@login_required
def vista_calendario(request):
    return render(request, 'clinica/calendario.html')

@api_key_required
def calendario_citas(request):
    """API que devuelve las citas en formato JSON para FullCalendar"""
    from django.http import JsonResponse
    from datetime import datetime
    
    # Obtenemos todas las citas activas
    citas = Cita.objects.filter(estatus__in=Cita.ESTATUS_ACTIVOS)
    
    eventos = []
    for cita in citas:
        # FullCalendar necesita fecha y hora combinadas
        start = datetime.combine(cita.fecha, cita.hora)
        
        # Si no tienes hora_fin, calculamos 1 hora por defecto
        # (Si ya tienes hora_fin en tu modelo, úsalo: cita.hora_fin)
        from datetime import timedelta
        end = start + timedelta(hours=1) 

        # Colores según estatus
        color = '#3788d8' # Azul default
        if cita.estatus == Cita.ESTATUS_CONFIRMADA:
            color = '#28a745' # Verde
        elif cita.estatus == Cita.ESTATUS_SIN_CONFIRMAR:
            color = '#26C6DA' # Tu color INTRA Primary
        elif cita.estatus == Cita.ESTATUS_REAGENDO:
            color = '#f59f00'
        elif cita.estatus == Cita.ESTATUS_INCIDENCIA:
            color = '#6c757d'

        eventos.append({
            'title': f"{cita.pacientes_display()} ({cita.terapeuta})",
            'start': start.isoformat(),
            'end': end.isoformat(),
            'color': color,
            'url': f"/pacientes/{cita.paciente.id}/" # Al dar click, lleva al paciente
        })

    return JsonResponse(eventos, safe=False)
# En clinica/views.py

@login_required
def crear_cita(request):
    from clinica.models import PenalizacionPaciente
    if request.method == 'POST':
        form = CitaForm(request.POST)
        if form.is_valid():
            cita = form.save(commit=False)

            # Validar solapamiento de terapeuta y consultorio
            if cita.fecha and cita.hora:
                hora_sig = (datetime.combine(cita.fecha, cita.hora) + timedelta(hours=1)).time()

                if cita.terapeuta_id and Cita.objects.filter(
                    terapeuta_id=cita.terapeuta_id,
                    fecha=cita.fecha,
                    hora=cita.hora,
                    estatus__in=Cita.ESTATUS_ACTIVOS,
                ).exists():
                    messages.error(
                        request,
                        f'El terapeuta ya tiene una cita agendada a las {cita.hora.strftime("%H:%M")}. '
                        f'Puedes agendar desde las {hora_sig.strftime("%H:%M")}.',
                    )
                    form.add_error('terapeuta', 'Terapeuta ocupado a esa hora.')

                if cita.consultorio_id and Cita.objects.filter(
                    fecha=cita.fecha,
                    consultorio_id=cita.consultorio_id,
                    hora=cita.hora,
                    estatus__in=Cita.ESTATUS_ACTIVOS,
                ).exists():
                    messages.error(
                        request,
                        f'El consultorio ya tiene una cita agendada a las {cita.hora.strftime("%H:%M")}. '
                        f'Puedes agendar desde las {hora_sig.strftime("%H:%M")}.',
                    )
                    form.add_error('consultorio', 'Consultorio ocupado a esa hora.')

            if not form.errors:
                # Aplicar penalización pendiente del paciente si existe
                pen = None
                if cita.paciente_id:
                    pen = (
                        PenalizacionPaciente.objects
                        .filter(paciente_id=cita.paciente_id, pagada=False)
                        .order_by('-fecha_creacion')
                        .first()
                    )
                    if pen:
                        cita.costo = (cita.costo or 0) + pen.monto

                cita.save()
                _registrar_actividad(request, 'cita_creada', 'cita',
                    f'Cita de {cita.paciente.nombre if cita.paciente else "?"} con '
                    f'{cita.terapeuta.nombre if cita.terapeuta else "Sin terapeuta"} '
                    f'el {cita.fecha.strftime("%d/%m/%Y")} a las {cita.hora.strftime("%H:%M")} agendada.',
                    terapeuta=cita.terapeuta, paciente=cita.paciente)

                # Marcar penalización como cobrada.
                # El pago al terapeuta se registra en nómina cuando la cita se marque como 'si_asistio'.
                if pen:
                    pen.pagada = True
                    pen.cita_cobro = cita
                    pen.save(update_fields=['pagada', 'cita_cobro'])

                # Para servicios grupales guardamos pacientes adicionales en la misma cita.
                if es_servicio_grupal(cita.servicio):
                    pacientes_extra = form.cleaned_data.get('pacientes_extra')
                    if pacientes_extra:
                        cita.pacientes_adicionales.set(
                            pacientes_extra.exclude(pk=cita.paciente_id)
                        )
                else:
                    cita.pacientes_adicionales.clear()
                _vincular_expediente_grupal(cita)

                # --- NUEVA MAGIA: Limpieza de la bandeja de entrada ---
                solicitud_id = request.GET.get('solicitud')
                if solicitud_id:
                    try:
                        solicitud = SolicitudCita.objects.get(id=solicitud_id)
                        solicitud.estado = 'aceptada'
                        solicitud.save()
                        if solicitud.terapeuta:
                            NotificacionTerapeuta.objects.create(
                                terapeuta=solicitud.terapeuta,
                                tipo=NotificacionTerapeuta.TIPO_CITA_ACEPTADA,
                                mensaje=f"Tu solicitud de cita para {solicitud.paciente_nombre} el {solicitud.fecha_deseada.strftime('%d/%m/%Y')} fue aceptada y agendada.",
                            )
                    except SolicitudCita.DoesNotExist:
                        pass
                    else:
                        _registrar_actividad(request, 'solicitud_aceptada', 'solicitud',
                            f'Solicitud de cita de {solicitud.paciente_nombre} aceptada y agendada.',
                            terapeuta=solicitud.terapeuta)
                # ------------------------------------------------------

                messages.success(request, 'Cita agendada correctamente.')
                return redirect('home')
        else:
            # Tu logica de rastreo de errores intacta
            print("ERRORES DETECTADOS:", form.errors) 
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{error}")
            
            # ATENCION: Aqui NO hacemos redirect. Dejamos que el flujo continue 
            # hacia abajo para que vuelva a mostrar el formulario, pero ahora 
            # marcado con los errores y conservando lo que el usuario escribio.
            
    else:
        # Peticion GET: El usuario apenas va a abrir la pagina
        datos_iniciales = {}
        
        if 'fecha' in request.GET:
            datos_iniciales['fecha'] = request.GET.get('fecha')
        if 'hora' in request.GET:
            datos_iniciales['hora'] = request.GET.get('hora')
        if 'paciente' in request.GET:
            paciente_param = request.GET.get('paciente')
            if str(paciente_param).isdigit():
                datos_iniciales['paciente'] = paciente_param
            else:
                paciente_match = resolver_paciente_por_nombre(paciente_param)
                if paciente_match:
                    datos_iniciales['paciente'] = paciente_match.id
            
        # TRUCO: Lo convertimos a entero para que el campo Choice de Django lo acepte sin quejarse
        if 'terapeuta' in request.GET:
            try:
                datos_iniciales['terapeuta'] = int(request.GET.get('terapeuta'))
            except ValueError:
                pass
            
        solicitud_id = request.GET.get('solicitud')
        if solicitud_id:
            try:
                solicitud = SolicitudCita.objects.get(id=solicitud_id)
                # Llenamos el formulario con lo que pidió el paciente
                datos_iniciales['fecha'] = solicitud.fecha_deseada
                
                if solicitud.hora_deseada:
                    datos_iniciales['hora'] = solicitud.hora_deseada
                    
                if solicitud.terapeuta:
                    datos_iniciales['terapeuta'] = solicitud.terapeuta.id

                if solicitud.consultorio:
                    datos_iniciales['consultorio'] = solicitud.consultorio.id

                if solicitud.paciente:
                    datos_iniciales['paciente'] = solicitud.paciente.id
                else:
                    paciente_match = resolver_paciente_por_nombre(solicitud.paciente_nombre)
                    if paciente_match:
                        datos_iniciales['paciente'] = paciente_match.id

                if solicitud.division:
                    datos_iniciales['division'] = solicitud.division.id
                if solicitud.servicio:
                    datos_iniciales['servicio'] = solicitud.servicio.id
            except SolicitudCita.DoesNotExist:
                pass
        form = CitaForm(initial=datos_iniciales)

    # ESTO ES VITAL: Renderizamos la plantilla HTML en lugar de redirigir.
    # Asi el usuario puede ver el formulario para llenarlo o corregirlo.
    return render(request, 'clinica/crear_cita.html', {'form': form})


@api_key_required
def api_penalizacion_paciente(request):
    """Devuelve la penalización pendiente de un paciente, si existe."""
    from clinica.models import PenalizacionPaciente
    from django.http import JsonResponse
    paciente_id = request.GET.get('paciente_id')
    if not paciente_id:
        return JsonResponse({'penalizacion': None})
    pen = (
        PenalizacionPaciente.objects
        .filter(paciente_id=paciente_id, pagada=False)
        .select_related('cita_origen__servicio')
        .order_by('-fecha_creacion')
        .first()
    )
    if not pen:
        return JsonResponse({'penalizacion': None})
    return JsonResponse({
        'penalizacion': {
            'monto': str(pen.monto),
            'fecha_cita': pen.cita_origen.fecha.strftime('%d/%m/%Y'),
            'servicio': pen.cita_origen.servicio.nombre if pen.cita_origen.servicio else '',
        }
    })


@api_key_required
def api_pacientes_relacionados(request):
    paciente_id = request.GET.get('paciente_id')
    if not paciente_id:
        return JsonResponse({'relacionados': []})

    try:
        paciente = Paciente.objects.get(id=paciente_id)
    except Paciente.DoesNotExist:
        return JsonResponse({'relacionados': []})

    relacionados = Paciente.objects.exclude(id=paciente.id).order_by('nombre').values('id', 'nombre')
    return JsonResponse({'relacionados': list(relacionados)})

# En clinica/views.py

@api_key_required
def verificar_disponibilidad(request):
    fecha_str = request.GET.get('fecha')
    hora_str = request.GET.get('hora')
    consultorio_id = request.GET.get('consultorio')
    terapeuta_id = request.GET.get('terapeuta')
    cita_id = request.GET.get('cita_id')

    if not (fecha_str and hora_str and terapeuta_id):
        return JsonResponse({'available': True, 'msg': ''})

    _DIAS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']

    try:
        if len(hora_str) > 5:
            hora_str = hora_str[:5]

        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        hora_obj = datetime.strptime(hora_str, '%H:%M').time()

        # 1. Validar bloqueos del terapeuta
        bloqueo = obtener_bloqueo_terapeuta_en_fecha(terapeuta_id, fecha_obj, hora_obj)
        if bloqueo:
            return JsonResponse({'available': False, 'msg': bloqueo.mensaje_bloqueo()})

        # 2. Validar horario solo si el terapeuta tiene alguno configurado
        dia_semana = fecha_obj.weekday()
        tiene_horarios = Horario.objects.filter(terapeuta_id=terapeuta_id).exists()
        if not tiene_horarios:
            return JsonResponse({'available': True, 'msg': 'Disponible para agendar'})

        horarios_dia = list(Horario.objects.filter(terapeuta_id=terapeuta_id, dia=dia_semana))
        if not horarios_dia:
            return JsonResponse({
                'available': False,
                'msg': f'El terapeuta no tiene horario configurado los {_DIAS[dia_semana]}.',
            })
        horarios_activos = [h for h in horarios_dia if h.hora_inicio <= hora_obj <= h.hora_fin]
        if not horarios_activos:
            return JsonResponse({
                'available': False,
                'msg': f'Las {hora_str} está fuera del horario del terapeuta los {_DIAS[dia_semana]}.',
            })

        # 3. Validar que el terapeuta no tenga ya una cita a esa hora
        qs_excluir = Cita.objects.exclude(pk=cita_id) if cita_id else Cita.objects.all()
        duplicado = qs_excluir.filter(
            terapeuta_id=terapeuta_id,
            fecha=fecha_obj,
            hora=hora_obj,
            estatus__in=Cita.ESTATUS_ACTIVOS,
        ).exists()
        if duplicado:
            hora_siguiente = (datetime.combine(fecha_obj, hora_obj) + timedelta(hours=1)).time()
            return JsonResponse({
                'available': False,
                'msg': (
                    f'El terapeuta ya tiene una cita agendada a las {hora_str}. '
                    f'Puedes agendar desde las {hora_siguiente.strftime("%H:%M")}.'
                ),
            })

        # 4. Validar que el consultorio pertenezca a ALGUNA de las sedes activas
        if consultorio_id:
            from .models import Consultorio as ConsultorioModel
            _SEDES = dict(Horario.SEDE_CHOICES)
            try:
                cons = ConsultorioModel.objects.get(id=consultorio_id)
                if cons.sede:
                    match = any(not h.sede or h.sede == cons.sede for h in horarios_activos)
                    if not match:
                        sedes = ' o '.join(
                            _SEDES.get(h.sede, h.sede) for h in horarios_activos if h.sede
                        )
                        sede_c = _SEDES.get(cons.sede, cons.sede)
                        return JsonResponse({
                            'available': False,
                            'msg': f'El consultorio "{cons}" es de {sede_c}, pero el terapeuta trabaja en {sedes} a esa hora.',
                        })
            except ConsultorioModel.DoesNotExist:
                pass

        # 5. Validar que el consultorio no esté ocupado a esa hora
        if consultorio_id:
            ocupado = qs_excluir.filter(
                fecha=fecha_obj,
                consultorio_id=consultorio_id,
                hora=hora_obj,
                estatus__in=Cita.ESTATUS_ACTIVOS,
            ).exists()
            if ocupado:
                hora_siguiente = (datetime.combine(fecha_obj, hora_obj) + timedelta(hours=1)).time()
                return JsonResponse({
                    'available': False,
                    'msg': (
                        f'El consultorio ya tiene una cita agendada a las {hora_str}. '
                        f'Puedes agendar desde las {hora_siguiente.strftime("%H:%M")}.'
                    ),
                })

        return JsonResponse({'available': True, 'msg': 'Disponible para agendar'})

    except ValueError as e:
        print(f"Error de formato en verificar_disponibilidad: {e}")
        return JsonResponse({'available': True, 'msg': 'Disponible'})
    
# En clinica/views.py

@login_required
def editar_cita(request, cita_id):
    cita = get_object_or_404(Cita, id=cita_id)

    fecha_anterior     = cita.fecha
    hora_anterior      = cita.hora
    terapeuta_anterior = cita.terapeuta

    if request.method == 'POST':
        form = CitaForm(request.POST, instance=cita)
        if form.is_valid():
            nueva = form.save(commit=False)
            if nueva.fecha and nueva.hora:
                hora_sig = (datetime.combine(nueva.fecha, nueva.hora) + timedelta(hours=1)).time()
                otras = Cita.objects.exclude(pk=cita_id)

                # Solo validar solapamiento si el terapeuta, la fecha o la hora cambiaron.
                # Si únicamente cambió el estatus u otro campo, la cita ya ocupaba ese slot.
                slot_modificado = (
                    nueva.fecha != fecha_anterior
                    or nueva.hora != hora_anterior
                    or nueva.terapeuta_id != (terapeuta_anterior.id if terapeuta_anterior else None)
                )

                if slot_modificado and nueva.terapeuta_id and otras.filter(
                    terapeuta_id=nueva.terapeuta_id,
                    fecha=nueva.fecha,
                    hora=nueva.hora,
                    estatus__in=Cita.ESTATUS_ACTIVOS,
                ).exists():
                    messages.error(
                        request,
                        f'El terapeuta ya tiene una cita agendada a las {nueva.hora.strftime("%H:%M")}. '
                        f'Puedes agendar desde las {hora_sig.strftime("%H:%M")}.',
                    )
                    form.add_error('terapeuta', 'Terapeuta ocupado a esa hora.')

                if slot_modificado and nueva.consultorio_id and otras.filter(
                    fecha=nueva.fecha,
                    consultorio_id=nueva.consultorio_id,
                    hora=nueva.hora,
                    estatus__in=Cita.ESTATUS_ACTIVOS,
                ).exists():
                    messages.error(
                        request,
                        f'El consultorio ya tiene una cita agendada a las {nueva.hora.strftime("%H:%M")}. '
                        f'Puedes agendar desde las {hora_sig.strftime("%H:%M")}.',
                    )
                    form.add_error('consultorio', 'Consultorio ocupado a esa hora.')

            if form.errors:
                return render(request, 'clinica/editar_cita.html', {'form': form, 'cita': cita})

            cita = nueva
            cita.save()
            _registrar_actividad(request, 'cita_editada', 'cita',
                f'Cita de {cita.paciente.nombre if cita.paciente else "?"} con '
                f'{cita.terapeuta.nombre if cita.terapeuta else "Sin terapeuta"} '
                f'el {cita.fecha.strftime("%d/%m/%Y")} editada.',
                terapeuta=cita.terapeuta, paciente=cita.paciente)
            if es_servicio_grupal(cita.servicio):
                pacientes_extra = form.cleaned_data.get('pacientes_extra')
                cita.pacientes_adicionales.set(
                    (pacientes_extra or Paciente.objects.none()).exclude(pk=cita.paciente_id)
                )
            else:
                cita.pacientes_adicionales.clear()
            _vincular_expediente_grupal(cita)

            cambio_fecha     = cita.fecha != fecha_anterior
            cambio_hora      = cita.hora  != hora_anterior
            cambio_terapeuta = cita.terapeuta_id != (terapeuta_anterior.id if terapeuta_anterior else None)

            editor_es_terapeuta = hasattr(request.user, 'perfil_terapeuta')

            if (cambio_fecha or cambio_hora) and cita.terapeuta and not (
                editor_es_terapeuta and request.user.perfil_terapeuta == cita.terapeuta
            ):
                cambios = []
                if cambio_fecha:
                    cambios.append(f"fecha: {fecha_anterior.strftime('%d/%m/%Y')} → {cita.fecha.strftime('%d/%m/%Y')}")
                if cambio_hora:
                    cambios.append(f"hora: {hora_anterior.strftime('%H:%M')} → {cita.hora.strftime('%H:%M')}")
                NotificacionTerapeuta.objects.create(
                    terapeuta=cita.terapeuta,
                    tipo=NotificacionTerapeuta.TIPO_CITA_MODIFICADA,
                    mensaje=f"La cita de {cita.paciente} fue modificada. {' | '.join(cambios)}.",
                )

            if cambio_terapeuta and terapeuta_anterior and not (
                editor_es_terapeuta and request.user.perfil_terapeuta == terapeuta_anterior
            ):
                NotificacionTerapeuta.objects.create(
                    terapeuta=terapeuta_anterior,
                    tipo=NotificacionTerapeuta.TIPO_CITA_MODIFICADA,
                    mensaje=f"La cita de {cita.paciente} del {fecha_anterior.strftime('%d/%m/%Y')} a las {hora_anterior.strftime('%H:%M')} fue reasignada a otro terapeuta.",
                )
            if cambio_terapeuta and cita.terapeuta and not (
                editor_es_terapeuta and request.user.perfil_terapeuta == cita.terapeuta
            ):
                NotificacionTerapeuta.objects.create(
                    terapeuta=cita.terapeuta,
                    tipo=NotificacionTerapeuta.TIPO_CITA_MODIFICADA,
                    mensaje=f"Se te asignó una cita con {cita.paciente} el {cita.fecha.strftime('%d/%m/%Y')} a las {cita.hora.strftime('%H:%M')}.",
                )

            messages.success(request, '¡Cita actualizada correctamente! ')
            
            origen = request.GET.get('next', 'home')
            if origen == 'paciente':
                return redirect('detalle_paciente', paciente_id=cita.paciente.id)
            return redirect('home')
        else:
            # 👇 AQUÍ ESTÁ LA MAGIA: Mostramos el error exacto
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f" {error}")
                    else:
                        messages.error(request, f"Error en {field}: {error}")
    else:
        datos_iniciales = {}
        if 'fecha' in request.GET:
            datos_iniciales['fecha'] = request.GET.get('fecha')
        if 'hora' in request.GET:
            datos_iniciales['hora'] = request.GET.get('hora')
        if 'paciente' in request.GET:
            paciente_param = request.GET.get('paciente')
            if str(paciente_param).isdigit():
                datos_iniciales['paciente'] = paciente_param
            else:
                paciente_match = resolver_paciente_por_nombre(paciente_param)
                if paciente_match:
                    datos_iniciales['paciente'] = paciente_match.id
        if 'terapeuta' in request.GET:
            try:
                datos_iniciales['terapeuta'] = int(request.GET.get('terapeuta'))
            except (TypeError, ValueError):
                pass
        form = CitaForm(instance=cita, initial=datos_iniciales)

    return render(request, 'clinica/editar_cita.html', {
        'form': form, 
        'cita': cita
    }
    )


@login_required
def eliminar_cita(request, cita_id):
    if request.method != 'POST':
        return redirect('editar_cita', cita_id=cita_id)

    cita = get_object_or_404(Cita, id=cita_id)
    paciente_id = cita.paciente_id
    pac_nombre = cita.paciente.nombre if cita.paciente else '?'
    ter_nombre = cita.terapeuta.nombre if cita.terapeuta else 'Sin terapeuta'
    fecha_str  = cita.fecha.strftime('%d/%m/%Y') if cita.fecha else '?'
    hora_str   = cita.hora.strftime('%H:%M') if cita.hora else '?'
    cita.delete()
    _registrar_actividad(request, 'cita_eliminada', 'cita',
        f'Cita de {pac_nombre} con {ter_nombre} el {fecha_str} a las {hora_str} eliminada.')
    messages.success(request, 'Cita eliminada correctamente.')

    origen = request.GET.get('next', 'home')
    if origen == 'paciente' and paciente_id:
        return redirect('detalle_paciente', paciente_id=paciente_id)
    return redirect('home')


@login_required
def eliminar_paciente(request, paciente_id):
    if request.method != 'POST':
        return redirect('detalle_paciente', paciente_id=paciente_id)

    paciente = get_object_or_404(Paciente, id=paciente_id)
    nombre = paciente.nombre
    paciente.delete()
    _registrar_actividad(request, 'paciente_eliminado', 'paciente',
        f'Expediente de {nombre} eliminado del sistema.')
    messages.success(request, f'Expediente de {nombre} eliminado correctamente.')
    return redirect('lista_pacientes')


@api_key_required
def api_citas_calendario(request):
    start_str = request.GET.get('start', "")
    end_str = request.GET.get('end', "")

    citas = Cita.objects.select_related(
        'division', 'servicio', 'terapeuta', 'consultorio', 'paciente'
    ).prefetch_related(
        'pacientes_adicionales'
    )

    #Filtrar por las fechas que pide FullCalendar
    if start_str:
        try:
            start_date = datetime.fromisoformat(start_str).date()
            citas = citas.filter(fecha__gte=start_date)
        except ValueError:
            pass
    if end_str:
        try:
            end_date = datetime.fromisoformat(end_str).date()
            citas = citas.filter(fecha__lte=end_date)
        except ValueError:
            pass

    eventos = []
    
    for cita in citas:
        # FullCalendar necesita fecha y hora juntas en formato ISO
        start_datetime = datetime.combine(cita.fecha, cita.hora)
        # Asumimos que la cita dura 1 hora por defecto para pintar el bloque
        end_datetime = start_datetime + timedelta(hours=1)
        configuracion_estatus = obtener_configuracion_estatus_cita(cita.estatus)
        pacientes = cita.titulo_cita()
        consultorio = str(cita.consultorio) if cita.consultorio else 'Sin consultorio'
        terapeuta = str(cita.terapeuta) if cita.terapeuta else 'Sin terapeuta'
        servicio = str(cita.servicio) if cita.servicio else 'Sin servicio'
        division = str(cita.division) if cita.division else 'Sin division'
        costo = str(cita.costo) if cita.costo is not None else ''

        eventos.append({
            'id': cita.id,
            'title': pacientes,
            'start': start_datetime.isoformat(),
            'end': end_datetime.isoformat(),
            'backgroundColor': configuracion_estatus['bg'],
            'borderColor': configuracion_estatus['border'],
            'textColor': configuracion_estatus['text'],
            'classNames': ['evento-cita'],
            'extendedProps': {
                'paciente': pacientes,
                'paciente_upper': pacientes.upper(),
                'terapeuta': terapeuta,
                'servicio': servicio,
                'consultorio': consultorio,
                'division': division,
                'situacion': configuracion_estatus['label'],
                'situacion_color': configuracion_estatus['color'],
                'situacion_bg': configuracion_estatus['bg'],
                'situacion_border': configuracion_estatus['border'],
                'situacion_text': configuracion_estatus['text'],
                'tipo_paciente': cita.tipo_paciente,
                'tipo_paciente_label': cita.get_tipo_paciente_display(),
                'costo': costo,
                'notas': cita.notas or '',
                'editar_url': f"/citas/{cita.id}/editar/?next=calendario",
            },
        })
        
    return JsonResponse(eventos, safe=False)

@api_key_required
def api_terapeutas_paciente(request):
    paciente_id = request.GET.get('paciente_id')
    if not paciente_id:
        return JsonResponse({'terapeutas': []})
    
    # Buscamos todas las citas del paciente y sacamos los terapeutas
    citas = Cita.objects.filter(
        Q(paciente_id=paciente_id) | Q(pacientes_adicionales__id=paciente_id)
    ).select_related('terapeuta').distinct()
    
    terapeutas_unicos = set()
    for cita in citas:
        if cita.terapeuta:
            terapeutas_unicos.add(str(cita.terapeuta))
            
    return JsonResponse({'terapeutas': list(terapeutas_unicos)})

@login_required
def portal_terapeuta(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home') 
    
    mi_perfil = request.user.perfil_terapeuta
    hoy = date.today()
    
    # --- TRUCO PARA LA FECHA PERFECTA EN ESPAÑOL ---
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    fecha_bonita = f"{hoy.day} de {meses[hoy.month - 1]} del {hoy.year}"
    
    citas_hoy = Cita.objects.filter(
        terapeuta=mi_perfil,
        fecha=hoy,
    ).order_by('hora')

    citas_proximas = Cita.objects.filter(
        terapeuta=mi_perfil,
        fecha__gt=hoy,
    ).order_by('fecha', 'hora')[:10] 

    try:
        semana_offset = int(request.GET.get('semana_offset', 0))
    except (TypeError, ValueError):
        semana_offset = 0

    inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
    fin_semana = inicio_semana + timedelta(days=6)
    citas_semana = list(
        Cita.objects.filter(
            terapeuta=mi_perfil,
            fecha__gte=inicio_semana,
            fecha__lte=fin_semana,
        )
        .select_related('consultorio', 'servicio', 'paciente')
        .prefetch_related('pacientes_adicionales')
        .order_by('fecha', 'hora')
    )
    citas_por_fecha = defaultdict(list)
    for cita in citas_semana:
        citas_por_fecha[cita.fecha].append(cita)

    dias_semana_cortos = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    meses_cortos = ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 'jul', 'ago', 'sep', 'oct', 'nov', 'dic']
    agenda_semanal = []
    for offset in range(7):
        fecha_item = inicio_semana + timedelta(days=offset)
        agenda_semanal.append({
            'fecha': fecha_item,
            'nombre_corto': dias_semana_cortos[fecha_item.weekday()],
            'mes_corto': meses_cortos[fecha_item.month - 1],
            'es_hoy': fecha_item == hoy,
            'citas': citas_por_fecha.get(fecha_item, []),
        })
    
    mis_solicitudes = SolicitudCita.objects.filter(
        terapeuta=mi_perfil
    ).order_by('-fecha_creacion')[:5]

    mis_reagendos = SolicitudReagendo.objects.filter(
        terapeuta=mi_perfil
    ).select_related('cita', 'cita__paciente').order_by('-creado_en')[:10]
    bloqueos_futuros = mi_perfil.bloqueos_agenda.filter(
        activo=True,
    ).filter(
        Q(tipo_bloqueo=BloqueoAgendaTerapeuta.TIPO_PERMANENTE) |
        Q(fecha_fin__gte=hoy) |
        Q(fecha_fin__isnull=True, fecha_inicio__gte=hoy)
    ).order_by('alcance', 'dia_semana', 'fecha_inicio', 'hora_inicio')
    manual_portal = AccesoDirectoPortal.objects.filter(
        clave=AccesoDirectoPortal.CLAVE_MANUAL_PORTAL_MEDICO,
        activo=True,
    ).first()
    corte_pendiente_confirmacion = (
        CorteSemanal.objects.filter(
            terapeuta=mi_perfil,
            fecha_fin__lte=hoy,
            confirmacion_terapeuta__isnull=True,
        )
        .order_by('-fecha_fin')
        .first()
    )

    mis_solicitudes_expositor = SolicitudHorasExpositor.objects.filter(
        terapeuta=mi_perfil
    ).order_by('-creado_en')[:10]

    context = {
        'terapeuta': mi_perfil,
        'citas_hoy': citas_hoy,
        'citas_proximas': citas_proximas,
        'agenda_semanal': agenda_semanal,
        'agenda_inicio_semana': inicio_semana,
        'agenda_fin_semana': fin_semana,
        'agenda_semana_offset': semana_offset,
        'mostrar_agenda_semanal': request.GET.get('ver_agenda') == '1',
        'fecha_bonita': fecha_bonita,
        'mis_solicitudes': mis_solicitudes,
        'mis_reagendos': mis_reagendos,
        'mis_solicitudes_expositor': mis_solicitudes_expositor,
        'bloqueos_agenda': bloqueos_futuros,
        'bloqueo_form': BloqueoAgendaTerapeutaForm(),
        'manual_portal': manual_portal,
        'corte_pendiente_confirmacion': corte_pendiente_confirmacion,
    }
    
    return render(request, 'clinica/portal_terapeuta.html', context)


@login_required
def mi_disponibilidad_terapeuta(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    mi_perfil = request.user.perfil_terapeuta

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'agregar':
            dia = request.POST.get('dia')
            hora_inicio = request.POST.get('hora_inicio')
            hora_fin = request.POST.get('hora_fin')
            sede = request.POST.get('sede') or None
            if not (dia and hora_inicio and hora_fin and sede):
                messages.error(request, 'Completa todos los campos, incluyendo la sede.')
            elif hora_fin <= hora_inicio:
                messages.error(request, 'La hora de fin debe ser mayor que la de inicio.')
            else:
                Horario.objects.create(
                    terapeuta=mi_perfil,
                    dia=int(dia),
                    hora_inicio=hora_inicio,
                    hora_fin=hora_fin,
                    sede=sede,
                )
                messages.success(request, 'Franja horaria agregada correctamente.')
        elif action == 'eliminar':
            horario_id = request.POST.get('horario_id')
            horario = get_object_or_404(Horario, id=horario_id, terapeuta=mi_perfil)
            horario.delete()
            messages.success(request, 'Franja horaria eliminada.')
        return redirect('mi_disponibilidad_terapeuta')

    horarios_qs = list(Horario.objects.filter(terapeuta=mi_perfil).order_by('dia', 'hora_inicio'))
    dias_data = []
    for dia_num, dia_nombre in Horario.DIAS_SEMANA:
        slots = [h for h in horarios_qs if h.dia == dia_num]
        dias_data.append({'num': dia_num, 'nombre': dia_nombre, 'slots': slots})

    return render(request, 'clinica/mi_disponibilidad_terapeuta.html', {
        'terapeuta': mi_perfil,
        'dias_data': dias_data,
        'sede_choices': Horario.SEDE_CHOICES,
        'total': len(horarios_qs),
    })


@login_required
def descargar_manual_portal_medico(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    manual = AccesoDirectoPortal.objects.filter(
        clave=AccesoDirectoPortal.CLAVE_MANUAL_PORTAL_MEDICO,
        activo=True,
    ).first()
    if not manual or not manual.tiene_archivo:
        messages.error(request, 'El manual del sistema no esta disponible por el momento.')
        return redirect('portal_terapeuta')

    response = HttpResponse(
        bytes(manual.contenido),
        content_type=manual.tipo_mime or 'application/octet-stream',
    )
    response['Content-Disposition'] = f'attachment; filename="{manual.nombre_archivo or "manual_sistema"}"'
    return response


@login_required
def recursos_propios(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'subir':
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            archivo = request.FILES.get('archivo')
            if not nombre or not archivo:
                messages.error(request, 'El nombre y el archivo son obligatorios.')
            else:
                RecursoPropio.objects.create(
                    nombre=nombre,
                    descripcion=descripcion,
                    nombre_archivo=archivo.name,
                    tipo_mime=archivo.content_type or 'application/octet-stream',
                    contenido=archivo.read(),
                    subido_por=request.user,
                )
                messages.success(request, f'"{nombre}" agregado correctamente.')
            return redirect('recursos_propios')

        if accion == 'eliminar':
            recurso_id = request.POST.get('recurso_id')
            RecursoPropio.objects.filter(pk=recurso_id, subido_por=request.user).delete()
            messages.success(request, 'Recurso eliminado.')
            return redirect('recursos_propios')

    recursos = RecursoPropio.objects.filter(subido_por=request.user)
    return render(request, 'clinica/recursos_propios.html', {'recursos': recursos})


@login_required
def descargar_recurso_propio(request, recurso_id):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')
    recurso = get_object_or_404(RecursoPropio, pk=recurso_id, subido_por=request.user)
    response = HttpResponse(bytes(recurso.contenido), content_type=recurso.tipo_mime or 'application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{recurso.nombre_archivo}"'
    return response


@login_required
def crear_bloqueo_terapeuta(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    if request.method != 'POST':
        return redirect('portal_terapeuta')

    mi_perfil = request.user.perfil_terapeuta
    form = BloqueoAgendaTerapeutaForm(request.POST)
    if form.is_valid():
        bloqueo = form.save(commit=False)
        bloqueo.terapeuta = mi_perfil
        bloqueo.creado_por = request.user
        bloqueo.save()
        _registrar_actividad(request, 'bloqueo_creado', 'disponibilidad',
            f'{mi_perfil.nombre} creó un bloqueo de agenda '
            f'({bloqueo.get_tipo_bloqueo_display()}, desde {bloqueo.fecha_inicio.strftime("%d/%m/%Y")}).',
            terapeuta=mi_perfil)
        messages.success(request, 'Bloqueo de agenda guardado correctamente.')
    else:
        for _, errors in form.errors.items():
            for error in errors:
                messages.error(request, error)

    return redirect('portal_terapeuta')


@login_required
def eliminar_bloqueo_terapeuta(request, bloqueo_id):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    bloqueo = get_object_or_404(
        BloqueoAgendaTerapeuta,
        id=bloqueo_id,
        terapeuta=request.user.perfil_terapeuta,
    )
    ter_perfil = bloqueo.terapeuta
    fecha_str  = bloqueo.fecha_inicio.strftime('%d/%m/%Y') if bloqueo.fecha_inicio else '?'
    bloqueo.delete()
    _registrar_actividad(request, 'bloqueo_eliminado', 'disponibilidad',
        f'{ter_perfil.nombre if ter_perfil else "?"} eliminó un bloqueo de agenda desde {fecha_str}.',
        terapeuta=ter_perfil)
    messages.success(request, 'Bloqueo eliminado correctamente.')
    return redirect('portal_terapeuta')


@login_required
def portal_consultoria(request):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')

    mi_perfil = request.user.perfil_consultoria
    hoy = date.today()
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    fecha_bonita = f"{hoy.day} de {meses[hoy.month - 1]} del {hoy.year}"

    divisiones_consultoria = mi_perfil.divisiones.all().order_by('nombre')
    division_ids = list(divisiones_consultoria.values_list('id', flat=True))

    fecha_inicio_str = request.GET.get('fecha_inicio') or request.GET.get('fecha') or hoy.isoformat()
    fecha_fin_str = request.GET.get('fecha_fin') or request.GET.get('fecha') or fecha_inicio_str
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        fecha_inicio = hoy
    try:
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        fecha_fin = fecha_inicio
    if fecha_fin < fecha_inicio:
        fecha_inicio, fecha_fin = fecha_fin, fecha_inicio

    division_filtro = request.GET.get('division') or ''
    estatus_filtro = request.GET.get('estatus') or ''
    terapeuta_filtro = request.GET.get('terapeuta') or ''
    q_filtro = (request.GET.get('q') or '').strip()

    citas_base = (
        Cita.objects.filter(
            Q(division_id__in=division_ids) |
            Q(paciente__division_id__in=division_ids) |
            Q(pacientes_adicionales__division_id__in=division_ids)
        )
        .select_related('paciente', 'paciente__division', 'division', 'terapeuta', 'consultorio', 'servicio')
        .prefetch_related('pacientes_adicionales', 'pacientes_adicionales__division')
        .distinct()
    )

    terapeutas_filtro = (
        Terapeuta.objects
        .filter(cita__in=citas_base)
        .distinct()
        .order_by('nombre')
    )

    citas_filtradas = citas_base.filter(fecha__gte=fecha_inicio, fecha__lte=fecha_fin)
    if division_filtro:
        citas_filtradas = citas_filtradas.filter(
            Q(division_id=division_filtro) |
            Q(paciente__division_id=division_filtro) |
            Q(pacientes_adicionales__division_id=division_filtro)
        ).distinct()
    if estatus_filtro:
        citas_filtradas = citas_filtradas.filter(estatus=estatus_filtro)
    if terapeuta_filtro:
        citas_filtradas = citas_filtradas.filter(terapeuta_id=terapeuta_filtro)
    if q_filtro:
        citas_filtradas = citas_filtradas.filter(
            Q(paciente__nombre__icontains=q_filtro) |
            Q(pacientes_adicionales__nombre__icontains=q_filtro) |
            Q(notas__icontains=q_filtro)
        ).distinct()

    citas_hoy = citas_filtradas.order_by('fecha', 'hora', 'paciente__nombre')
    citas_proximas = citas_base.filter(fecha__gt=fecha_fin).order_by('fecha', 'hora')[:10]

    try:
        semana_offset = int(request.GET.get('semana_offset', 0))
    except (TypeError, ValueError):
        semana_offset = 0

    inicio_semana = fecha_inicio - timedelta(days=fecha_inicio.weekday()) + timedelta(weeks=semana_offset)
    fin_semana = inicio_semana + timedelta(days=6)
    citas_semana = list(
        citas_base.filter(fecha__gte=inicio_semana, fecha__lte=fin_semana)
        .order_by('fecha', 'hora')
    )
    citas_por_fecha = defaultdict(list)
    for cita in citas_semana:
        citas_por_fecha[cita.fecha].append(cita)

    dias_semana_cortos = ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab', 'Dom']
    meses_cortos = ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 'jul', 'ago', 'sep', 'oct', 'nov', 'dic']
    agenda_semanal = []
    for offset in range(7):
        fecha_item = inicio_semana + timedelta(days=offset)
        agenda_semanal.append({
            'fecha': fecha_item,
            'nombre_corto': dias_semana_cortos[fecha_item.weekday()],
            'mes_corto': meses_cortos[fecha_item.month - 1],
            'es_hoy': fecha_item == hoy,
            'citas': citas_por_fecha.get(fecha_item, []),
        })

    manual_portal = AccesoDirectoPortal.objects.filter(
        clave=AccesoDirectoPortal.CLAVE_MANUAL_PORTAL_MEDICO,
        activo=True,
    ).first()

    context = {
        'consultoria': mi_perfil,
        'terapeuta': mi_perfil,
        'citas_hoy': citas_hoy,
        'citas_proximas': citas_proximas,
        'divisiones_consultoria': divisiones_consultoria,
        'terapeutas_filtro': terapeutas_filtro,
        'estatus_choices': Cita.ESTATUS_CHOICES,
        'filtros_consultoria': {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'division': division_filtro,
            'estatus': estatus_filtro,
            'terapeuta': terapeuta_filtro,
            'q': q_filtro,
        },
        'total_citas_filtradas': citas_hoy.count(),
        'agenda_semanal': agenda_semanal,
        'agenda_inicio_semana': inicio_semana,
        'agenda_fin_semana': fin_semana,
        'agenda_semana_offset': semana_offset,
        'mostrar_agenda_semanal': request.GET.get('ver_agenda') == '1',
        'fecha_bonita': fecha_bonita,
        'mis_solicitudes': SolicitudCita.objects.none(),
        'mis_reagendos': SolicitudReagendo.objects.none(),
        'mis_solicitudes_expositor': SolicitudHorasExpositor.objects.none(),
        'bloqueos_agenda': [],
        'bloqueo_form': BloqueoAgendaTerapeutaForm(),
        'manual_portal': manual_portal,
        'corte_pendiente_confirmacion': None,
    }
    return render(request, 'clinica/portal_consultoria.html', context)


@login_required
def mi_disponibilidad_consultoria(request):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')

    if request.method == 'POST':
        messages.info(request, 'La disponibilidad de Consultoria ya tiene vista propia; falta conectar su almacenamiento independiente.')
        return redirect('mi_disponibilidad_consultoria')

    dias_data = [
        {'num': dia_num, 'nombre': dia_nombre, 'slots': []}
        for dia_num, dia_nombre in Horario.DIAS_SEMANA
    ]
    return render(request, 'clinica/mi_disponibilidad_consultoria.html', {
        'consultoria': request.user.perfil_consultoria,
        'terapeuta': request.user.perfil_consultoria,
        'dias_data': dias_data,
        'sede_choices': Horario.SEDE_CHOICES,
        'total': 0,
    })


@login_required
def descargar_manual_portal_consultoria(request):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')

    manual = AccesoDirectoPortal.objects.filter(
        clave=AccesoDirectoPortal.CLAVE_MANUAL_PORTAL_MEDICO,
        activo=True,
    ).first()
    if not manual or not manual.tiene_archivo:
        messages.error(request, 'El manual del sistema no esta disponible por el momento.')
        return redirect('portal_consultoria')

    response = HttpResponse(
        bytes(manual.contenido),
        content_type=manual.tipo_mime or 'application/octet-stream',
    )
    response['Content-Disposition'] = f'attachment; filename="{manual.nombre_archivo or "manual_sistema"}"'
    return response


@login_required
def recursos_consultoria(request):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'subir':
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            archivo = request.FILES.get('archivo')
            if not nombre or not archivo:
                messages.error(request, 'El nombre y el archivo son obligatorios.')
            else:
                RecursoPropio.objects.create(
                    nombre=nombre,
                    descripcion=descripcion,
                    nombre_archivo=archivo.name,
                    tipo_mime=archivo.content_type or 'application/octet-stream',
                    contenido=archivo.read(),
                    subido_por=request.user,
                )
                messages.success(request, f'"{nombre}" agregado correctamente.')
            return redirect('recursos_consultoria')

        if accion == 'eliminar':
            recurso_id = request.POST.get('recurso_id')
            RecursoPropio.objects.filter(pk=recurso_id, subido_por=request.user).delete()
            messages.success(request, 'Recurso eliminado.')
            return redirect('recursos_consultoria')

    recursos = RecursoPropio.objects.filter(subido_por=request.user)
    return render(request, 'clinica/recursos_consultoria.html', {'recursos': recursos})


@login_required
def descargar_recurso_consultoria(request, recurso_id):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')
    recurso = get_object_or_404(RecursoPropio, pk=recurso_id, subido_por=request.user)
    response = HttpResponse(bytes(recurso.contenido), content_type=recurso.tipo_mime or 'application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{recurso.nombre_archivo}"'
    return response


@login_required
def crear_bloqueo_consultoria(request):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')
    messages.info(request, 'La gestion de bloqueos de Consultoria ya esta separada; falta conectar su modelo propio.')
    return redirect('portal_consultoria')


@login_required
def eliminar_bloqueo_consultoria(request, bloqueo_id):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')
    messages.info(request, 'La gestion de bloqueos de Consultoria ya esta separada; falta conectar su modelo propio.')
    return redirect('portal_consultoria')


@login_required
def confirmar_nomina_consultoria(request, corte_id):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')
    messages.info(request, 'La nomina de Consultoria usara un flujo propio.')
    return redirect('portal_consultoria')


@login_required
def solicitar_horas_expositor_consultoria(request):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')
    messages.info(request, 'La solicitud de horas expositor de Consultoria ya tiene endpoint propio; falta conectar su almacenamiento.')
    return redirect('portal_consultoria')


@login_required
def reportar_incidente_consultoria(request):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')
    messages.info(request, 'El reporte de incidencias de Consultoria ya tiene endpoint propio; falta conectar su almacenamiento.')
    return redirect('portal_consultoria')


@login_required
def expedientes_consultoria(request):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')
    messages.info(request, 'Los expedientes de Consultoria usaran una vista propia.')
    return redirect('portal_consultoria')


@login_required
def registrar_paciente_consultoria(request):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')
    messages.info(request, 'El alta de pacientes de Consultoria usara una vista propia.')
    return redirect('portal_consultoria')


@login_required
def solicitar_cita_consultoria(request):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')

    mi_perfil = request.user.perfil_consultoria
    divisiones = mi_perfil.divisiones.all().order_by('nombre')
    division_ids = list(divisiones.values_list('id', flat=True))
    pacientes = (
        Paciente.objects
        .filter(
            Q(division_id__in=division_ids) |
            Q(empresa__division_id__in=division_ids) |
            Q(citas__division_id__in=division_ids) |
            Q(citas_como_adicional__division_id__in=division_ids)
        )
        .select_related('division', 'empresa', 'empresa__division')
        .distinct()
        .order_by('nombre')
    )
    aconsejados = []
    for paciente in pacientes:
        division_base = paciente.division or (paciente.empresa.division if paciente.empresa_id and paciente.empresa else None)
        if not division_base or division_base.id not in division_ids:
            division_base = (
                Division.objects
                .filter(Q(cita__paciente=paciente) | Q(cita__pacientes_adicionales=paciente), id__in=division_ids)
                .distinct()
                .order_by('nombre')
                .first()
            )
        aconsejados.append({
            'paciente': paciente,
            'division_id': division_base.id if division_base else '',
            'division_nombre': division_base.nombre if division_base else 'Sin division',
        })
    terapeutas = Terapeuta.objects.filter(
        activo=True, horario__isnull=False
    ).distinct().order_by('nombre')
    consultorios = Consultorio.objects.filter(activo=True).order_by('nombre')
    servicios = Servicio.objects.all().order_by('nombre')

    if request.method == 'POST':
        paciente_id = request.POST.get('paciente')
        division_id = request.POST.get('division')
        terapeuta_id = request.POST.get('terapeuta') or None
        consultorio_id = request.POST.get('consultorio') or None
        servicio_id = request.POST.get('servicio') or None
        fecha_str = request.POST.get('fecha_deseada', '').strip()
        hora_str = request.POST.get('hora_deseada', '').strip()
        notas = request.POST.get('notas_consultoria', '').strip()
        errores = []
        paciente = None
        division = None
        fecha_obj = None
        hora_obj = None

        try:
            paciente = pacientes.get(pk=paciente_id)
        except Exception:
            errores.append('Selecciona un aconsejado valido.')

        if division_id:
            try:
                division = divisiones.get(pk=division_id)
            except Exception:
                errores.append('Selecciona una division valida.')
        elif paciente:
            division = paciente.division or (paciente.empresa.division if paciente.empresa_id and paciente.empresa else None)
            if not division or division.id not in division_ids:
                errores.append('El aconsejado no tiene una division disponible para Consultoria.')
        else:
            errores.append('Selecciona una division.')

        if paciente and division:
            pertenece_a_division = (
                paciente.division_id == division.id or
                (paciente.empresa_id and paciente.empresa and paciente.empresa.division_id == division.id) or
                Cita.objects.filter(
                    Q(paciente=paciente) | Q(pacientes_adicionales=paciente),
                    division=division,
                ).exists()
            )
            if not pertenece_a_division:
                errores.append('El aconsejado seleccionado no pertenece a esa division.')

        if not fecha_str:
            errores.append('La fecha es requerida.')
        else:
            try:
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            except ValueError:
                errores.append('Fecha invalida.')

        if hora_str:
            try:
                hora_obj = datetime.strptime(hora_str, '%H:%M').time()
            except ValueError:
                errores.append('Hora invalida.')

        if errores:
            for error in errores:
                messages.error(request, error)
        else:
            SolicitudCita.objects.create(
                paciente_nombre=paciente.nombre,
                telefono=paciente.telefono or '',
                fecha_deseada=fecha_obj,
                hora_deseada=hora_obj,
                terapeuta_id=terapeuta_id,
                consultorio_id=consultorio_id,
                servicio_id=servicio_id,
                paciente=paciente,
                empresa=paciente.empresa,
                division=division,
                notas_paciente=f'SOLICITADO POR CONSULTORIA ({mi_perfil.nombre}): {notas}' if notas else f'SOLICITADO POR CONSULTORIA ({mi_perfil.nombre})',
                estado='pendiente',
            )
            messages.success(request, f'Solicitud enviada para {paciente.nombre}. Recepcion la confirmara pronto.')
            return redirect('portal_consultoria')

    return render(request, 'clinica/solicitar_cita_consultoria.html', {
        'consultoria': mi_perfil,
        'divisiones': divisiones,
        'pacientes': pacientes,
        'aconsejados': aconsejados,
        'terapeutas': terapeutas,
        'consultorios': consultorios,
        'servicios': servicios,
    })


@login_required
def expediente_consultoria_detalle(request, paciente_id):
    if not hasattr(request.user, 'perfil_consultoria'):
        return redirect('home')
    messages.info(request, 'El detalle de expediente de Consultoria usara una vista propia.')
    return redirect('portal_consultoria')

@login_required
def portal_paciente(request):
    # 1. Verificamos si el usuario actual tiene un perfil de paciente
    if not hasattr(request.user, 'perfil_paciente'):
        return redirect('home') 
    
    # 2. Identificamos al paciente exacto
    mi_perfil = request.user.perfil_paciente
    hoy = date.today()
    
    # 3. Filtramos sus citas futuras y pasadas
    citas_proximas = Cita.objects.filter(
        Q(paciente=mi_perfil) | Q(pacientes_adicionales=mi_perfil),
        fecha__gte=hoy
    ).distinct().order_by('fecha', 'hora')
    
    historial = Cita.objects.filter(
        Q(paciente=mi_perfil) | Q(pacientes_adicionales=mi_perfil),
        fecha__lt=hoy
    ).distinct().order_by('-fecha', '-hora')
    
  
    mis_solicitudes = SolicitudCita.objects.filter(
        paciente_nombre=mi_perfil.nombre
    ).order_by('-fecha_creacion')[:5] 
    
    context = {
        'paciente': mi_perfil,
        'citas_proximas': citas_proximas,
        'historial': historial,
        'mis_solicitudes': mis_solicitudes, # <--- Lo agregamos al paquete
    }
    
    return render(request, 'clinica/portal_paciente.html', context)

@login_required
def solicitar_cita_paciente(request):
    # Verificamos que sea un paciente
    if not hasattr(request.user, 'perfil_paciente'):
        return redirect('home')
        
    mi_perfil = request.user.perfil_paciente
    
    if request.method == 'POST':
        # Atrapamos lo que el paciente lleno en el formulario HTML
        fecha = request.POST.get('fecha_deseada')
        hora = request.POST.get('hora_deseada')
        terapeuta_id = request.POST.get('terapeuta')
        notas = request.POST.get('notas_paciente')
        
        consultorio_id = request.POST.get('consultorio')

        # Creamos el ticket en nuestra "Sala de Espera" (SolicitudCita)
        SolicitudCita.objects.create(
            paciente_nombre=mi_perfil.nombre,
            telefono=mi_perfil.telefono, # Asumiendo que tu modelo Paciente tiene campo telefono
            fecha_deseada=fecha,
            hora_deseada=hora if hora else None,
            terapeuta_id=terapeuta_id if terapeuta_id else None,
            consultorio_id=consultorio_id if consultorio_id else None,
            notas_paciente=notas,
            estado='pendiente'
        )
        
        # Le avisamos que todo salio bien y lo regresamos a su portal
        messages.success(request, '¡Tu solicitud ha sido enviada! Recepción la revisará y te confirmará pronto.')
        return redirect('portal_paciente')
        
    # Si apenas va a abrir la pagina (GET), le mandamos la lista de terapeutas y consultorios
    terapeutas = Terapeuta.objects.filter(activo=True, horario__isnull=False).distinct().order_by('nombre')
    from .models import Consultorio
    consultorios = Consultorio.objects.filter(activo=True)
    return render(request, 'clinica/solicitar_cita.html', {'terapeutas': terapeutas, 'consultorios': consultorios})

@login_required
def rechazar_solicitud(request, solicitud_id):
    if request.method == 'POST':
        try:
            from .models import SolicitudCita # Aseguramos que este import global funcione
            solicitud = SolicitudCita.objects.get(id=solicitud_id)
            
            # Atrapamos lo que escribio la recepcionista
            motivo = request.POST.get('motivo_rechazo', '')
            
            solicitud.estado = 'rechazada'
            solicitud.motivo_rechazo = motivo
            solicitud.save()
            _registrar_actividad(request, 'solicitud_rechazada', 'solicitud',
                f'Solicitud de cita de {solicitud.paciente_nombre} para el '
                f'{solicitud.fecha_deseada.strftime("%d/%m/%Y")} rechazada.',
                terapeuta=solicitud.terapeuta)
            if solicitud.terapeuta:
                motivo_texto = f" Motivo: {motivo}" if motivo else ""
                NotificacionTerapeuta.objects.create(
                    terapeuta=solicitud.terapeuta,
                    tipo=NotificacionTerapeuta.TIPO_CITA_RECHAZADA,
                    mensaje=f"Tu solicitud de cita para {solicitud.paciente_nombre} el {solicitud.fecha_deseada.strftime('%d/%m/%Y')} fue rechazada.{motivo_texto}",
                )
            messages.warning(request, f'La solicitud de {solicitud.paciente_nombre} fue rechazada correctamente.')
        except SolicitudCita.DoesNotExist:
            pass
            
    return redirect('home')

@login_required
def solicitar_cita_terapeuta(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')
        
    mi_perfil = request.user.perfil_terapeuta
    
    # Importamos los modelos necesarios
    from .models import SolicitudCita, Paciente
    
    if request.method == 'POST':
        # Ahora este dato vendra del menu desplegable (exactamente como esta escrito en la BD)
        paciente = request.POST.get('paciente_nombre')
        telefono = request.POST.get('telefono', '')
        fecha = request.POST.get('fecha_deseada')
        hora = request.POST.get('hora_deseada')
        notas = request.POST.get('notas_terapeuta', '')
        consultorio_id = request.POST.get('consultorio')
        servicio_id = request.POST.get('servicio')

        SolicitudCita.objects.create(
            paciente_nombre=paciente,
            telefono=telefono,
            fecha_deseada=fecha,
            hora_deseada=hora if hora else None,
            terapeuta=mi_perfil,
            consultorio_id=consultorio_id if consultorio_id else None,
            servicio_id=servicio_id if servicio_id else None,
            notas_paciente=f"SOLICITADO POR TERAPEUTA: {notas}",
            estado='pendiente'
        )
        
        messages.success(request, 'Solicitud enviada a Recepción. Espera su confirmación.')
        return redirect('portal_terapeuta')
        
    # --- NUEVO: Traemos la lista de pacientes ordenados alfabeticamente ---
    pacientes = Paciente.objects.all().order_by('nombre')
    from .models import Consultorio, Servicio
    consultorios = Consultorio.objects.filter(activo=True)
    servicios = Servicio.objects.all().order_by('nombre')

    return render(request, 'clinica/solicitar_cita_terapeuta.html', {
        'pacientes': pacientes,
        'consultorios': consultorios,
        'servicios': servicios,
        'mi_perfil': mi_perfil,
    })

@api_key_required
def api_disponibilidad_terapeuta(request):
    fecha_str = request.GET.get('fecha')
    terapeuta_id = request.GET.get('terapeuta')

    if not (fecha_str and terapeuta_id):
        return JsonResponse({'horarios': []})

    try:
        from datetime import datetime, timedelta
        from .models import Horario, Cita
        
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        dia_semana = fecha_obj.weekday()
        bloqueos = list(
            BloqueoAgendaTerapeuta.objects.filter(
                terapeuta_id=terapeuta_id,
                activo=True,
                fecha_inicio__lte=fecha_obj,
            ).filter(
                Q(tipo_bloqueo=BloqueoAgendaTerapeuta.TIPO_PERMANENTE) |
                Q(fecha_fin__gte=fecha_obj)
            ).order_by('hora_inicio', 'fecha_inicio')
        )

        bloqueo_total = next(
            (bloqueo for bloqueo in bloqueos if bloqueo.bloquea_fecha_hora(fecha_obj)),
            None,
        )

        if bloqueo_total and not bloqueo_total.es_bloqueo_parcial():
            return JsonResponse({
                'bloqueado': True,
                'mensaje': bloqueo_total.mensaje_bloqueo(),
                'horarios': [],
            })

        # 1. Buscar si el terapeuta trabaja ese dia
        horarios_laborales = Horario.objects.filter(
            terapeuta_id=terapeuta_id, 
            dia=dia_semana
        )

        if not horarios_laborales.exists():
            return JsonResponse({'mensaje': 'El terapeuta no tiene horario laboral registrado para este dia.', 'horarios': []})

        # 2. Calcular capacidad por slot (cuántos horarios cubren esa hora)
        from collections import Counter
        capacidad = Counter()
        for hl in horarios_laborales:
            hora_actual = datetime.combine(fecha_obj, hl.hora_inicio)
            hora_fin_dt = datetime.combine(fecha_obj, hl.hora_fin)
            while hora_actual <= hora_fin_dt:
                capacidad[hora_actual.time()] += 1
                hora_actual += timedelta(minutes=60)

        # 3. Contar citas activas por hora
        citas_ocupadas = list(Cita.objects.filter(
            terapeuta_id=terapeuta_id,
            fecha=fecha_obj,
            estatus__in=Cita.ESTATUS_ACTIVOS,
        ).values_list('hora', flat=True))
        citas_por_hora = Counter(citas_ocupadas)

        # 4. Un slot es libre si citas_ocupadas < capacidad y no está bloqueado
        horarios_libres = []
        for slot, cap in capacidad.items():
            if citas_por_hora.get(slot, 0) >= cap:
                continue
            bloqueo_slot = obtener_bloqueo_terapeuta_en_fecha(terapeuta_id, fecha_obj, slot)
            if bloqueo_slot:
                continue
            horarios_libres.append(slot.strftime('%H:%M'))

        mensaje = ''
        if any(b.es_bloqueo_parcial() and b.aplica_en_fecha(fecha_obj) for b in bloqueos):
            mensaje = 'Hay horas bloqueadas por el terapeuta en esta fecha.'

        sedes_del_dia = list(
            horarios_laborales.exclude(sede__isnull=True)
            .values_list('sede', flat=True)
            .distinct()
        )
        consultorios_disponibles = list(
            Consultorio.objects.filter(sede__in=sedes_del_dia, activo=True)
            .values('id', 'nombre')
            .order_by('nombre')
        ) if sedes_del_dia else []

        return JsonResponse({
            'horarios': sorted(list(set(horarios_libres))),
            'mensaje': mensaje,
            'consultorios': consultorios_disponibles,
        })

    except Exception as e:
        print(f"Error en radar: {e}")
        return JsonResponse({'horarios': [], 'error': 'Error procesando datos'})


@login_required
def checkout_cita(request, cita_id):
    """
    Procesa el cierre de sesión de una cita desde el portal del terapeuta.
    Solo acepta POST (el formulario se envía desde el modal en portal_terapeuta.html).
    Opcionalmente crea una SolicitudCita de seguimiento si el terapeuta lo indicó.
    """
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    mi_perfil = request.user.perfil_terapeuta
    cita = get_object_or_404(Cita, id=cita_id, terapeuta=mi_perfil)

    if not cita.es_finalizable:
        messages.warning(request, f'La cita de {cita.paciente.nombre} ya fue finalizada anteriormente.')
        return redirect('portal_terapeuta')

    if request.method != 'POST':
        return redirect('portal_terapeuta')

    form = CheckoutCitaForm(request.POST)
    if form.is_valid():
        # Actualizar cita
        cita.estatus = form.cleaned_data['estatus']
        metodo = form.cleaned_data.get('metodo_pago')
        costo = form.cleaned_data.get('costo')
        if metodo:
            cita.metodo_pago = metodo
        if costo is not None:
            cita.costo = costo
        cita.save()
        estatus_display = dict(Cita.ESTATUS_CHOICES).get(cita.estatus, cita.estatus)
        _registrar_actividad(request, 'cita_checkout', 'cita',
            f'{mi_perfil.nombre} cerró la sesión de {cita.paciente.nombre if cita.paciente else "?"} '
            f'el {cita.fecha.strftime("%d/%m/%Y")}: {estatus_display}.',
            terapeuta=mi_perfil, paciente=cita.paciente)

        # Crear solicitud de seguimiento si fue solicitada
        if form.cleaned_data.get('solicitar_siguiente') and form.cleaned_data.get('siguiente_fecha'):
            SolicitudCita.objects.create(
                paciente_nombre=cita.paciente.nombre,
                telefono=cita.paciente.telefono or '',
                fecha_deseada=form.cleaned_data['siguiente_fecha'],
                hora_deseada=form.cleaned_data.get('siguiente_hora'),
                terapeuta=mi_perfil,
                notas_paciente=form.cleaned_data.get('notas_recepcion') or '',
                estado='pendiente',
            )
            messages.success(request, f'Sesión de {cita.paciente.nombre} cerrada. Solicitud de siguiente cita enviada a Recepción.')
        else:
            messages.success(request, f'Sesión de {cita.paciente.nombre} registrada correctamente.')
    else:
        messages.error(request, 'Hubo un error al procesar el cierre de sesión. Intenta de nuevo.')

    return redirect('portal_terapeuta')


# =============================================================================
# SPRINT 3 — VISTAS DE RECEPCIÓN
# =============================================================================

@login_required
def bitacora_diaria(request):
    """
    Bitácora de citas de un día concreto para el staff.
    GET ?fecha=YYYY-MM-DD  → muestra ese día (default: hoy).
    Acceso exclusivo a is_superuser.
    """
    if not request.user.is_superuser:
        return redirect('home')

    from datetime import date as date_type
    fecha_str = request.GET.get('fecha', '')
    try:
        fecha = date_type.fromisoformat(fecha_str) if fecha_str else date_type.today()
    except ValueError:
        fecha = date_type.today()

    citas_qs = (
        Cita.objects
        .filter(fecha=fecha)
        .select_related('paciente', 'terapeuta', 'servicio', 'consultorio', 'division')
        .prefetch_related('pacientes_adicionales')
        .order_by('hora', 'terapeuta__nombre')
    )

    # Evaluar queryset una sola vez para cómputos Python eficientes
    citas = list(citas_qs)

    # Indicador "Solicitó Seguimiento": SolicitudCita creada ese día para ese par (terapeuta, paciente)
    sol_set = set(
        SolicitudCita.objects
        .filter(fecha_creacion__date=fecha)
        .values_list('terapeuta_id', 'paciente_nombre')
    )
    for cita in citas:
        key = (cita.terapeuta_id, cita.paciente.nombre if cita.paciente else '')
        cita.solicito_seguimiento = key in sol_set

    # Stats del día
    total         = len(citas)
    asistieron    = sum(1 for c in citas if c.estatus == Cita.ESTATUS_SI_ASISTIO)
    no_asistieron = sum(1 for c in citas if c.estatus in (Cita.ESTATUS_NO_ASISTIO, Cita.ESTATUS_CANCELO))
    por_cerrar    = sum(1 for c in citas if c.es_finalizable)
    monto_dia     = sum(
        (c.costo or Decimal('0')) for c in citas if c.estatus == Cita.ESTATUS_SI_ASISTIO
    )

    solicitudes_pendientes_count = SolicitudCita.objects.filter(estado='pendiente').count()

    # ── Exportar Excel ────────────────────────────────────────────────────────
    if request.GET.get('export') == 'xlsx':
        bita_headers = [
            'Hora', 'Consultorio', 'Paciente', 'Terapeuta',
            'Servicio', 'Estatus', 'Método de Pago', 'Costo ($)', 'Solicitó Seguimiento',
        ]
        bita_rows = [
            [
                c.hora.strftime('%H:%M'),
                str(c.consultorio) if c.consultorio else '—',
                c.paciente.nombre if c.paciente else '—',
                str(c.terapeuta) if c.terapeuta else '—',
                str(c.servicio) if c.servicio else '—',
                c.get_estatus_display(),
                c.metodo_pago or '—',
                float(c.costo) if c.costo is not None else 0,
                'Sí' if c.solicito_seguimiento else 'No',
            ]
            for c in citas
        ]
        return _build_excel_response(
            filename=f"bitacora_{fecha.isoformat()}.xlsx",
            sheet_title=f"Bitácora {fecha.isoformat()}",
            headers=bita_headers,
            rows=bita_rows,
            col_widths=[8, 18, 24, 22, 22, 14, 16, 10, 18],
        )
    # ─────────────────────────────────────────────────────────────────────────

    meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    fecha_bonita = f"{fecha.day} de {meses[fecha.month - 1]} de {fecha.year}"

    return render(request, 'clinica/bitacora_diaria.html', {
        'citas':                        citas,
        'fecha':                        fecha,
        'fecha_iso':                    fecha.isoformat(),
        'fecha_bonita':                 fecha_bonita,
        'total':                        total,
        'asistieron':                   asistieron,
        'no_asistieron':                no_asistieron,
        'por_cerrar':                   por_cerrar,
        'monto_dia':                    monto_dia,
        'solicitudes_pendientes_count': solicitudes_pendientes_count,
    })


@login_required
def reporte_general(request):
    """
    Reporte histórico de citas con filtros por rango de fechas y terapeuta.
    GET ?fecha_inicio=&fecha_fin=&terapeuta_id=&export=csv
    Acceso exclusivo a is_superuser.
    """
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('home')

    from datetime import date as date_type
    hoy = date_type.today()

    # Leer filtros (default: mes actual)
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str    = request.GET.get('fecha_fin', '')
    terapeuta_id     = request.GET.get('terapeuta_id', '')

    try:
        fecha_inicio = date_type.fromisoformat(fecha_inicio_str) if fecha_inicio_str else hoy.replace(day=1)
    except ValueError:
        fecha_inicio = hoy.replace(day=1)

    try:
        fecha_fin = date_type.fromisoformat(fecha_fin_str) if fecha_fin_str else hoy
    except ValueError:
        fecha_fin = hoy

    citas = (
        Cita.objects
        .filter(fecha__range=(fecha_inicio, fecha_fin))
        .select_related('paciente', 'terapeuta', 'servicio', 'consultorio', 'division')
        .order_by('-fecha', 'hora')
    )
    if terapeuta_id:
        citas = citas.filter(terapeuta_id=terapeuta_id)

    # ── Exportar Excel / CSV ──────────────────────────────────────────────────
    export = request.GET.get('export', '')
    if export in ('xlsx', 'csv'):
        headers = [
            'Fecha', 'Hora', 'Tipo', 'Paciente', 'Terapeuta',
            'Servicio', 'División', 'Consultorio',
            'Estatus', 'Método de Pago', 'Costo ($)',
        ]
        rows = [
            [
                c.fecha.strftime('%d/%m/%Y'),
                c.hora.strftime('%H:%M'),
                c.get_tipo_paciente_display(),
                c.paciente.nombre if c.paciente else '—',
                str(c.terapeuta) if c.terapeuta else '—',
                str(c.servicio)  if c.servicio  else '—',
                str(c.division)  if c.division  else '—',
                str(c.consultorio) if c.consultorio else '—',
                c.get_estatus_display(),
                c.metodo_pago or '—',
                float(c.costo) if c.costo is not None else 0,
            ]
            for c in citas
        ]
        if export == 'xlsx':
            return _build_excel_response(
                filename=f"reporte_intra_{fecha_inicio}_{fecha_fin}.xlsx",
                sheet_title="Reporte General",
                headers=headers,
                rows=rows,
                col_widths=[14, 8, 8, 24, 22, 22, 14, 18, 16, 16, 10],
            )
        # CSV fallback
        nombre = f"reporte_intra_{fecha_inicio}_{fecha_fin}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{nombre}"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    # ── Stats del rango ───────────────────────────────────────────────────────
    total      = citas.count()
    asistieron = citas.filter(estatus=Cita.ESTATUS_SI_ASISTIO).count()
    monto_total = (
        citas.filter(estatus=Cita.ESTATUS_SI_ASISTIO)
             .aggregate(t=Sum('costo'))['t'] or Decimal('0')
    )

    terapeutas = Terapeuta.objects.filter(activo=True).order_by('nombre')

    return render(request, 'clinica/reporte_general.html', {
        'citas':            citas,
        'fecha_inicio':     fecha_inicio,
        'fecha_fin':        fecha_fin,
        'fecha_inicio_iso': fecha_inicio.isoformat(),
        'fecha_fin_iso':    fecha_fin.isoformat(),
        'terapeuta_id_sel': terapeuta_id,
        'terapeutas':       terapeutas,
        'total':            total,
        'asistieron':       asistieron,
        'monto_total':      monto_total,
    })


@login_required
def estadisticas_ausentismo(request):
    """
    Ranking de pacientes con más cancelaciones o inasistencias.
    Acceso exclusivo a is_superuser.
    """
    if not request.user.is_superuser:
        return redirect('home')

    hoy = date.today()

    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str    = request.GET.get('fecha_fin', '')

    try:
        fecha_inicio = date.fromisoformat(fecha_inicio_str) if fecha_inicio_str else None
    except ValueError:
        fecha_inicio = None

    try:
        fecha_fin = date.fromisoformat(fecha_fin_str) if fecha_fin_str else hoy
    except ValueError:
        fecha_fin = hoy

    citas_qs = Cita.objects.filter(paciente__isnull=False, fecha__lte=fecha_fin)
    if fecha_inicio:
        citas_qs = citas_qs.filter(fecha__gte=fecha_inicio)

    pacientes_stats = (
        citas_qs
        .values('paciente__id', 'paciente__nombre')
        .annotate(
            cancelaciones=Count('id', filter=Q(estatus=Cita.ESTATUS_CANCELO)),
            inasistencias=Count('id', filter=Q(estatus=Cita.ESTATUS_NO_ASISTIO)),
            total_ausencias=Count('id', filter=Q(estatus__in=[
                Cita.ESTATUS_CANCELO, Cita.ESTATUS_NO_ASISTIO
            ])),
            total_citas=Count('id'),
        )
        .filter(total_ausencias__gt=0)
        .order_by('-total_ausencias', '-cancelaciones')
    )

    return render(request, 'clinica/estadisticas_ausentismo.html', {
        'pacientes_stats':  pacientes_stats,
        'fecha_inicio':     fecha_inicio,
        'fecha_fin':        fecha_fin,
        'fecha_inicio_iso': fecha_inicio.isoformat() if fecha_inicio else '',
        'fecha_fin_iso':    fecha_fin.isoformat(),
        'total_pacientes':  pacientes_stats.count(),
    })


# =============================================================================
# SPRINT 4 — NÓMINA SEMANAL
# =============================================================================

def _semana_actual():
    """Devuelve (viernes, jueves) del periodo de corte en curso (Vie→Jue)."""
    hoy = date.today()
    # (weekday - 4) % 7: días transcurridos desde el viernes más reciente
    # Lun=0→3, Mar=1→4, Mié=2→5, Jue=3→6, Vie=4→0, Sáb=5→1, Dom=6→2
    viernes = hoy - timedelta(days=(hoy.weekday() - 4) % 7)
    return viernes, viernes + timedelta(days=6)


def _parse_fechas_semana(request):
    """Extrae fecha_inicio y fecha_fin de los GET params, con fallback a semana actual."""
    lunes, domingo = _semana_actual()
    try:
        fi = date.fromisoformat(request.GET.get('fecha_inicio', ''))
    except ValueError:
        fi = lunes
    try:
        ff = date.fromisoformat(request.GET.get('fecha_fin', ''))
    except ValueError:
        ff = domingo
    return fi, ff


@login_required
def nomina_lista(request):
    """
    Panel de nómina semanal: muestra todos los terapeutas activos con
    sus sesiones, ingreso clínica, pago calculado y estado del corte.
    """
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('home')

    fecha_inicio, fecha_fin = _parse_fechas_semana(request)
    prev_inicio = fecha_inicio - timedelta(days=7)
    next_inicio = fecha_inicio + timedelta(days=7)

    terapeutas = Terapeuta.objects.filter(activo=True).order_by('nombre')

    # Cortes existentes para esta semana (una sola consulta)
    cortes_map = {
        c.terapeuta_id: c
        for c in CorteSemanal.objects.filter(
            fecha_inicio=fecha_inicio,
            terapeuta__in=terapeutas,
        )
    }

    # Stats de citas por terapeuta (una sola consulta agregada)
    stats_map = {
        row['terapeuta_id']: row
        for row in Cita.objects.filter(
            fecha__range=(fecha_inicio, fecha_fin),
            estatus=Cita.ESTATUS_SI_ASISTIO,
            terapeuta__in=terapeutas,
        ).values('terapeuta_id').annotate(
            total_citas=Count('id'),
            ingreso_clinica=Sum('costo'),
        )
    }

    filas = []
    total_clinica_global = Decimal('0')
    total_terapeutas_global = Decimal('0')
    total_citas_global = 0

    for t in terapeutas:
        corte = cortes_map.get(t.id)
        stats = stats_map.get(t.id, {})
        citas_count   = stats.get('total_citas', 0)
        ingreso       = stats.get('ingreso_clinica') or Decimal('0')
        total_pago    = corte.total_pago if corte else None

        # Solo incluir terapeutas con actividad o con corte generado
        if not citas_count and not corte:
            continue

        total_clinica_global    += ingreso
        total_terapeutas_global += total_pago or Decimal('0')
        total_citas_global      += citas_count

        filas.append({
            'terapeuta':      t,
            'citas_count':    citas_count,
            'ingreso_clinica': ingreso,
            'total_pago':     total_pago,
            'corte':          corte,
        })

    meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    semana_label = (
        f"{fecha_inicio.day} de {meses[fecha_inicio.month-1]}"
        f" al {fecha_fin.day} de {meses[fecha_fin.month-1]} de {fecha_fin.year}"
    )

    return render(request, 'clinica/nomina_lista.html', {
        'filas':                    filas,
        'fecha_inicio':             fecha_inicio,
        'fecha_fin':                fecha_fin,
        'fecha_inicio_iso':         fecha_inicio.isoformat(),
        'fecha_fin_iso':            fecha_fin.isoformat(),
        'prev_inicio_iso':          prev_inicio.isoformat(),
        'prev_fin_iso':             (prev_inicio + timedelta(days=6)).isoformat(),
        'next_inicio_iso':          next_inicio.isoformat(),
        'next_fin_iso':             (next_inicio + timedelta(days=6)).isoformat(),
        'semana_label':             semana_label,
        'total_clinica_global':     total_clinica_global,
        'total_terapeutas_global':  total_terapeutas_global,
        'total_citas_global':       total_citas_global,
        'aprobados':  sum(1 for f in filas if f['corte'] and f['corte'].estatus == CorteSemanal.ESTATUS_APROBADO),
        'borradores': sum(1 for f in filas if f['corte'] and f['corte'].estatus == CorteSemanal.ESTATUS_BORRADOR),
        'confirmaciones_aceptadas': sum(1 for f in filas if f['corte'] and f['corte'].confirmacion_terapeuta == CorteSemanal.CONFIRMACION_ACEPTADO),
        'confirmaciones_incidencia': sum(1 for f in filas if f['corte'] and f['corte'].confirmacion_terapeuta == CorteSemanal.CONFIRMACION_INCIDENCIA),
        'hoy': date.today(),
    })


@login_required
def nomina_exportar_reporte_general(request):
    """
    Genera un Excel en formato "Reporte General" para un rango de fechas.
    GET ?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD
    Acceso exclusivo a superusuarios y staff.
    """
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('home')

    from datetime import date as date_type
    hoy = date_type.today()

    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str    = request.GET.get('fecha_fin', '')

    try:
        fecha_inicio = date_type.fromisoformat(fecha_inicio_str) if fecha_inicio_str else hoy.replace(day=1)
    except ValueError:
        fecha_inicio = hoy.replace(day=1)

    try:
        fecha_fin = date_type.fromisoformat(fecha_fin_str) if fecha_fin_str else hoy
    except ValueError:
        fecha_fin = hoy

    meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

    citas = (
        Cita.objects
        .filter(
            fecha__range=(fecha_inicio, fecha_fin),
            estatus__in=(Cita.ESTATUS_SI_ASISTIO, Cita.ESTATUS_NO_ASISTIO),
        )
        .select_related('paciente', 'terapeuta', 'servicio', 'consultorio', 'division')
        .order_by('fecha', 'hora')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    _ALT2_FILL  = PatternFill("solid", fgColor="F0F4F8")

    # Fila 1: encabezados
    headers = [
        'Dia', 'Periodo', 'Hora', 'División', 'Paciente', 'Sexo', 'Servicio',
        'Terapeuta', 'Consultorio', 'Pago', 'Método de pago', 'Tarjeta',
        'Folio fiscal', 'Notas', 'Fecha de pago',
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font   = _HEADER_FONT
        cell.fill   = _TEAL_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Datos a partir de fila 2
    for row_idx, c in enumerate(citas, start=2):
        fill = _ALT2_FILL if row_idx % 2 == 0 else None
        periodo = f"{meses[c.fecha.month - 1]} {c.fecha.year}"
        valores = [
            c.fecha.day,
            periodo,
            c.hora.strftime('%H:%M') if c.hora else '',
            str(c.division)    if c.division    else '',
            c.paciente.nombre  if c.paciente    else '',
            c.paciente.sexo    if c.paciente    else '',
            str(c.servicio)    if c.servicio    else '',
            str(c.terapeuta)   if c.terapeuta   else '',
            str(c.consultorio) if c.consultorio else '',
            float(c.costo)     if c.costo is not None else 0,
            c.metodo_pago      if c.metodo_pago else '',
            '',                                              # Tarjeta (no en modelo)
            c.folio_fiscal     if c.folio_fiscal else '',
            c.notas            if c.notas        else '',
            c.fecha,                                         # Fecha de pago = fecha cita
        ]
        for col_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _BODY_FONT
            if fill:
                cell.fill = fill

    # Formatear columna de Fecha de pago como fecha
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=15).number_format = 'DD/MM/YYYY'

    # Anchos de columna
    col_widths = [6, 14, 7, 16, 28, 10, 24, 24, 14, 10, 16, 16, 20, 30, 14]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"reporte_general_{fecha_inicio.isoformat()}_{fecha_fin.isoformat()}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def nomina_exportar_dispersion(request):
    """
    Genera el "Formato de Relación Horas Pendientes de Pago" (dispersión) en Excel.
    GET ?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD
    Acceso exclusivo a superusuarios y staff.
    """
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('home')

    from datetime import date as date_type
    hoy = date_type.today()

    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str    = request.GET.get('fecha_fin', '')

    try:
        fecha_inicio = date_type.fromisoformat(fecha_inicio_str) if fecha_inicio_str else hoy.replace(day=1)
    except ValueError:
        fecha_inicio = hoy.replace(day=1)
    try:
        fecha_fin = date_type.fromisoformat(fecha_fin_str) if fecha_fin_str else hoy
    except ValueError:
        fecha_fin = hoy

    from .models import Terapeuta as TerapeutaModel, ReglaTerapeuta

    meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

    # ── Helpers de cálculo inline ─────────────────────────────────────────────
    def _monto_sesion(cita, regla):
        tiene_adicionales = cita.pacientes_adicionales.exists()
        if tiene_adicionales and regla.pago_pareja is not None:
            return regla.pago_pareja
        if regla.pago_individual is not None:
            return regla.pago_individual
        if regla.pago_por_sesion is not None:
            return regla.pago_por_sesion
        if regla.tabulador_base and regla.tabulador_base.pago_base is not None:
            return regla.tabulador_base.pago_base
        return Decimal('0.00')

    def _gasolina(total_sesiones, regla):
        total = Decimal('0.00')
        if regla.bono_umbral_monto and regla.bono_umbral_pacientes:
            veces = total_sesiones // regla.bono_umbral_pacientes
            if veces > 0:
                total += regla.bono_umbral_monto * veces
        if regla.bono_por_paciente:
            total += regla.bono_por_paciente * total_sesiones
        if not regla.bono_umbral_monto and regla.tabulador_base:
            tab = regla.tabulador_base
            if tab.bono_monto and tab.bono_umbral_pacientes:
                veces = total_sesiones // tab.bono_umbral_pacientes
                if veces > 0:
                    total += tab.bono_monto * veces
        return total

    def _info_vales(regla):
        if regla.bono_umbral_monto and regla.bono_umbral_pacientes:
            return f"${int(regla.bono_umbral_monto)} / {regla.bono_umbral_pacientes}px"
        if regla.tabulador_base and regla.tabulador_base.bono_monto and regla.tabulador_base.bono_umbral_pacientes:
            return f"${int(regla.tabulador_base.bono_monto)} / {regla.tabulador_base.bono_umbral_pacientes}px"
        return "NA"

    # ── Construir filas agrupadas por terapeuta ───────────────────────────────
    all_rows = []
    grand_total    = Decimal('0.00')
    grand_gasolina = Decimal('0.00')

    terapeutas = TerapeutaModel.objects.filter(activo=True).order_by('nombre')

    for terapeuta in terapeutas:
        try:
            regla = terapeuta.regla_pago
        except ReglaTerapeuta.DoesNotExist:
            continue

        citas = (
            Cita.objects
            .filter(
                terapeuta=terapeuta,
                fecha__range=(fecha_inicio, fecha_fin),
                estatus=Cita.ESTATUS_SI_ASISTIO,
            )
            .select_related('paciente', 'servicio', 'consultorio')
            .prefetch_related('pacientes_adicionales')
            .order_by('fecha', 'hora')
        )

        lista = list(citas)
        if not lista:
            continue

        # Montos desde LineaNomina si existen (respeta ediciones manuales)
        lineas_map = {
            l.cita_id: l.monto
            for l in LineaNomina.objects.filter(
                tipo=LineaNomina.TIPO_SESION,
                cita__in=lista,
            )
        }

        # Gasolina desde cortes existentes en el rango; si no, calcular
        cortes_gas = CorteSemanal.objects.filter(
            terapeuta=terapeuta,
            fecha_inicio__lte=fecha_fin,
            fecha_fin__gte=fecha_inicio,
        ).aggregate(t=Sum('total_bonos'))['t'] or Decimal('0.00')
        gas = cortes_gas if lineas_map else _gasolina(len(lista), regla)

        info = _info_vales(regla)
        subtotal        = Decimal('0.00')
        filas_terapeuta = []

        for cita in lista:
            # Usar monto guardado si existe; si no, calcular desde tabulador
            monto = lineas_map.get(cita.id) or _monto_sesion(cita, regla)
            subtotal += monto
            periodo = f"{meses[cita.fecha.month - 1]} {cita.fecha.year}"
            filas_terapeuta.append({
                'dia':        cita.fecha.day,
                'periodo':    periodo,
                'hora':       cita.hora.strftime('%H:%M') if cita.hora else '',
                'terapeuta':  terapeuta.nombre,
                'paciente':   cita.paciente.nombre if cita.paciente else '',
                'info_vales': info,
                'monto':      float(monto),
                'total':      None,
                'gasolina':   None,
                'metodo':     '',
            })

        filas_terapeuta[-1]['total']    = float(subtotal)
        filas_terapeuta[-1]['gasolina'] = float(gas) if gas else None

        grand_total    += subtotal
        grand_gasolina += gas
        all_rows.extend(filas_terapeuta)

    # ── Construir Excel ───────────────────────────────────────────────────────
    _GREEN_FILL  = PatternFill("solid", fgColor="00B050")
    _GREEN_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    _BOLD_FONT   = Font(bold=True, name="Calibri", size=11)
    _TITLE_FONT  = Font(bold=True, underline="single", name="Calibri", size=13)
    _ALT3_FILL   = PatternFill("solid", fgColor="F0F4F8")
    _BORDER_THIN = openpyxl.styles.borders.Border(
        bottom=openpyxl.styles.borders.Side(style='thin', color='CCCCCC')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dispersión"

    COLS = 10  # A–J

    # Fila 1: Título
    ws.merge_cells(f'A1:{get_column_letter(COLS)}1')
    title_cell = ws['A1']
    title_cell.value     = "FORMATO DE RELACION HORAS PENDIENTES DE PAGO"
    title_cell.font      = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Fila 2: Info periodo + totales
    ws['A2'] = "Inicio del periodo:"
    ws['A2'].font = Font(bold=True, name="Calibri", size=11)
    ws['C2'] = fecha_inicio.strftime('%d/%m/%Y')
    ws['C2'].font = Font(bold=True, color="0070C0", name="Calibri", size=11)

    total_cell = ws.cell(row=2, column=8, value=f"${grand_total:,.2f}")
    total_cell.font  = _GREEN_FONT
    total_cell.fill  = _GREEN_FILL
    total_cell.alignment = Alignment(horizontal="center")

    gas_cell = ws.cell(row=2, column=9, value=f"${grand_gasolina:,.2f}" if grand_gasolina else "")
    gas_cell.font  = _GREEN_FONT
    gas_cell.fill  = _GREEN_FILL
    gas_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Fila 3: Encabezados de columna
    col_headers = ['Dia', 'Periodo', 'Hora', 'Terapeuta', 'Paciente',
                   'Info Vales', 'Monto', 'Total', 'Gasolina', 'Metodo']
    _SUBHEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
    for col_idx, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font      = Font(bold=True, name="Calibri", size=10)
        cell.fill      = _SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # Filas de datos (desde fila 4)
    for row_idx, r in enumerate(all_rows, start=4):
        fill = _ALT3_FILL if row_idx % 2 == 0 else None
        valores = [
            r['dia'], r['periodo'], r['hora'], r['terapeuta'],
            r['paciente'], r['info_vales'], r['monto'],
            r['total'], r['gasolina'], r['metodo'],
        ]
        for col_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _BODY_FONT
            if fill:
                cell.fill = fill
            # Total y Gasolina en negrita
            if col_idx in (8, 9) and val is not None:
                cell.font = Font(bold=True, name="Calibri", size=10)
            # Formatear montos como moneda
            if col_idx in (7, 8, 9) and isinstance(val, float):
                cell.number_format = '"$"#,##0.00'

    # Anchos de columna
    col_widths = [5, 13, 7, 22, 32, 13, 10, 12, 11, 10]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"dispersion_{fecha_inicio.isoformat()}_{fecha_fin.isoformat()}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def nomina_detalle(request, terapeuta_id):
    """
    Desglose de nómina de un terapeuta para una semana.
    Si el CorteSemanal aún no existe, muestra un preview sin persistir.
    Si existe en borrador, permite agregar BonoExtra y aprobarlo.
    """
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('home')

    terapeuta    = get_object_or_404(Terapeuta, id=terapeuta_id)
    fecha_inicio, fecha_fin = _parse_fechas_semana(request)

    try:
        corte = CorteSemanal.objects.get(terapeuta=terapeuta, fecha_inicio=fecha_inicio)
    except CorteSemanal.DoesNotExist:
        corte = None

    error_preview = None

    if corte:
        lineas_sesion = list(
            corte.lineas.filter(tipo=LineaNomina.TIPO_SESION)
                        .select_related('cita__paciente', 'cita__servicio')
                        .order_by('cita__fecha', 'cita__hora')
        )
        lineas_bono = list(
            corte.lineas.filter(tipo__in=[LineaNomina.TIPO_BONO_UMBRAL, LineaNomina.TIPO_BONO_POR_PACIENTE])
        )
        lineas_penalizacion = list(
            corte.lineas.filter(tipo=LineaNomina.TIPO_PENALIZACION)
                        .select_related('cita__paciente', 'cita__servicio')
                        .order_by('cita__fecha')
        )
        bonos_extra = list(corte.bonos_extra.select_related('registrado_por').order_by('creado_en'))
        subtotal    = corte.subtotal_sesiones
        total_bonos = corte.total_bonos
        total_pago  = corte.total_pago
    else:
        preview = preview_nomina_semanal(terapeuta, fecha_inicio, fecha_fin)
        if 'error' in preview:
            error_preview = preview['error']
            lineas_sesion = lineas_bono = lineas_penalizacion = bonos_extra = []
            subtotal = total_bonos = total_pago = Decimal('0')
        else:
            lineas_sesion       = [l for l in preview['lineas'] if l['tipo'] == LineaNomina.TIPO_SESION]
            lineas_bono         = [l for l in preview['lineas'] if l['tipo'] in (LineaNomina.TIPO_BONO_UMBRAL, LineaNomina.TIPO_BONO_POR_PACIENTE)]
            lineas_penalizacion = []
            bonos_extra         = []
            subtotal      = preview['subtotal_sesiones']
            total_bonos   = preview['total_bonos']
            total_pago    = preview['total_pago']

    puede_editar  = corte and corte.estatus == CorteSemanal.ESTATUS_BORRADOR
    puede_aprobar = puede_editar

    meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    semana_label = (
        f"{fecha_inicio.day} de {meses[fecha_inicio.month-1]}"
        f" al {fecha_fin.day} de {meses[fecha_fin.month-1]} de {fecha_fin.year}"
    )

    # ── Exportar nómina Excel ─────────────────────────────────────────────────
    if request.GET.get('export') == 'xlsx' and not error_preview:
        nom_headers = ['Tipo', 'Concepto', 'Fecha', 'Hora', 'Paciente', 'Monto ($)']
        nom_rows = []
        for l in lineas_sesion:
            if isinstance(l, dict):
                fecha_c = l['cita'].fecha.strftime('%d/%m/%Y') if l.get('cita') and l['cita'].fecha else '—'
                hora_c  = l['cita'].hora.strftime('%H:%M') if l.get('cita') and l['cita'].hora else '—'
                pac     = l['cita'].paciente.nombre if l.get('cita') and l['cita'].paciente else '—'
            else:
                fecha_c = l.cita.fecha.strftime('%d/%m/%Y') if l.cita and l.cita.fecha else '—'
                hora_c  = l.cita.hora.strftime('%H:%M') if l.cita and l.cita.hora else '—'
                pac     = l.cita.paciente.nombre if l.cita and l.cita.paciente else '—'
            monto = float(l['monto'] if isinstance(l, dict) else l.monto)
            nom_rows.append(['Sesión', l['concepto'] if isinstance(l, dict) else l.concepto, fecha_c, hora_c, pac, monto])
        for l in lineas_bono:
            monto = float(l['monto'] if isinstance(l, dict) else l.monto)
            nom_rows.append(['Bono automático', l['concepto'] if isinstance(l, dict) else l.concepto, '—', '—', '—', monto])
        for b in bonos_extra:
            nom_rows.append(['Bono extra', b.concepto, '—', '—', '—', float(b.monto)])
        return _build_excel_response(
            filename=f"nomina_{terapeuta.nombre.replace(' ', '_')}_{fecha_inicio.isoformat()}.xlsx",
            sheet_title="Nómina Detalle",
            headers=nom_headers,
            rows=nom_rows,
            col_widths=[16, 35, 12, 8, 24, 12],
        )
    # ─────────────────────────────────────────────────────────────────────────

    return render(request, 'clinica/nomina_detalle.html', {
        'terapeuta':           terapeuta,
        'corte':               corte,
        'lineas_sesion':       lineas_sesion,
        'lineas_bono':         lineas_bono,
        'lineas_penalizacion': lineas_penalizacion,
        'bonos_extra':         bonos_extra,
        'subtotal':            subtotal,
        'total_bonos':         total_bonos,
        'total_pago':          total_pago,
        'fecha_inicio':        fecha_inicio,
        'fecha_fin':           fecha_fin,
        'fecha_inicio_iso':    fecha_inicio.isoformat(),
        'fecha_fin_iso':       fecha_fin.isoformat(),
        'semana_label':        semana_label,
        'puede_editar':        puede_editar,
        'puede_aprobar':       puede_aprobar,
        'error_preview':       error_preview,
    })


@login_required
def confirmar_nomina_terapeuta(request, corte_id):
    """El terapeuta revisa y confirma (o reporta incidencia en) su nómina semanal."""
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    terapeuta = request.user.perfil_terapeuta
    corte = get_object_or_404(CorteSemanal, id=corte_id, terapeuta=terapeuta)

    if corte.confirmacion_terapeuta is not None:
        messages.warning(request, 'Esta nómina ya fue respondida anteriormente.')
        return redirect('portal_terapeuta')

    if request.method == 'POST':
        accion = request.POST.get('accion', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()

        if accion == 'confirmo':
            corte.confirmacion_terapeuta = CorteSemanal.CONFIRMACION_ACEPTADO
            corte.confirmacion_terapeuta_en = timezone.now()
            corte.save(update_fields=['confirmacion_terapeuta', 'confirmacion_terapeuta_en'])
            messages.success(request, 'Nómina confirmada correctamente. ¡Gracias!')
            return redirect('portal_terapeuta')

        elif accion == 'algo_mal':
            if not descripcion:
                messages.error(request, 'Por favor describe el problema antes de enviar.')
                return redirect('confirmar_nomina_terapeuta', corte_id=corte_id)

            meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
            semana_label = (
                f"{corte.fecha_inicio.day} de {meses[corte.fecha_inicio.month - 1]}"
                f" al {corte.fecha_fin.day} de {meses[corte.fecha_fin.month - 1]}"
                f" de {corte.fecha_fin.year}"
            )
            ReporteIncidente.objects.create(
                terapeuta=terapeuta,
                tipo=ReporteIncidente.TIPO_QUEJA,
                titulo=f"Incidencia en nómina: {semana_label}",
                descripcion=descripcion,
            )
            corte.confirmacion_terapeuta = CorteSemanal.CONFIRMACION_INCIDENCIA
            corte.confirmacion_terapeuta_en = timezone.now()
            corte.save(update_fields=['confirmacion_terapeuta', 'confirmacion_terapeuta_en'])
            messages.success(request, 'Tu reporte fue enviado. El equipo revisará tu nómina a la brevedad.')
            return redirect('portal_terapeuta')

        messages.error(request, 'Acción no válida.')
        return redirect('confirmar_nomina_terapeuta', corte_id=corte_id)

    # GET — mostrar resumen de nómina para revisión
    lineas_sesion = list(
        corte.lineas.filter(tipo=LineaNomina.TIPO_SESION)
                    .select_related('cita__paciente', 'cita__servicio')
                    .order_by('cita__fecha', 'cita__hora')
    )
    lineas_bono = list(
        corte.lineas.filter(
            tipo__in=[LineaNomina.TIPO_BONO_UMBRAL, LineaNomina.TIPO_BONO_POR_PACIENTE]
        )
    )
    bonos_extra = list(corte.bonos_extra.order_by('creado_en'))

    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    semana_label = (
        f"{corte.fecha_inicio.day} de {meses[corte.fecha_inicio.month - 1]}"
        f" al {corte.fecha_fin.day} de {meses[corte.fecha_fin.month - 1]}"
        f" de {corte.fecha_fin.year}"
    )

    return render(request, 'clinica/confirmar_nomina_terapeuta.html', {
        'corte': corte,
        'lineas_sesion': lineas_sesion,
        'lineas_bono': lineas_bono,
        'bonos_extra': bonos_extra,
        'semana_label': semana_label,
    })


@login_required
def nomina_calcular(request, terapeuta_id):
    """POST: genera o recalcula el CorteSemanal en borrador."""
    if not request.user.is_superuser or request.method != 'POST':
        return redirect('nomina_lista')

    terapeuta = get_object_or_404(Terapeuta, id=terapeuta_id)
    fi_str = request.POST.get('fecha_inicio', '')
    ff_str = request.POST.get('fecha_fin', '')

    try:
        fi = date.fromisoformat(fi_str)
        ff = date.fromisoformat(ff_str)
    except ValueError:
        messages.error(request, 'Fechas inválidas.')
        return redirect('nomina_lista')

    try:
        corte = calcular_nomina_semanal(terapeuta, fi, ff)
        messages.success(
            request,
            f'Corte de {terapeuta.nombre} generado. '
            f'Sesiones: {corte.total_sesiones} | Total: ${corte.total_pago:,.2f}'
        )
    except ValueError as e:
        messages.error(request, str(e))

    url = reverse('nomina_detalle', args=[terapeuta_id])
    return redirect(f'{url}?fecha_inicio={fi_str}&fecha_fin={ff_str}')


@login_required
def nomina_aprobar(request, corte_id):
    """POST: aprueba y sella un CorteSemanal."""
    if not request.user.is_superuser or request.method != 'POST':
        return redirect('nomina_lista')

    corte = get_object_or_404(CorteSemanal, id=corte_id)
    fi_str = corte.fecha_inicio.isoformat()
    ff_str = corte.fecha_fin.isoformat()

    try:
        aprobar_corte_semanal(corte, request.user)
        _registrar_actividad(request, 'nomina_aprobada', 'nomina',
            f'Nómina de {corte.terapeuta.nombre} del {corte.fecha_inicio.strftime("%d/%m/%Y")} '
            f'al {corte.fecha_fin.strftime("%d/%m/%Y")} aprobada. Total: ${corte.total_pago:,.2f}.',
            terapeuta=corte.terapeuta)
        messages.success(
            request,
            f'Nómina de {corte.terapeuta.nombre} aprobada y sellada. '
            f'Total: ${corte.total_pago:,.2f}'
        )
    except ValueError as e:
        messages.error(request, str(e))

    url = reverse('nomina_detalle', args=[corte.terapeuta_id])
    return redirect(f'{url}?fecha_inicio={fi_str}&fecha_fin={ff_str}')


@login_required
def nomina_agregar_bono(request, corte_id):
    """POST: agrega un BonoExtra a un CorteSemanal en borrador y recalcula totales."""
    if not request.user.is_superuser or request.method != 'POST':
        return redirect('nomina_lista')

    corte = get_object_or_404(CorteSemanal, id=corte_id)
    fi_str = corte.fecha_inicio.isoformat()
    ff_str = corte.fecha_fin.isoformat()

    if corte.estatus != CorteSemanal.ESTATUS_BORRADOR:
        messages.error(request, 'Solo se pueden agregar bonos a cortes en borrador.')
        url = reverse('nomina_detalle', args=[corte.terapeuta_id])
        return redirect(f'{url}?fecha_inicio={fi_str}&fecha_fin={ff_str}')

    concepto = request.POST.get('concepto', '').strip()
    monto_str = request.POST.get('monto', '').strip()

    try:
        monto = Decimal(monto_str)
        if monto <= 0:
            raise ValueError('El monto debe ser mayor a cero.')
        if not concepto:
            raise ValueError('El concepto es obligatorio.')
    except Exception as e:
        messages.error(request, f'Error al agregar bono: {e}')
    else:
        BonoExtra.objects.create(
            corte=corte,
            concepto=concepto,
            monto=monto,
            registrado_por=request.user,
        )
        # Recalcular para que los totales del corte reflejen el nuevo BonoExtra
        try:
            calcular_nomina_semanal(corte.terapeuta, corte.fecha_inicio, corte.fecha_fin)
            messages.success(request, f'Bono "{concepto}" de ${monto:,.2f} agregado correctamente.')
        except ValueError as e:
            messages.warning(request, f'Bono guardado pero no se pudo recalcular el total: {e}')

    url = reverse('nomina_detalle', args=[corte.terapeuta_id])
    return redirect(f'{url}?fecha_inicio={fi_str}&fecha_fin={ff_str}')


# =============================================================================
# NÓMINA — Vista consolidada (todos los terapeutas)
# =============================================================================

@login_required
def nomina_todos_detalles(request):
    """
    Vista consolidada: muestra el desglose de sesiones de TODOS los terapeutas
    con actividad en la semana seleccionada.
    Las líneas de cortes en borrador tienen un botón de edición.
    """
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('home')

    fecha_inicio, fecha_fin = _parse_fechas_semana(request)

    meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    semana_label = (
        f"{fecha_inicio.day} de {meses[fecha_inicio.month-1]}"
        f" al {fecha_fin.day} de {meses[fecha_fin.month-1]} de {fecha_fin.year}"
    )

    terapeutas = Terapeuta.objects.filter(activo=True).order_by('nombre')

    cortes_map = {
        c.terapeuta_id: c
        for c in CorteSemanal.objects.filter(
            fecha_inicio=fecha_inicio,
            terapeuta__in=terapeutas,
        ).prefetch_related('lineas__cita__paciente', 'lineas__cita__servicio', 'bonos_extra')
    }

    bloques = []
    total_global = Decimal('0')

    for t in terapeutas:
        corte = cortes_map.get(t.id)

        if corte:
            lineas_sesion = list(
                corte.lineas.filter(tipo=LineaNomina.TIPO_SESION)
                            .select_related('cita__paciente', 'cita__servicio')
                            .order_by('cita__fecha', 'cita__hora')
            )
            lineas_bono  = list(corte.lineas.exclude(tipo=LineaNomina.TIPO_SESION))
            bonos_extra  = list(corte.bonos_extra.all())
            subtotal     = corte.subtotal_sesiones
            total_bonos  = corte.total_bonos
            total_pago   = corte.total_pago
            puede_editar = corte.estatus == CorteSemanal.ESTATUS_BORRADOR
        else:
            prev = preview_nomina_semanal(t, fecha_inicio, fecha_fin)
            if 'error' in prev or not prev.get('lineas'):
                continue
            lineas_sesion = [l for l in prev['lineas'] if l['tipo'] == LineaNomina.TIPO_SESION]
            lineas_bono   = [l for l in prev['lineas'] if l['tipo'] != LineaNomina.TIPO_SESION]
            bonos_extra   = []
            subtotal      = prev['subtotal_sesiones']
            total_bonos   = prev['total_bonos']
            total_pago    = prev['total_pago']
            puede_editar  = False

        if not lineas_sesion and not lineas_bono:
            continue

        total_global += total_pago or Decimal('0')

        bloques.append({
            'terapeuta':    t,
            'corte':        corte,
            'lineas_sesion': lineas_sesion,
            'lineas_bono':  lineas_bono,
            'bonos_extra':  bonos_extra,
            'subtotal':     subtotal,
            'total_bonos':  total_bonos,
            'total_pago':   total_pago,
            'puede_editar': puede_editar,
        })

    return render(request, 'clinica/nomina_todos_detalles.html', {
        'bloques':          bloques,
        'total_global':     total_global,
        'fecha_inicio':     fecha_inicio,
        'fecha_fin':        fecha_fin,
        'fecha_inicio_iso': fecha_inicio.isoformat(),
        'fecha_fin_iso':    fecha_fin.isoformat(),
        'semana_label':     semana_label,
    })


@login_required
def nomina_editar_linea(request, linea_id):
    """POST: edita el monto de una LineaNomina en borrador y recalcula el corte."""
    if not request.user.is_superuser or request.method != 'POST':
        return redirect('nomina_lista')

    linea = get_object_or_404(LineaNomina, id=linea_id)
    corte = linea.corte
    fi_str = request.POST.get('fecha_inicio', corte.fecha_inicio.isoformat())
    ff_str = request.POST.get('fecha_fin',    corte.fecha_fin.isoformat())

    if corte.estatus != CorteSemanal.ESTATUS_BORRADOR:
        messages.error(request, 'Solo se pueden editar líneas de cortes en borrador.')
        url = reverse('nomina_todos_detalles')
        return redirect(f'{url}?fecha_inicio={fi_str}&fecha_fin={ff_str}')

    from decimal import InvalidOperation
    monto_str = request.POST.get('monto', '').replace(',', '.')
    try:
        nuevo_monto = Decimal(monto_str)
        if nuevo_monto < 0:
            raise ValueError
    except (InvalidOperation, ValueError):
        messages.error(request, 'Monto inválido.')
        url = reverse('nomina_todos_detalles')
        return redirect(f'{url}?fecha_inicio={fi_str}&fecha_fin={ff_str}')

    linea.monto = nuevo_monto
    linea.save()

    # Recalcular snapshot del corte
    subtotal    = corte.lineas.filter(tipo=LineaNomina.TIPO_SESION).aggregate(t=Sum('monto'))['t'] or Decimal('0')
    bonos_auto  = corte.lineas.exclude(tipo=LineaNomina.TIPO_SESION).aggregate(t=Sum('monto'))['t'] or Decimal('0')
    bonos_extra = corte.bonos_extra.aggregate(t=Sum('monto'))['t'] or Decimal('0')
    corte.subtotal_sesiones = subtotal
    corte.total_bonos       = bonos_auto + bonos_extra
    corte.total_pago        = subtotal + bonos_auto + bonos_extra
    corte.save()

    messages.success(request, f'Monto actualizado a ${nuevo_monto:,.2f}.')
    next_view = request.POST.get('next_view', '')
    if next_view == 'detalle':
        url = reverse('nomina_detalle', args=[corte.terapeuta_id])
    else:
        url = reverse('nomina_todos_detalles')
    return redirect(f'{url}?fecha_inicio={fi_str}&fecha_fin={ff_str}')


@login_required
def nomina_editar_bono_extra(request, bono_id):
    """POST: edita concepto y monto de un BonoExtra y recalcula el corte."""
    if not request.user.is_superuser or request.method != 'POST':
        return redirect('nomina_lista')

    bono  = get_object_or_404(BonoExtra, id=bono_id)
    corte = bono.corte
    fi_str = request.POST.get('fecha_inicio', corte.fecha_inicio.isoformat())
    ff_str = request.POST.get('fecha_fin',    corte.fecha_fin.isoformat())

    if corte.estatus != CorteSemanal.ESTATUS_BORRADOR:
        messages.error(request, 'Solo se pueden editar bonos de cortes en borrador.')
        url = reverse('nomina_todos_detalles')
        return redirect(f'{url}?fecha_inicio={fi_str}&fecha_fin={ff_str}')

    from decimal import InvalidOperation
    concepto  = request.POST.get('concepto', '').strip()
    monto_str = request.POST.get('monto', '').replace(',', '.')
    try:
        nuevo_monto = Decimal(monto_str)
        if nuevo_monto < 0:
            raise ValueError
    except (InvalidOperation, ValueError):
        messages.error(request, 'Monto inválido.')
        url = reverse('nomina_todos_detalles')
        return redirect(f'{url}?fecha_inicio={fi_str}&fecha_fin={ff_str}')

    if concepto:
        bono.concepto = concepto
    bono.monto = nuevo_monto
    bono.save()

    # Recalcular snapshot del corte
    subtotal    = corte.lineas.filter(tipo=LineaNomina.TIPO_SESION).aggregate(t=Sum('monto'))['t'] or Decimal('0')
    bonos_auto  = corte.lineas.exclude(tipo=LineaNomina.TIPO_SESION).aggregate(t=Sum('monto'))['t'] or Decimal('0')
    bonos_extra = corte.bonos_extra.aggregate(t=Sum('monto'))['t'] or Decimal('0')
    corte.subtotal_sesiones = subtotal
    corte.total_bonos       = bonos_auto + bonos_extra
    corte.total_pago        = subtotal + bonos_auto + bonos_extra
    corte.save()

    messages.success(request, f'Bono actualizado a ${nuevo_monto:,.2f}.')
    next_view = request.POST.get('next_view', '')
    if next_view == 'detalle':
        url = reverse('nomina_detalle', args=[corte.terapeuta_id])
    else:
        url = reverse('nomina_todos_detalles')
    return redirect(f'{url}?fecha_inicio={fi_str}&fecha_fin={ff_str}')


# =============================================================================
# MÓDULO DE TABULADORES — Misión 2
# =============================================================================

@login_required
def tabuladores_config(request):
    """
    Vista principal de configuración de tabuladores.
    Muestra dos pestañas: Tabulador General (categorías) y Reglas Individuales.
    Acceso exclusivo a is_staff.
    """
    if not request.user.is_staff:
        return redirect('home')

    from .models import TabuladorGeneral, ReglaTerapeuta

    categorias = TabuladorGeneral.objects.order_by('numero')
    reglas = (
        ReglaTerapeuta.objects
        .select_related('terapeuta', 'tabulador_base')
        .order_by('terapeuta__nombre')
    )
    ids_con_regla = reglas.values_list('terapeuta_id', flat=True)
    sin_regla = Terapeuta.objects.filter(activo=True).exclude(id__in=ids_con_regla).order_by('nombre')
    return render(request, 'clinica/tabuladores_config.html', {
        'categorias': categorias,
        'reglas': reglas,
        'sin_regla': sin_regla,
    })


@login_required
def tabuladores_editar_categoria(request, categoria_id):
    """POST: actualiza un TabuladorGeneral desde el modal de edición."""
    if not request.user.is_staff or request.method != 'POST':
        return redirect('tabuladores_config')

    from .models import TabuladorGeneral
    from .forms import TabuladorGeneralForm

    categoria = get_object_or_404(TabuladorGeneral, id=categoria_id)
    form = TabuladorGeneralForm(request.POST, instance=categoria)
    if form.is_valid():
        form.save()
        messages.success(request, f'Categoría {categoria.numero} actualizada correctamente.')
    else:
        errores = '; '.join(
            f'{field}: {", ".join(errs)}'
            for field, errs in form.errors.items()
        )
        messages.error(request, f'Error al guardar: {errores}')
    return redirect('tabuladores_config')


@login_required
@login_required
def tabuladores_crear_regla(request, terapeuta_id):
    """POST: crea una ReglaTerapeuta nueva para un terapeuta que aún no tiene."""
    if not request.user.is_staff or request.method != 'POST':
        return redirect('tabuladores_config')

    from .models import ReglaTerapeuta
    from .forms import ReglaTerapeutaForm

    terapeuta = get_object_or_404(Terapeuta, id=terapeuta_id)
    if hasattr(terapeuta, 'regla_pago'):
        messages.error(request, f'{terapeuta.nombre} ya tiene una regla configurada.')
        return redirect('tabuladores_config')

    regla = ReglaTerapeuta(terapeuta=terapeuta)
    form = ReglaTerapeutaForm(request.POST, instance=regla)
    if form.is_valid():
        form.save()
        messages.success(request, f'Regla creada para {terapeuta.nombre}.')
    else:
        errores = '; '.join(
            f'{field}: {", ".join(errs)}'
            for field, errs in form.errors.items()
        )
        messages.error(request, f'Error al guardar: {errores}')
    return redirect('tabuladores_config')


@login_required
def tabuladores_editar_regla(request, regla_id):
    """POST: actualiza una ReglaTerapeuta desde el modal de edición."""
    if not request.user.is_staff or request.method != 'POST':
        return redirect('tabuladores_config')

    from .models import ReglaTerapeuta
    from .forms import ReglaTerapeutaForm

    regla = get_object_or_404(ReglaTerapeuta, id=regla_id)
    form = ReglaTerapeutaForm(request.POST, instance=regla)
    if form.is_valid():
        form.save()
        messages.success(request, f'Regla de {regla.terapeuta.nombre} actualizada correctamente.')
    else:
        errores = '; '.join(
            f'{field}: {", ".join(errs)}'
            for field, errs in form.errors.items()
        )
        messages.error(request, f'Error al guardar: {errores}')
    return redirect('tabuladores_config')


# ─────────────────────────────────────────────────────────────
# DISPONIBILIDAD SEMANAL
# ─────────────────────────────────────────────────────────────

DIAS_SEMANA = [
    (0, 'Lunes'), (1, 'Martes'), (2, 'Miércoles'),
    (3, 'Jueves'), (4, 'Viernes'), (5, 'Sábado'), (6, 'Domingo'),
]

@login_required
def disponibilidad_semanal(request):
    terapeutas = Terapeuta.objects.filter(activo=True).order_by('nombre')
    terapeutas_data = []
    for t in terapeutas:
        horarios_qs = list(Horario.objects.filter(terapeuta=t).order_by('dia', 'hora_inicio'))
        dias_data = []
        for dia_num, dia_nombre in DIAS_SEMANA:
            slots = [h for h in horarios_qs if h.dia == dia_num]
            dias_data.append({'num': dia_num, 'nombre': dia_nombre, 'slots': slots})
        terapeutas_data.append({
            'terapeuta': t,
            'dias': dias_data,
            'total': len(horarios_qs),
        })
    return render(request, 'clinica/disponibilidad_semanal.html', {
        'terapeutas_data': terapeutas_data,
    })


@login_required
def agregar_disponibilidad(request):
    if request.method != 'POST':
        return redirect('disponibilidad_semanal')
    terapeuta = get_object_or_404(Terapeuta, id=request.POST.get('terapeuta'))
    dia = request.POST.get('dia')
    hora_inicio = request.POST.get('hora_inicio')
    hora_fin = request.POST.get('hora_fin')
    sede = request.POST.get('sede') or None
    if not (dia and hora_inicio and hora_fin and sede):
        messages.error(request, 'Completa todos los campos del horario, incluyendo la sede.')
        return redirect('disponibilidad_semanal')
    if hora_fin <= hora_inicio:
        messages.error(request, 'La hora de fin debe ser mayor que la de inicio.')
        return redirect('disponibilidad_semanal')
    Horario.objects.create(terapeuta=terapeuta, dia=int(dia), hora_inicio=hora_inicio, hora_fin=hora_fin, sede=sede)
    dia_display = dict(Horario.DIAS_SEMANA).get(int(dia), dia)
    _registrar_actividad(request, 'disponib_agregada', 'disponibilidad',
        f'Disponibilidad de {terapeuta.nombre} agregada: {dia_display} de {hora_inicio} a {hora_fin}.',
        terapeuta=terapeuta)
    messages.success(request, f'Horario agregado para {terapeuta.nombre}.')
    return redirect('disponibilidad_semanal')


@login_required
def eliminar_disponibilidad(request, horario_id):
    if request.method != 'POST':
        return redirect('disponibilidad_semanal')
    horario = get_object_or_404(Horario, id=horario_id)
    nombre      = horario.terapeuta.nombre
    ter_obj     = horario.terapeuta
    dia_display = dict(Horario.DIAS_SEMANA).get(horario.dia, str(horario.dia))
    horario.delete()
    _registrar_actividad(request, 'disponib_eliminada', 'disponibilidad',
        f'Disponibilidad de {nombre} eliminada: {dia_display}.',
        terapeuta=ter_obj)
    messages.success(request, f'Horario eliminado para {nombre}.')
    return redirect('disponibilidad_semanal')


@api_key_required
def api_consultorios_por_horario(request):
    """Devuelve los consultorios válidos para un terapeuta en una fecha y hora dadas."""
    from .models import Consultorio as ConsultorioModel
    terapeuta_id = request.GET.get('terapeuta')
    fecha_str    = request.GET.get('fecha')
    hora_str     = request.GET.get('hora')

    activos = ConsultorioModel.objects.filter(activo=True)

    if not (terapeuta_id and fecha_str and hora_str):
        todos = list(activos.values('id', 'nombre').order_by('nombre'))
        return JsonResponse({'consultorios': todos, 'filtrado': False})

    try:
        if len(hora_str) > 5:
            hora_str = hora_str[:5]
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        hora_obj  = datetime.strptime(hora_str, '%H:%M').time()
        dia_semana = fecha_obj.weekday()

        horarios_dia = list(Horario.objects.filter(terapeuta_id=terapeuta_id, dia=dia_semana))
        horario_activo = next(
            (h for h in horarios_dia if h.hora_inicio <= hora_obj <= h.hora_fin), None
        )

        if horario_activo and horario_activo.sede:
            consultorios = list(activos.filter(sede=horario_activo.sede).values('id', 'nombre').order_by('nombre'))
            return JsonResponse({'consultorios': consultorios, 'filtrado': True, 'sede': horario_activo.sede})

        todos = list(activos.values('id', 'nombre').order_by('nombre'))
        return JsonResponse({'consultorios': todos, 'filtrado': False})

    except Exception as e:
        print(f'Error en api_consultorios_por_horario: {e}')
        todos = list(activos.values('id', 'nombre').order_by('nombre'))
        return JsonResponse({'consultorios': todos, 'filtrado': False})


# =============================================================================
# PORTAL EMPRESA
# =============================================================================

@login_required
def portal_empresa(request):
    if not hasattr(request.user, 'perfil_empresa'):
        return redirect('home')

    mi_empresa = request.user.perfil_empresa
    hoy = date.today()

    # Solo nombre y teléfono — no expediente
    pacientes = mi_empresa.pacientes.only('id', 'nombre', 'telefono').order_by('nombre')

    proximas_citas_qs = Cita.objects.filter(
        paciente__empresa=mi_empresa,
        fecha__gte=hoy,
        estatus__in=Cita.ESTATUS_ACTIVOS,
    ).select_related('paciente', 'terapeuta').order_by('fecha', 'hora')

    return render(request, 'clinica/portal_empresa.html', {
        'empresa': mi_empresa,
        'pacientes': pacientes,
        'total_pacientes': pacientes.count(),
        'proximas_citas': proximas_citas_qs[:10],
        'total_proximas_citas': proximas_citas_qs.count(),
        'hoy': hoy,
    })


@login_required
def registrar_paciente_empresa(request):
    if not hasattr(request.user, 'perfil_empresa'):
        return redirect('home')

    mi_empresa = request.user.perfil_empresa

    if request.method == 'POST':
        form = PacienteEmpresaForm(request.POST)
        if form.is_valid():
            paciente = form.save(commit=False)
            paciente.empresa = mi_empresa
            if mi_empresa.division:
                paciente.division = mi_empresa.division
            paciente.save()
            messages.success(request, f'Paciente {paciente.nombre} registrado correctamente.')
            return redirect('portal_empresa')
    else:
        form = PacienteEmpresaForm()

    return render(request, 'clinica/registrar_paciente_empresa.html', {
        'form': form,
        'empresa': mi_empresa,
    })


@login_required
def agendar_cita_empresa(request):
    if not hasattr(request.user, 'perfil_empresa'):
        return redirect('home')

    mi_empresa = request.user.perfil_empresa
    terapeutas = Terapeuta.objects.filter(activo=True, horario__isnull=False).distinct().order_by('nombre')

    if request.method == 'POST':
        primera_vez = request.POST.get('primera_vez') == '1'

        if primera_vez:
            paciente_id = request.POST.get('paciente')
            fecha_str   = request.POST.get('fecha', '').strip()
            hora_str    = request.POST.get('hora', '').strip()
            paciente    = None
            fecha_obj   = None
            hora_obj    = None
            errores     = []

            try:
                paciente = mi_empresa.pacientes.get(pk=paciente_id)
            except Exception:
                errores.append('Selecciona un colaborador válido.')

            if not fecha_str:
                errores.append('La fecha es requerida.')
            else:
                try:
                    fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                except ValueError:
                    errores.append('Fecha inválida.')

            if hora_str:
                try:
                    hora_obj = datetime.strptime(hora_str, '%H:%M').time()
                except ValueError:
                    pass

            if errores:
                for e in errores:
                    messages.error(request, e)
            else:
                SolicitudCita.objects.create(
                    paciente_nombre=paciente.nombre,
                    telefono=paciente.telefono or '',
                    fecha_deseada=fecha_obj,
                    hora_deseada=hora_obj,
                    paciente=paciente,
                    empresa=mi_empresa,
                    division=mi_empresa.division,
                    estado='pendiente',
                )
                messages.success(request, f'Solicitud de primera cita enviada para {paciente.nombre}. Recepción asignará terapeuta disponible.')
                return redirect('portal_empresa')

            form = CitaEmpresaForm(empresa=mi_empresa)
        else:
            form = CitaEmpresaForm(request.POST, empresa=mi_empresa)
            if form.is_valid():
                data = form.cleaned_data
                paciente = data['paciente']
                SolicitudCita.objects.create(
                    paciente_nombre=paciente.nombre,
                    telefono=paciente.telefono or '',
                    fecha_deseada=data['fecha'],
                    hora_deseada=data['hora'],
                    terapeuta=data.get('terapeuta'),
                    consultorio=data.get('consultorio'),
                    paciente=paciente,
                    empresa=mi_empresa,
                    division=data.get('division') or mi_empresa.division,
                    servicio=data.get('servicio'),
                    estado='pendiente',
                )
                messages.success(request, f'Solicitud enviada para {paciente.nombre}. Recepción la confirmará pronto.')
                return redirect('portal_empresa')
    else:
        form = CitaEmpresaForm(empresa=mi_empresa)

    return render(request, 'clinica/agendar_cita_empresa.html', {
        'form': form,
        'empresa': mi_empresa,
        'terapeutas': terapeutas,
    })


@login_required
def terapeutas_paciente_empresa(request):
    if not hasattr(request.user, 'perfil_empresa'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    mi_empresa  = request.user.perfil_empresa
    paciente_id = request.GET.get('paciente_id')

    if not paciente_id:
        return JsonResponse({'terapeutas': []})

    try:
        paciente = mi_empresa.pacientes.get(pk=paciente_id)
    except Exception:
        return JsonResponse({'terapeutas': []})

    terapeutas = (
        Terapeuta.objects
        .filter(cita__paciente=paciente, activo=True)
        .distinct()
        .values('id', 'nombre')
    )

    return JsonResponse({'terapeutas': list(terapeutas)})


@login_required
def citas_en_proceso_empresa(request):
    if not hasattr(request.user, 'perfil_empresa'):
        return redirect('home')

    mi_empresa = request.user.perfil_empresa
    hoy = date.today()

    solicitudes_pendientes = SolicitudCita.objects.filter(
        empresa=mi_empresa,
        estado='pendiente',
    ).select_related('paciente', 'terapeuta', 'consultorio').order_by('fecha_deseada', 'hora_deseada')

    solicitudes_rechazadas = SolicitudCita.objects.filter(
        empresa=mi_empresa,
        estado='rechazada',
    ).select_related('paciente', 'terapeuta').order_by('-fecha_creacion')[:10]

    citas_proximas = Cita.objects.filter(
        paciente__empresa=mi_empresa,
        fecha__gte=hoy,
        estatus__in=Cita.ESTATUS_ACTIVOS,
    ).select_related('paciente', 'terapeuta', 'consultorio').order_by('fecha', 'hora')

    return render(request, 'clinica/citas_en_proceso_empresa.html', {
        'empresa': mi_empresa,
        'solicitudes_pendientes': solicitudes_pendientes,
        'solicitudes_rechazadas': solicitudes_rechazadas,
        'citas_proximas': citas_proximas,
        'hoy': hoy,
    })


# ---------------------------------------------------------------------------
# REAGENDOS — flujo terapeuta propone, recepción decide
# ---------------------------------------------------------------------------

@login_required
def solicitar_reagendo(request, cita_id):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    mi_perfil = request.user.perfil_terapeuta
    cita = get_object_or_404(Cita, id=cita_id, terapeuta=mi_perfil)

    if request.method != 'POST':
        return redirect('portal_terapeuta')

    # Evitar duplicados: solo una solicitud pendiente por cita
    if SolicitudReagendo.objects.filter(cita=cita, estado='pendiente').exists():
        messages.warning(request, 'Ya existe una solicitud de reagendo pendiente para esa cita.')
        return redirect('portal_terapeuta')

    fecha_str = request.POST.get('fecha_propuesta', '').strip()
    hora_str  = request.POST.get('hora_propuesta', '').strip()
    motivo    = request.POST.get('motivo', '').strip()

    from datetime import datetime as dt
    try:
        fecha = dt.strptime(fecha_str, '%Y-%m-%d').date()
        hora  = dt.strptime(hora_str,  '%H:%M').time()
    except ValueError:
        messages.error(request, 'Fecha u hora inválida. Intenta de nuevo.')
        return redirect('portal_terapeuta')

    SolicitudReagendo.objects.create(
        cita=cita,
        terapeuta=mi_perfil,
        fecha_propuesta=fecha,
        hora_propuesta=hora,
        motivo=motivo,
    )
    cita.estatus = Cita.ESTATUS_REAGENDO
    cita.save(update_fields=['estatus'])
    _registrar_actividad(request, 'reagendo_solicitado', 'reagendo',
        f'{mi_perfil.nombre} solicitó reagendar la cita de '
        f'{cita.paciente.nombre if cita.paciente else "?"} al '
        f'{fecha.strftime("%d/%m/%Y")} a las {hora.strftime("%H:%M")}.',
        terapeuta=mi_perfil, paciente=cita.paciente)
    messages.success(request, 'Solicitud de reagendo enviada a Recepción.')
    return redirect('portal_terapeuta')


@login_required
def aprobar_reagendo(request, solicitud_id):
    if request.method != 'POST':
        return redirect('home')

    from .forms import verificar_empalme_paciente
    from django.db.models import Q as _Q

    solicitud = get_object_or_404(SolicitudReagendo, id=solicitud_id, estado='pendiente')
    cita = solicitud.cita
    nueva_fecha = solicitud.fecha_propuesta
    nueva_hora  = solicitud.hora_propuesta

    # Verificar empalme de todos los pacientes de la cita
    todos_pacientes = [cita.paciente] + list(cita.pacientes_adicionales.all())
    conflictos = []
    for p in todos_pacientes:
        c = verificar_empalme_paciente(p, nueva_fecha, nueva_hora, excluir_cita_id=cita.pk)
        if c:
            terapeuta_str = c.terapeuta.nombre if c.terapeuta else 'sin terapeuta'
            conflictos.append(
                f"{p.nombre} ya tiene cita el {nueva_fecha:%d/%m/%Y} a las {nueva_hora:%H:%M} "
                f"con {terapeuta_str} ({c.get_estatus_display()})."
            )

    if conflictos:
        for msg in conflictos:
            messages.error(request, f"No se puede reagendar: {msg}")
        return redirect('home')

    cita.fecha   = nueva_fecha
    cita.hora    = nueva_hora
    cita.estatus = Cita.ESTATUS_CONFIRMADA
    cita.save(update_fields=['fecha', 'hora', 'estatus'])

    solicitud.estado = 'aprobada'
    solicitud.nota_recepcion = request.POST.get('nota_recepcion', '').strip()
    solicitud.save(update_fields=['estado', 'nota_recepcion'])
    _registrar_actividad(request, 'reagendo_aprobado', 'reagendo',
        f'Reagendo de {solicitud.terapeuta.nombre} para '
        f'{cita.paciente.nombre if cita.paciente else "?"} aprobado. '
        f'Nueva fecha: {nueva_fecha.strftime("%d/%m/%Y")} a las {nueva_hora.strftime("%H:%M")}.',
        terapeuta=solicitud.terapeuta, paciente=cita.paciente)

    NotificacionTerapeuta.objects.create(
        terapeuta=solicitud.terapeuta,
        tipo=NotificacionTerapeuta.TIPO_REAGENDO_APROBADO,
        mensaje=f"Tu solicitud de reagendo para la cita del paciente {cita.paciente} fue aprobada. Nueva fecha: {cita.fecha:%d/%m/%Y} a las {cita.hora:%H:%M}.",
    )

    messages.success(request, f'Reagendo aprobado. Cita movida al {cita.fecha:%d/%m/%Y} a las {cita.hora:%H:%M}.')
    return redirect('home')


@login_required
def rechazar_reagendo(request, solicitud_id):
    if request.method != 'POST':
        return redirect('home')

    solicitud = get_object_or_404(SolicitudReagendo, id=solicitud_id, estado='pendiente')
    solicitud.estado = 'rechazada'
    solicitud.nota_recepcion = request.POST.get('nota_recepcion', '').strip()
    solicitud.save(update_fields=['estado', 'nota_recepcion'])

    cita = solicitud.cita
    cita.estatus = Cita.ESTATUS_CONFIRMADA
    cita.save(update_fields=['estatus'])
    _registrar_actividad(request, 'reagendo_rechazado', 'reagendo',
        f'Reagendo de {solicitud.terapeuta.nombre} para '
        f'{cita.paciente.nombre if cita.paciente else "?"} rechazado.',
        terapeuta=solicitud.terapeuta, paciente=cita.paciente)

    nota_texto = f" Nota: {solicitud.nota_recepcion}" if solicitud.nota_recepcion else ""
    NotificacionTerapeuta.objects.create(
        terapeuta=solicitud.terapeuta,
        tipo=NotificacionTerapeuta.TIPO_REAGENDO_RECHAZADO,
        mensaje=f"Tu solicitud de reagendo para la cita del paciente {cita.paciente} fue rechazada. La cita se mantiene el {cita.fecha:%d/%m/%Y} a las {cita.hora:%H:%M}.{nota_texto}",
    )

    messages.info(request, 'Solicitud de reagendo rechazada. La cita se mantuvo en su fecha original.')
    return redirect('home')


@login_required
def precios_servicios(request):
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('home')

    from clinica.models import Servicio
    servicios = Servicio.objects.all().order_by('nombre')

    if request.method == 'POST':
        errores = []
        for servicio in servicios:
            key = f'precio_{servicio.id}'
            valor = request.POST.get(key, '').strip()
            if valor == '':
                servicio.precio = None
                servicio.save(update_fields=['precio'])
            else:
                try:
                    servicio.precio = Decimal(valor)
                    servicio.save(update_fields=['precio'])
                except Exception:
                    errores.append(f'Precio inválido para "{servicio.nombre}": {valor}')
        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            messages.success(request, 'Precios actualizados correctamente.')
        return redirect('precios_servicios')

    return render(request, 'clinica/precios_servicios.html', {'servicios': servicios})


@login_required
def reportar_incidente(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    if request.method != 'POST':
        return redirect('portal_terapeuta')

    tipo = request.POST.get('tipo', '').strip()
    titulo = request.POST.get('titulo', '').strip()
    descripcion = request.POST.get('descripcion', '').strip()

    tipos_validos = [ReporteIncidente.TIPO_QUEJA, ReporteIncidente.TIPO_SUGERENCIA, ReporteIncidente.TIPO_INCIDENTE]
    if tipo not in tipos_validos or not titulo or not descripcion:
        messages.error(request, 'Por favor completa todos los campos del reporte.')
        return redirect('portal_terapeuta')

    ReporteIncidente.objects.create(
        terapeuta=request.user.perfil_terapeuta,
        tipo=tipo,
        titulo=titulo,
        descripcion=descripcion,
    )
    tipo_display = {'queja': 'queja', 'sugerencia': 'sugerencia', 'incidente': 'incidente'}.get(tipo, tipo)
    _registrar_actividad(request, 'incidente_reportado', 'incidente',
        f'{request.user.perfil_terapeuta.nombre} reportó una {tipo_display}: "{titulo}".',
        terapeuta=request.user.perfil_terapeuta)
    messages.success(request, 'Tu reporte fue enviado correctamente. El equipo lo revisará pronto.')
    return redirect('portal_terapeuta')


@login_required
def lista_incidentes(request):
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('home')

    incidentes = ReporteIncidente.objects.select_related('terapeuta').all()

    estado_filter = request.GET.get('estado', '')
    if estado_filter in [ReporteIncidente.ESTADO_PENDIENTE, ReporteIncidente.ESTADO_REVISADO]:
        incidentes = incidentes.filter(estado=estado_filter)

    if request.method == 'POST':
        incidente_id = request.POST.get('incidente_id')
        nuevo_estado = request.POST.get('estado')
        incidente = get_object_or_404(ReporteIncidente, id=incidente_id)
        if nuevo_estado in [ReporteIncidente.ESTADO_PENDIENTE, ReporteIncidente.ESTADO_REVISADO]:
            incidente.estado = nuevo_estado
            incidente.save(update_fields=['estado'])
        return redirect('lista_incidentes')

    pendientes_count = ReporteIncidente.objects.filter(estado=ReporteIncidente.ESTADO_PENDIENTE).count()
    return render(request, 'clinica/lista_incidentes.html', {
        'incidentes': incidentes,
        'estado_filter': estado_filter,
        'pendientes_count': pendientes_count,
    })


@login_required
def responder_incidente(request):
    if not (request.user.is_superuser or request.user.is_staff):
        return redirect('home')
    if request.method != 'POST':
        return redirect('lista_incidentes')

    incidente_id = request.POST.get('incidente_id')
    respuesta = request.POST.get('respuesta', '').strip()

    if not respuesta:
        messages.error(request, 'La respuesta no puede estar vacía.')
        return redirect('lista_incidentes')

    incidente = get_object_or_404(ReporteIncidente, id=incidente_id)
    incidente.respuesta = respuesta
    incidente.respuesta_fecha = timezone.now()
    incidente.estado = ReporteIncidente.ESTADO_REVISADO
    incidente.save(update_fields=['respuesta', 'respuesta_fecha', 'estado'])
    _registrar_actividad(request, 'incidente_respondido', 'incidente',
        f'Incidente "{incidente.titulo}" de {incidente.terapeuta.nombre} respondido.',
        terapeuta=incidente.terapeuta)

    NotificacionTerapeuta.objects.create(
        terapeuta=incidente.terapeuta,
        tipo=NotificacionTerapeuta.TIPO_RESPUESTA_INCIDENTE,
        mensaje=f'Recepción respondió tu reporte "{incidente.titulo}": {respuesta}',
    )

    messages.success(request, 'Respuesta enviada correctamente.')
    return redirect('lista_incidentes')


@login_required
def expedientes_grupales_lista(request):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    terapeuta = request.user.perfil_terapeuta
    paciente_ids = _pacientes_ids_terapeuta(terapeuta)
    expedientes = (
        ExpedienteGrupal.objects
        .filter(pacientes__id__in=paciente_ids)
        .prefetch_related('pacientes')
        .distinct()
        .order_by('-fecha_apertura')
    )
    return render(request, 'clinica/expedientes_grupales_lista.html', {
        'expedientes': expedientes,
        'terapeuta': terapeuta,
    })


@login_required
def expediente_grupal_detalle(request, expediente_id):
    if not hasattr(request.user, 'perfil_terapeuta'):
        return redirect('home')

    terapeuta = request.user.perfil_terapeuta
    expediente = get_object_or_404(ExpedienteGrupal, id=expediente_id)

    paciente_ids_terapeuta = _pacientes_ids_terapeuta(terapeuta)
    paciente_ids_grupal = set(expediente.pacientes.values_list('id', flat=True))
    if not paciente_ids_terapeuta.intersection(paciente_ids_grupal):
        messages.error(request, 'No tienes acceso a este expediente.')
        return redirect('expedientes_terapeuta')

    apertura = getattr(expediente, 'apertura', None)
    pacientes_grupo = list(expediente.pacientes.all())

    citas = expediente.citas.select_related(
        'terapeuta', 'servicio', 'consultorio'
    ).prefetch_related('pacientes_adicionales').order_by('-fecha', '-hora')

    from django.db.models import Min
    reporte_ids = (
        ReporteSesion.objects
        .filter(cita__expediente_grupal=expediente)
        .values('cita_id')
        .annotate(primer_id=Min('id'))
        .values_list('primer_id', flat=True)
    )
    reportes = (
        ReporteSesion.objects
        .filter(id__in=reporte_ids)
        .select_related('paciente', 'terapeuta')
        .order_by('-fecha', '-creado_en')
    )

    notas = expediente.notas.select_related('terapeuta').all()

    hoy = date.today()
    cita_hoy = expediente.citas.filter(fecha=hoy).order_by('hora').last()

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'guardar_apertura':
            form_apertura = AperturaExpedienteGrupalForm(request.POST, instance=apertura)
            if form_apertura.is_valid():
                ap = form_apertura.save(commit=False)
                ap.expediente = expediente
                ap.save()
                messages.success(request, 'Apertura guardada correctamente.')
            else:
                messages.error(request, 'Revisa los datos de la apertura.')
            return redirect('expediente_grupal_detalle', expediente_id=expediente.id)

        if accion == 'guardar_reporte_grupal':
            form_reporte = ReporteSesionForm(request.POST)
            if form_reporte.is_valid():
                for paciente in pacientes_grupo:
                    num_sesion = ReporteSesion.objects.filter(
                        terapeuta=terapeuta, paciente=paciente
                    ).count() + 1
                    reporte = form_reporte.save(commit=False)
                    reporte.pk = None
                    reporte.paciente = paciente
                    reporte.terapeuta = terapeuta
                    reporte.numero_sesion = num_sesion
                    reporte.cita = cita_hoy
                    reporte.save()
                messages.success(request, f'Reporte guardado para {len(pacientes_grupo)} paciente(s).')
            else:
                messages.error(request, 'Revisa los datos del reporte.')
            return redirect('expediente_grupal_detalle', expediente_id=expediente.id)

        if accion == 'agregar_nota':
            contenido = request.POST.get('contenido', '').strip()
            if contenido:
                NotaExpedienteGrupal.objects.create(
                    expediente=expediente,
                    terapeuta=terapeuta,
                    contenido=contenido,
                )
                messages.success(request, 'Nota agregada al expediente.')
            else:
                messages.error(request, 'La nota no puede estar vacía.')
            return redirect('expediente_grupal_detalle', expediente_id=expediente.id)

        if accion == 'eliminar_nota':
            nota_id = request.POST.get('nota_id')
            NotaExpedienteGrupal.objects.filter(
                id=nota_id, expediente=expediente, terapeuta=terapeuta
            ).delete()
            messages.success(request, 'Nota eliminada.')
            return redirect('expediente_grupal_detalle', expediente_id=expediente.id)

    form_apertura = AperturaExpedienteGrupalForm(instance=apertura)
    form_reporte = ReporteSesionForm(initial={
        'fecha': hoy,
        'hora_inicio': cita_hoy.hora if cita_hoy else None,
    })

    return render(request, 'clinica/expediente_grupal_detalle.html', {
        'expediente': expediente,
        'apertura': apertura,
        'form_apertura': form_apertura,
        'form_reporte': form_reporte,
        'pacientes_grupo': pacientes_grupo,
        'citas': citas,
        'reportes': reportes,
        'notas': notas,
        'terapeuta': terapeuta,
        'cita_hoy': cita_hoy,
    })


@api_key_required
def api_notificaciones_terapeuta(request):
    """Devuelve las notificaciones no leídas del terapeuta como JSON."""
    from django.http import JsonResponse
    if not hasattr(request.user, 'perfil_terapeuta'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    terapeuta = request.user.perfil_terapeuta
    notifs = NotificacionTerapeuta.objects.filter(
        terapeuta=terapeuta, leida=False
    ).order_by('-creada_en')[:20]

    data = [
        {
            'id': n.id,
            'tipo': n.tipo,
            'mensaje': n.mensaje,
            'creada_en': n.creada_en.strftime('%d/%m/%Y %H:%M'),
        }
        for n in notifs
    ]
    return JsonResponse({'notificaciones': data, 'total': len(data)})


@login_required
def marcar_notificaciones_leidas(request):
    """Marca como leídas todas las notificaciones del terapeuta."""
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    if not hasattr(request.user, 'perfil_terapeuta'):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    terapeuta = request.user.perfil_terapeuta
    NotificacionTerapeuta.objects.filter(
        terapeuta=terapeuta, leida=False
    ).update(leida=True)
    return JsonResponse({'ok': True})


@login_required
def agregar_nota_recepcion(request):
    if request.method == 'POST':
        texto = request.POST.get('texto', '').strip()
        if texto:
            NotaRecepcion.objects.create(texto=texto, creado_por=request.user)
    return redirect('home')


@login_required
def eliminar_nota_recepcion(request, nota_id):
    nota = get_object_or_404(NotaRecepcion, id=nota_id)
    nota.delete()
    return redirect('home')


@login_required
def toggle_reaccion_nota(request, nota_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'method not allowed'}, status=405)
    nota = get_object_or_404(NotaRecepcion, id=nota_id)
    emoji = request.POST.get('emoji', '').strip()
    if not emoji:
        return JsonResponse({'error': 'emoji required'}, status=400)
    obj, created = ReaccionNota.objects.get_or_create(nota=nota, usuario=request.user, emoji=emoji)
    if not created:
        obj.delete()
        yo = False
    else:
        yo = True
    count = ReaccionNota.objects.filter(nota=nota, emoji=emoji).count()
    return JsonResponse({'emoji': emoji, 'count': count, 'yo': yo})


@login_required
def agregar_comentario_nota(request, nota_id):
    if request.method == 'POST':
        nota = get_object_or_404(NotaRecepcion, id=nota_id)
        texto = request.POST.get('texto', '').strip()
        if texto:
            ComentarioNota.objects.create(nota=nota, creado_por=request.user, texto=texto)
    return redirect('home')


@login_required
def solicitar_horas_expositor(request):
    if request.method != 'POST':
        return redirect('portal_terapeuta')
    try:
        mi_perfil = request.user.perfil_terapeuta
    except Exception:
        return redirect('home')

    horas_raw = request.POST.get('horas', '').strip()
    lugar     = request.POST.get('lugar', '').strip()
    notas     = request.POST.get('notas', '').strip()

    if not horas_raw or not lugar or not notas:
        messages.error(request, 'Todos los campos son obligatorios.')
        return redirect('portal_terapeuta')

    try:
        horas = int(horas_raw)
        if horas < 1:
            raise ValueError
    except ValueError:
        messages.error(request, 'Las horas deben ser un número entero mayor a 0.')
        return redirect('portal_terapeuta')

    SolicitudHorasExpositor.objects.create(
        terapeuta=mi_perfil,
        horas=horas,
        lugar=lugar,
        notas=notas,
    )
    messages.success(request, f'Solicitud de {horas} hora{"s" if horas > 1 else ""} expositor enviada a recepción.')
    return redirect('portal_terapeuta')


@login_required
def responder_horas_expositor(request, solicitud_id):
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('home')
    if request.method != 'POST':
        return redirect('home')

    from .services import calcular_nomina_semanal
    solicitud = get_object_or_404(SolicitudHorasExpositor, id=solicitud_id)
    if solicitud.estado != SolicitudHorasExpositor.ESTADO_PENDIENTE:
        messages.warning(request, 'Esta solicitud ya fue procesada.')
        return redirect('home')

    accion = request.POST.get('accion')

    if accion == 'aceptar':
        terapeuta = solicitud.terapeuta
        monto_por_hora = Decimal('0.00')
        try:
            regla = terapeuta.regla_pago
            if regla.pago_individual is not None:
                monto_por_hora = regla.pago_individual
            elif regla.pago_por_sesion is not None:
                monto_por_hora = regla.pago_por_sesion
            elif regla.tabulador_base and regla.tabulador_base.pago_base is not None:
                monto_por_hora = regla.tabulador_base.pago_base
        except Exception:
            pass

        monto_total = monto_por_hora * solicitud.horas

        hoy_local  = date.today()
        ini_semana = hoy_local - timedelta(days=hoy_local.weekday())
        fin_semana = ini_semana + timedelta(days=6)

        corte, _ = CorteSemanal.objects.get_or_create(
            terapeuta=terapeuta,
            fecha_inicio=ini_semana,
            defaults={'fecha_fin': fin_semana, 'estatus': CorteSemanal.ESTATUS_BORRADOR},
        )
        if corte.estatus != CorteSemanal.ESTATUS_BORRADOR:
            corte_alt = (
                CorteSemanal.objects
                .filter(terapeuta=terapeuta, estatus=CorteSemanal.ESTATUS_BORRADOR)
                .order_by('-fecha_inicio')
                .first()
            )
            if not corte_alt:
                messages.error(
                    request,
                    f'No hay un corte en borrador para {terapeuta.nombre}. '
                    'Crea uno primero en la sección de nómina.',
                )
                return redirect('home')
            corte = corte_alt

        LineaNomina.objects.create(
            corte=corte,
            cita=None,
            tipo=LineaNomina.TIPO_EXPOSITOR,
            concepto=f'Horas Expositor: {solicitud.horas}h en "{solicitud.lugar}"',
            monto=monto_total,
        )
        try:
            calcular_nomina_semanal(terapeuta, corte.fecha_inicio, corte.fecha_fin)
        except ValueError:
            pass

        solicitud.estado = SolicitudHorasExpositor.ESTADO_ACEPTADA
        solicitud.save()
        _registrar_actividad(request, 'expositor_respondido', 'nomina',
            f'Solicitud de {solicitud.horas}h expositor de {terapeuta.nombre} en "{solicitud.lugar}" aceptada.',
            terapeuta=terapeuta)

        NotificacionTerapeuta.objects.create(
            terapeuta=terapeuta,
            tipo=NotificacionTerapeuta.TIPO_EXPOSITOR_ACEPTADO,
            mensaje=(
                f'Tu solicitud de {solicitud.horas} hora{"s" if solicitud.horas > 1 else ""} expositor '
                f'en "{solicitud.lugar}" fue aceptada. '
                f'Se agregaron ${monto_total:,.2f} a tu nómina.'
            ),
        )
        messages.success(
            request,
            f'Solicitud aceptada. Se agregaron ${monto_total:,.2f} a la nómina de {terapeuta.nombre}.',
        )

    elif accion == 'rechazar':
        solicitud.estado = SolicitudHorasExpositor.ESTADO_RECHAZADA
        solicitud.save()
        _registrar_actividad(request, 'expositor_respondido', 'nomina',
            f'Solicitud de {solicitud.horas}h expositor de {solicitud.terapeuta.nombre} en "{solicitud.lugar}" rechazada.',
            terapeuta=solicitud.terapeuta)
        NotificacionTerapeuta.objects.create(
            terapeuta=solicitud.terapeuta,
            tipo=NotificacionTerapeuta.TIPO_EXPOSITOR_RECHAZADO,
            mensaje=(
                f'Tu solicitud de {solicitud.horas} hora{"s" if solicitud.horas > 1 else ""} expositor '
                f'en "{solicitud.lugar}" fue rechazada por recepción.'
            ),
        )
        messages.success(request, f'Solicitud rechazada. Se notificó a {solicitud.terapeuta.nombre}.')

    return redirect('home')


@login_required
def citas_tardias(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('home')

    hoy = date.today()
    desde = hoy - timedelta(days=7)

    ESTATUS_VALIDOS = {
        Cita.ESTATUS_SI_ASISTIO:    'Si asistió',
        Cita.ESTATUS_NO_ASISTIO:    'No asistió',
        Cita.ESTATUS_CANCELO:       'Canceló',
        Cita.ESTATUS_INCIDENCIA:    'Incidencia',
    }

    if request.method == 'POST':
        cita_id   = request.POST.get('cita_id')
        nuevo_est = request.POST.get('estatus')
        if nuevo_est in ESTATUS_VALIDOS:
            cita_obj = get_object_or_404(Cita, id=cita_id)
            cita_obj.estatus  = nuevo_est
            cita_obj.sin_bono = True
            cita_obj.save(update_fields=['estatus', 'sin_bono'])
            messages.success(
                request,
                f'Cita de {cita_obj.paciente} actualizada a "{ESTATUS_VALIDOS[nuevo_est]}" (sin bono).',
            )
        else:
            messages.error(request, 'Estatus no válido.')
        return redirect('citas_tardias')

    citas_qs = (
        Cita.objects
        .filter(
            fecha__gte=desde,
            fecha__lt=hoy,
            estatus__in=[Cita.ESTATUS_CONFIRMADA, Cita.ESTATUS_SIN_CONFIRMAR],
        )
        .select_related('paciente', 'terapeuta', 'consultorio')
        .order_by('fecha', 'hora')
    )

    return render(request, 'clinica/citas_tardias.html', {
        'citas': citas_qs,
        'estatus_opciones': ESTATUS_VALIDOS,
        'desde': desde,
        'hoy': hoy,
    })


@login_required
def catalogo_terapeutas(request):
    return render(request, 'clinica/catalogo_terapeutas.html')


# ── API Catálogo ──────────────────────────────────────────────────────────────

@api_key_required
def api_catalogo_list(request):
    if request.method == 'GET':
        perfiles = (PerfilCatalogo.objects
                    .filter(activo=True)
                    .select_related('terapeuta')
                    .order_by('nombre'))
        data = [
            {
                'id':               p.id,
                'nombre':           p.nombre,
                'titulo':           p.titulo,
                'cedula':           p.cedula,
                'preparacion':      p.preparacion,
                'formacion':        p.formacion,
                'terapeuta_id':     p.terapeuta_id,
                'terapeuta_nombre': p.terapeuta.nombre if p.terapeuta else None,
            }
            for p in perfiles
        ]
        return JsonResponse(data, safe=False)

    if request.method == 'POST':
        try:
            body = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        nombre = body.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'El nombre es obligatorio'}, status=400)

        terapeuta = None
        tid = body.get('terapeuta_id')
        if tid:
            try:
                terapeuta = Terapeuta.objects.get(id=int(tid))
            except (Terapeuta.DoesNotExist, ValueError):
                pass

        perfil = PerfilCatalogo.objects.create(
            nombre=nombre,
            titulo=body.get('titulo', 'Licenciatura'),
            cedula=body.get('cedula', '') or 'Sin cédula registrada',
            preparacion=body.get('preparacion', ''),
            formacion=body.get('formacion', ''),
            terapeuta=terapeuta,
        )
        return JsonResponse({'id': perfil.id}, status=201)

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@api_key_required
def api_catalogo_detail(request, perfil_id):
    perfil = get_object_or_404(PerfilCatalogo, id=perfil_id)

    if request.method == 'PUT':
        try:
            body = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        perfil.nombre      = body.get('nombre', perfil.nombre).strip() or perfil.nombre
        perfil.titulo      = body.get('titulo', perfil.titulo)
        perfil.cedula      = body.get('cedula', perfil.cedula) or 'Sin cédula registrada'
        perfil.preparacion = body.get('preparacion', perfil.preparacion)
        perfil.formacion   = body.get('formacion', perfil.formacion)

        tid = body.get('terapeuta_id')
        if tid is None or tid == '' or tid == 0:
            perfil.terapeuta = None
        else:
            try:
                perfil.terapeuta = Terapeuta.objects.get(id=int(tid))
            except (Terapeuta.DoesNotExist, ValueError):
                pass

        perfil.save()
        return JsonResponse({'ok': True})

    if request.method == 'DELETE':
        perfil.activo = False
        perfil.save()
        return JsonResponse({'ok': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@api_key_required
def api_terapeutas_activos(request):
    vinculados = set(
        PerfilCatalogo.objects
        .filter(terapeuta__isnull=False, activo=True)
        .values_list('terapeuta_id', flat=True)
    )
    terapeutas = Terapeuta.objects.filter(activo=True).order_by('nombre')
    data = [
        {
            'id':           t.id,
            'nombre':       t.nombre,
            'ya_vinculado': t.id in vinculados,
        }
        for t in terapeutas
    ]
    return JsonResponse(data, safe=False)


@login_required
def trazabilidad_admin(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('home')

    from django.contrib.auth.models import User as AuthUser
    from django.core.paginator import Paginator
    from datetime import datetime as _dt

    qs = (
        RegistroActividad.objects
        .select_related('usuario', 'terapeuta_afectado', 'paciente_afectado')
        .all()
    )

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    usuario_id  = request.GET.get('usuario', '').strip()
    categoria   = request.GET.get('categoria', '').strip()
    accion      = request.GET.get('accion', '').strip()
    q           = request.GET.get('q', '').strip()

    if fecha_desde:
        try:
            qs = qs.filter(timestamp__date__gte=_dt.strptime(fecha_desde, '%Y-%m-%d').date())
        except ValueError:
            fecha_desde = ''
    if fecha_hasta:
        try:
            qs = qs.filter(timestamp__date__lte=_dt.strptime(fecha_hasta, '%Y-%m-%d').date())
        except ValueError:
            fecha_hasta = ''
    if usuario_id:
        qs = qs.filter(usuario_id=usuario_id)
    if categoria:
        qs = qs.filter(categoria=categoria)
    if accion:
        qs = qs.filter(accion=accion)
    if q:
        qs = qs.filter(descripcion__icontains=q)

    total_registros  = qs.count()
    hoy              = timezone.now().date()
    registros_hoy    = qs.filter(timestamp__date=hoy).count()
    usuarios_activos = (
        RegistroActividad.objects.filter(timestamp__date=hoy, usuario__isnull=False)
        .values('usuario').distinct().count()
    )

    paginator = Paginator(qs, 50)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    usuarios_con_actividad = (
        AuthUser.objects
        .filter(registros_actividad__isnull=False)
        .distinct()
        .order_by('username')
    )

    return render(request, 'clinica/trazabilidad_admin.html', {
        'page_obj':               page_obj,
        'total_registros':        total_registros,
        'registros_hoy':          registros_hoy,
        'usuarios_activos':       usuarios_activos,
        'usuarios_con_actividad': usuarios_con_actividad,
        'categoria_choices':      RegistroActividad.CAT_CHOICES,
        'accion_choices':         RegistroActividad.ACCION_CHOICES,
        'f_fecha_desde':          fecha_desde,
        'f_fecha_hasta':          fecha_hasta,
        'f_usuario':              usuario_id,
        'f_categoria':            categoria,
        'f_accion':               accion,
        'f_q':                    q,
        'hoy':                    hoy,
    })


# =============================================================================
# API endpoints externos (protegidos con api_key_required)
# =============================================================================

@api_key_required
def api_notas_recepcion(request):
    from .models import NotaRecepcion
    if request.method == 'GET':
        notas = NotaRecepcion.objects.all().order_by('-creado_en')
        data = [{
            'id': n.id,
            'texto': n.texto,
            'creado_por': str(n.creado_por),
            'creado_en': n.creado_en.isoformat(),
        } for n in notas]
        return JsonResponse(data, safe=False)
    elif request.method == 'POST':
        import json
        body = json.loads(request.body)
        nota = NotaRecepcion.objects.create(
            texto=body['texto'],
            creado_por=request.user
        )
        return JsonResponse({
            'id': nota.id,
            'texto': nota.texto,
            'creado_por': str(nota.creado_por),
            'creado_en': nota.creado_en.isoformat(),
        }, status=201)


@api_key_required
def api_nota_recepcion_detail(request, nota_id):
    from .models import NotaRecepcion
    try:
        nota = NotaRecepcion.objects.get(id=nota_id)
    except NotaRecepcion.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    if request.method == 'DELETE':
        nota.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@api_key_required
def api_nomina_semanal(request):
    from .models import CorteSemanal, LineaNomina
    from django.utils.dateparse import parse_date
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    cortes = CorteSemanal.objects.select_related('terapeuta').all()
    if fecha_inicio:
        cortes = cortes.filter(fecha_inicio__gte=parse_date(fecha_inicio))
    if fecha_fin:
        cortes = cortes.filter(fecha_fin__lte=parse_date(fecha_fin))
    data = []
    for c in cortes:
        lineas = LineaNomina.objects.filter(corte=c).select_related('cita__paciente', 'cita__servicio')
        detalles = [{
            'paciente': l.cita.paciente.nombre if l.cita and l.cita.paciente else '',
            'fecha': l.cita.fecha.isoformat() if l.cita else '',
            'servicio': l.cita.servicio.nombre if l.cita and l.cita.servicio else '',
            'monto': float(l.monto),
        } for l in lineas]
        data.append({
            'id': c.id,
            'terapeuta': c.terapeuta.nombre,
            'fecha_inicio': c.fecha_inicio.isoformat(),
            'fecha_fin': c.fecha_fin.isoformat(),
            'total_sesiones': c.total_sesiones,
            'subtotal_sesiones': float(c.subtotal_sesiones),
            'total_bonos': float(c.total_bonos),
            'total_pago': float(c.total_pago),
            'estatus': c.estatus,
            'confirmacion_terapeuta': c.confirmacion_terapeuta,
            'notas': c.notas,
            'detalles': detalles,
        })
    return JsonResponse(data, safe=False)


@api_key_required
def api_reporte_general(request):
    """Igual que la pantalla /reporte-general/ pero en JSON, para que
    portal_grupointra (Finanzas) sincronice el Reporte General de Recepción
    sin depender del Excel exportado a mano. Regresa las citas del rango tal
    cual (sin aplicar la regla de qué cuenta como ingreso: eso es
    responsabilidad de quien consume la API)."""
    from django.utils.dateparse import parse_date
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    citas = Cita.objects.select_related('paciente', 'terapeuta', 'servicio', 'consultorio', 'division').all()
    if fecha_inicio:
        citas = citas.filter(fecha__gte=parse_date(fecha_inicio))
    if fecha_fin:
        citas = citas.filter(fecha__lte=parse_date(fecha_fin))
    data = [{
        'id': c.id,
        'fecha': c.fecha.isoformat(),
        'hora': c.hora.strftime('%H:%M'),
        'tipo_cita': c.get_tipo_paciente_display(),
        'paciente': c.paciente.nombre if c.paciente_id else '',
        'terapeuta': str(c.terapeuta) if c.terapeuta_id else '',
        'servicio': str(c.servicio) if c.servicio_id else '',
        'division': str(c.division) if c.division_id else '',
        'consultorio': str(c.consultorio) if c.consultorio_id else '',
        'estatus': c.estatus,
        'metodo_pago': c.metodo_pago or '',
        'costo': float(c.costo) if c.costo is not None else 0.0,
    } for c in citas]
    return JsonResponse(data, safe=False)


@api_key_required
def api_disponibilidad_semanal(request):
    from .models import Terapeuta, Horario
    terapeutas = Terapeuta.objects.filter(activo=True)
    data = []
    for t in terapeutas:
        horarios = Horario.objects.filter(terapeuta=t)
        data.append({
            'terapeuta_id': t.id,
            'terapeuta': t.nombre,
            'horarios': [{
                'dia': h.dia,
                'hora_inicio': h.hora_inicio.strftime('%H:%M'),
                'hora_fin': h.hora_fin.strftime('%H:%M'),
                'sede': h.sede,
            } for h in horarios],
        })
    return JsonResponse(data, safe=False)


# ───────────────────────── INSTRUMENTOS — RESPUESTA PÚBLICA ─────────────────────────

def _agrupar_preguntas_display(preguntas):
    """Agrupa preguntas con el mismo `clave` no vacío en bloques de display.

    Retorna lista de dicts:
      {'tipo': 'pregunta', 'pregunta': p}
      {'tipo': 'grupo', 'clave': '...', 'titulo': '...', 'preguntas': [...]}
    """
    items = []
    seen_clave = {}  # clave → index en items (para agrupar)
    for p in preguntas:
        clave = (p.clave or '').strip()
        if clave and clave in seen_clave:
            items[seen_clave[clave]]['preguntas'].append(p)
        elif clave:
            idx = len(items)
            items.append({
                'tipo': 'grupo',
                'clave': clave,
                'titulo': p.titulo_grupo or '',
                'preguntas': [p],
            })
            seen_clave[clave] = idx
        else:
            items.append({'tipo': 'pregunta', 'pregunta': p})
    return items


def _resolver_opcion_instrumento(pregunta, valor_recibido):
    """Busca, dentro de las opciones de la pregunta, la que coincide con el valor
    recibido del formulario y regresa (texto_legible, valor_numerico_para_el_motor)."""
    if not valor_recibido:
        return "", None

    for opcion in (pregunta.opciones or []):
        if str(opcion.get("valor")) == str(valor_recibido):
            etiqueta = opcion.get("etiqueta") or str(valor_recibido)
            try:
                numerico = Decimal(str(opcion.get("valor")))
            except (InvalidOperation, TypeError, ValueError):
                numerico = None
            return etiqueta, numerico

    return valor_recibido, None


def responder_instrumento(request, token):
    """Vista pública (sin login): el paciente abre su link único y contesta el
    instrumento. Se identifica solo por el token UUID — nadie más puede adivinarlo."""
    envio = get_object_or_404(
        EnvioInstrumento.objects.select_related("instrumento", "paciente"),
        token=token,
    )
    instrumento = envio.instrumento

    if envio.estado == EnvioInstrumento.ESTADO_RESPONDIDO:
        return render(request, "clinica/instrumento_publico.html", {
            "envio": envio, "instrumento": instrumento, "pantalla": "ya_respondido",
        })

    if envio.estado == EnvioInstrumento.ESTADO_CANCELADO:
        return render(request, "clinica/instrumento_publico.html", {
            "envio": envio, "instrumento": instrumento, "pantalla": "cancelado",
        })

    preguntas = list(instrumento.preguntas.all().order_by("orden", "id"))

    if request.method == "POST":
        nuevas_respuestas = []
        faltantes = []

        for pregunta in preguntas:
            campo = f"pregunta_{pregunta.id}"

            if pregunta.tipo_respuesta == PreguntaInstrumento.TIPO_OPCION_MULTIPLE:
                seleccionados = request.POST.getlist(campo)
                etiquetas, numericos = [], []
                for valor_crudo in seleccionados:
                    etiqueta, numerico = _resolver_opcion_instrumento(pregunta, valor_crudo)
                    etiquetas.append(etiqueta)
                    if numerico is not None:
                        numericos.append(numerico)
                valor_legible = ", ".join(etiquetas)
                valor_numerico = sum(numericos, Decimal("0")) if numericos else None

            elif pregunta.tipo_respuesta == PreguntaInstrumento.TIPO_TEXTO_LIBRE:
                valor_legible = request.POST.get(campo, "").strip()
                valor_numerico = None

            else:
                valor_crudo = request.POST.get(campo, "").strip()
                valor_legible, valor_numerico = _resolver_opcion_instrumento(pregunta, valor_crudo)

            if pregunta.requerida and not valor_legible:
                faltantes.append(pregunta.id)
                continue

            if valor_legible:
                nuevas_respuestas.append(RespuestaInstrumento(
                    envio=envio, pregunta=pregunta,
                    valor=valor_legible, valor_numerico=valor_numerico,
                ))

        if faltantes:
            import json as _json
            valores_previos = _json.dumps({
                str(p.id): request.POST.get(f"pregunta_{p.id}", "")
                for p in preguntas
                if request.POST.get(f"pregunta_{p.id}", "")
            })
            return render(request, "clinica/instrumento_publico.html", {
                "envio": envio, "instrumento": instrumento,
                "preguntas": preguntas,
                "display_items": _agrupar_preguntas_display(preguntas),
                "pantalla": "formulario",
                "error": "Por favor responde todas las preguntas obligatorias antes de continuar.",
                "faltantes": set(faltantes),
                "valores_previos": valores_previos,
            })

        with transaction.atomic():
            RespuestaInstrumento.objects.filter(envio=envio).delete()
            RespuestaInstrumento.objects.bulk_create(nuevas_respuestas)
            envio.estado = EnvioInstrumento.ESTADO_RESPONDIDO
            envio.respondido_en = timezone.now()
            calcular_resultado_instrumento(envio)
            envio.save()

        return render(request, "clinica/instrumento_publico.html", {
            "envio": envio, "instrumento": instrumento, "pantalla": "gracias",
        })

    return render(request, "clinica/instrumento_publico.html", {
        "envio": envio, "instrumento": instrumento,
        "preguntas": preguntas,
        "display_items": _agrupar_preguntas_display(preguntas),
        "pantalla": "formulario",
    })


# ─────────────────────────────────────────────────────────────────────────
# WhatsApp (Meta Cloud API) — recordatorios, confirmación, encuesta, reactivación
# ─────────────────────────────────────────────────────────────────────────

MAP_WHATSAPP_MANUAL = {
    MensajeWhatsApp.TIPO_RECORDATORIO_5D: ('recordatorio_cita_5_dias',
        lambda d: [d['nombre_paciente'], d['fecha'], d['hora'],
                   d['sucursal'], d['terapeuta'], d['servicio']]),
    MensajeWhatsApp.TIPO_RECORDATORIO_3D: ('recordatorio_cita_3_dias',
        lambda d: [d['nombre_paciente'], d['fecha'], d['hora'],
                   d['sucursal'], d['terapeuta'], d['servicio']]),
    MensajeWhatsApp.TIPO_CONFIRMACION_1D: ('confirmacion_cita_1_dia',
        lambda d: [d['fecha'], d['hora'], d['sucursal'], d['terapeuta'],
                   d['nombre_paciente'], d['servicio'], d['monto'], d['direccion']]),
    MensajeWhatsApp.TIPO_ENCUESTA: ('encuesta_conformidad',
        lambda d: [d['nombre_paciente'], d['servicio'], d['terapeuta']]),
}

# Campo de Cita que se marca al enviar exitosamente, para que el cron no reenvíe.
CAMPO_MARCA_WHATSAPP = {
    MensajeWhatsApp.TIPO_RECORDATORIO_5D: 'recordatorio_5d_enviado_en',
    MensajeWhatsApp.TIPO_RECORDATORIO_3D: 'recordatorio_3d_enviado_en',
    MensajeWhatsApp.TIPO_CONFIRMACION_1D: 'recordatorio_1d_enviado_en',
    MensajeWhatsApp.TIPO_ENCUESTA: 'encuesta_enviada_en',
}


@login_required
@require_POST
def whatsapp_enviar_manual(request, cita_id, tipo):
    """Botón manual en la ficha de cita: envía cualquiera de los 4 tipos de mensaje."""
    if tipo not in MAP_WHATSAPP_MANUAL:
        return JsonResponse({'error': 'Tipo inválido'}, status=400)

    cita = get_object_or_404(
        Cita.objects.select_related('paciente', 'terapeuta', 'servicio', 'consultorio'),
        pk=cita_id,
    )
    if not cita.paciente.telefono:
        return JsonResponse({'error': 'El paciente no tiene teléfono registrado'}, status=400)

    nombre_template, get_params = MAP_WHATSAPP_MANUAL[tipo]
    datos = wa.construir_parametros_cita(cita)

    try:
        resp = wa.enviar_template(cita.paciente.telefono, nombre_template, get_params(datos))
        exitoso = 'messages' in resp
    except Exception as e:
        resp = {'error': str(e)}
        exitoso = False

    campo = CAMPO_MARCA_WHATSAPP.get(tipo)
    if exitoso and campo:
        setattr(cita, campo, timezone.now())
        cita.save(update_fields=[campo])

    MensajeWhatsApp.objects.create(
        cita=cita,
        paciente=cita.paciente,
        telefono=cita.paciente.telefono,
        tipo=tipo,
        origen='manual',
        texto=wa.renderizar_template(nombre_template, get_params(datos)),
        exitoso=exitoso,
        respuesta_api=resp,
        enviado_por=request.user,
    )
    return JsonResponse({'ok': exitoso, 'meta_response': resp})


@login_required
@require_POST
def whatsapp_reactivacion_seguimiento(request):
    """
    Botón en la vista de pacientes sin seguimiento.
    Body JSON: { "paciente_ids": [1, 2, 3] }
    """
    data = json.loads(request.body)
    ids = data.get('paciente_ids', [])
    pacientes = Paciente.objects.filter(pk__in=ids).exclude(telefono='')

    resultados = []
    for paciente in pacientes:
        parametros = [paciente.nombre, 'INTRA']
        try:
            resp = wa.enviar_template(paciente.telefono, 'reactivacion_paciente', parametros)
            exitoso = 'messages' in resp
        except Exception as e:
            resp = {'error': str(e)}
            exitoso = False

        MensajeWhatsApp.objects.create(
            paciente=paciente,
            telefono=paciente.telefono,
            tipo=MensajeWhatsApp.TIPO_REACTIVACION,
            origen='manual',
            texto=wa.renderizar_template('reactivacion_paciente', parametros),
            exitoso=exitoso,
            respuesta_api=resp,
            enviado_por=request.user,
        )
        resultados.append({'paciente_id': paciente.pk, 'ok': exitoso})

    return JsonResponse({'resultados': resultados})


def _registrar_mensaje_entrante(mensaje: dict):
    """Guarda un mensaje entrante de Meta, si no se había registrado ya (Meta reintenta envíos)."""
    wa_message_id = mensaje.get('id')
    if not wa_message_id or MensajeWhatsAppEntrante.objects.filter(wa_message_id=wa_message_id).exists():
        return

    wa_id = mensaje.get('from', '')
    tipo = mensaje.get('type')
    if tipo == 'text':
        texto = mensaje.get('text', {}).get('body', '')
    elif tipo == 'button':
        texto = mensaje.get('button', {}).get('text', '')
    elif tipo == 'interactive':
        interactive = mensaje.get('interactive', {})
        respuesta = interactive.get('button_reply') or interactive.get('list_reply') or {}
        texto = respuesta.get('title', '')
    else:
        texto = ''

    paciente = wa.buscar_paciente_por_wa_id(wa_id)
    MensajeWhatsAppEntrante.objects.create(
        wa_message_id=wa_message_id,
        wa_id=wa_id,
        texto=texto,
        paciente=paciente,
    )

    if paciente and _es_confirmacion(texto):
        _confirmar_cita_pendiente(paciente)


def _es_confirmacion(texto: str) -> bool:
    """True si el texto contiene la palabra CONFIRMO (sin acentos, sin importar mayúsculas)."""
    normalizado = unicodedata.normalize('NFKD', texto or '').encode('ascii', 'ignore').decode('ascii').upper()
    return 'CONFIRMO' in normalizado


def _confirmar_cita_pendiente(paciente):
    """
    Si el paciente responde CONFIRMO después de haber recibido el template de
    confirmación (1 día antes), marca esa cita como confirmada automáticamente.
    Solo aplica a la cita ligada al último envío de ese template, y solo si
    sigue sin confirmar y no es una fecha ya pasada.
    """
    ultimo_envio = (
        MensajeWhatsApp.objects
        .filter(paciente=paciente, tipo=MensajeWhatsApp.TIPO_CONFIRMACION_1D, exitoso=True)
        .select_related('cita')
        .order_by('-enviado_en')
        .first()
    )
    if not ultimo_envio or not ultimo_envio.cita:
        return

    cita = ultimo_envio.cita
    if cita.estatus == Cita.ESTATUS_SIN_CONFIRMAR and cita.fecha >= timezone.localdate():
        cita.estatus = Cita.ESTATUS_CONFIRMADA
        cita.save(update_fields=['estatus'])


@csrf_exempt
def whatsapp_webhook(request):
    """
    GET:  verificación inicial de Meta (handshake)
    POST: mensajes entrantes — se guardan para revisión manual del personal
          (panel "Mensajes recibidos" en whatsapp_recordatorios.html)
    """
    if request.method == 'GET':
        mode = request.GET.get('hub.mode')
        token = request.GET.get('hub.verify_token')
        challenge = request.GET.get('hub.challenge')
        if mode == 'subscribe' and token == settings.WHATSAPP_VERIFY_TOKEN:
            return HttpResponse(challenge, content_type='text/plain')
        return HttpResponse(status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            for entry in data.get('entry', []):
                for change in entry.get('changes', []):
                    for mensaje in change.get('value', {}).get('messages', []):
                        _registrar_mensaje_entrante(mensaje)
        except Exception:
            logging.getLogger(__name__).exception('Error procesando webhook de WhatsApp')
        # Siempre 200: si Meta recibe un error, reintenta el mismo payload repetidamente.
        return HttpResponse(status=200)

    return HttpResponse(status=405)


@api_key_required
@require_POST
def whatsapp_trigger_cron(request):
    """Disparado por cron-job.org (cron gratuito externo) con header X-API-Key."""
    from django.core import management
    management.call_command('enviar_recordatorios_whatsapp')
    return JsonResponse({'ok': True})


@login_required
def pacientes_sin_seguimiento(request):
    """
    Pacientes cuya última cita con ESTATUS_SI_ASISTIO fue antes de `desde`
    y que no tienen ninguna cita futura activa.
    Filtro vía querystring: ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD
    """
    hoy = timezone.localdate()
    desde_str = request.GET.get('desde', '')
    hasta_str = request.GET.get('hasta', '')
    desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str else None
    hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else hoy

    pacientes_con_cita_futura = Cita.objects.filter(
        fecha__gte=hoy,
        estatus__in=Cita.ESTATUS_ACTIVOS,
    ).values_list('paciente_id', flat=True)

    pacientes = (
        Paciente.objects.exclude(telefono='')
        .exclude(pk__in=pacientes_con_cita_futura)
        .annotate(ultima_asistencia=Max(
            'citas__fecha', filter=Q(citas__estatus=Cita.ESTATUS_SI_ASISTIO)
        ))
        .filter(ultima_asistencia__isnull=False, ultima_asistencia__lte=hasta)
    )
    if desde:
        pacientes = pacientes.filter(ultima_asistencia__gte=desde)

    return render(request, 'clinica/pacientes_sin_seguimiento.html', {
        'pacientes': pacientes.order_by('-ultima_asistencia'),
        'desde': desde_str,
        'hasta': hasta_str or hoy.isoformat(),
    })


@login_required
def whatsapp_recordatorios(request):
    """
    Panel para enviar manualmente, por lote, los mensajes que normalmente
    dispara el cron diario: recordatorios (5d/3d), confirmación (1d) y
    encuesta de conformidad, a las citas que aún no los tienen marcados.
    """
    hoy = timezone.localdate()

    base = (
        Cita.objects
        .select_related('paciente', 'terapeuta', 'servicio', 'consultorio')
        .exclude(paciente__telefono='')
    )

    grupos = [
        {
            'tipo': MensajeWhatsApp.TIPO_RECORDATORIO_5D,
            'titulo': 'Recordatorio 5 días',
            'subtitulo': 'Citas del ' + (hoy + timedelta(days=5)).strftime('%d/%m/%Y'),
            'icono': 'bi-calendar-week',
            'color': '#0D6EFD',
            'citas': base.filter(
                fecha=hoy + timedelta(days=5),
                estatus__in=Cita.ESTATUS_ACTIVOS,
                recordatorio_5d_enviado_en__isnull=True,
            ).order_by('hora'),
        },
        {
            'tipo': MensajeWhatsApp.TIPO_RECORDATORIO_3D,
            'titulo': 'Recordatorio 3 días',
            'subtitulo': 'Citas del ' + (hoy + timedelta(days=3)).strftime('%d/%m/%Y'),
            'icono': 'bi-calendar3',
            'color': '#3B82F6',
            'citas': base.filter(
                fecha=hoy + timedelta(days=3),
                estatus__in=Cita.ESTATUS_ACTIVOS,
                recordatorio_3d_enviado_en__isnull=True,
            ).order_by('hora'),
        },
        {
            'tipo': MensajeWhatsApp.TIPO_CONFIRMACION_1D,
            'titulo': 'Confirmación 1 día',
            'subtitulo': 'Citas del ' + (hoy + timedelta(days=1)).strftime('%d/%m/%Y'),
            'icono': 'bi-check2-circle',
            'color': '#0EA5E9',
            'citas': base.filter(
                fecha=hoy + timedelta(days=1),
                estatus__in=Cita.ESTATUS_ACTIVOS,
                recordatorio_1d_enviado_en__isnull=True,
            ).order_by('hora'),
        },
        {
            'tipo': MensajeWhatsApp.TIPO_ENCUESTA,
            'titulo': 'Encuesta de satisfacción',
            'subtitulo': 'Asistencias de los últimos 14 días',
            'icono': 'bi-emoji-smile',
            'color': '#0D9488',
            'citas': base.filter(
                fecha__lt=hoy,
                fecha__gte=hoy - timedelta(days=14),
                estatus=Cita.ESTATUS_SI_ASISTIO,
                encuesta_enviada_en__isnull=True,
            ).order_by('-fecha', 'hora'),
        },
    ]

    # Una "conversación" pendiente = un paciente (o número sin identificar) con
    # al menos un mensaje entrante sin atender. Se agrupa en Python porque el
    # volumen es bajo (mensajes pendientes de un consultorio) y así se evita una
    # subconsulta para traer el último texto por grupo.
    conversaciones_por_clave = {}
    pendientes_qs = (
        MensajeWhatsAppEntrante.objects
        .filter(atendido=False)
        .select_related('paciente')
        .order_by('-recibido_en')
    )
    for mensaje in pendientes_qs:
        clave = mensaje.paciente_id or f'wa:{mensaje.wa_id}'
        conversacion = conversaciones_por_clave.get(clave)
        if conversacion is None:
            conversaciones_por_clave[clave] = {
                'paciente_id': mensaje.paciente_id,
                'wa_id': mensaje.wa_id,
                'nombre': mensaje.paciente.nombre if mensaje.paciente else None,
                'ultimo_texto': mensaje.texto,
                'ultimo_en': mensaje.recibido_en,
                'pendientes': 1,
            }
        else:
            conversacion['pendientes'] += 1

    conversaciones = sorted(conversaciones_por_clave.values(), key=lambda c: c['ultimo_en'], reverse=True)

    return render(request, 'clinica/whatsapp_recordatorios.html', {
        'grupos': grupos,
        'hoy': hoy,
        'automatizacion_activa': ConfiguracionWhatsApp.get_actual().automatizacion_activa,
        'conversaciones': conversaciones,
    })


@login_required
def whatsapp_conversacion(request):
    """
    Historial completo (entrante + saliente) con un paciente, para el panel de chat.
    Recibe ?paciente_id=N o ?wa_id=528... (cuando el número no matcheó a ningún paciente).
    """
    paciente_id = request.GET.get('paciente_id')
    wa_id = request.GET.get('wa_id')

    mensajes = []
    paciente = None
    telefono = wa_id or ''

    if paciente_id:
        paciente = get_object_or_404(Paciente, pk=paciente_id)
        telefono = paciente.telefono
        entrantes = MensajeWhatsAppEntrante.objects.filter(paciente_id=paciente_id)
        salientes = MensajeWhatsApp.objects.filter(paciente_id=paciente_id)
    else:
        entrantes = MensajeWhatsAppEntrante.objects.filter(wa_id=wa_id, paciente__isnull=True)
        salientes = MensajeWhatsApp.objects.none()

    for m in entrantes:
        mensajes.append({'origen': 'entrante', 'texto': m.texto, 'fecha': m.recibido_en})
    for m in salientes:
        mensajes.append({
            'origen': 'saliente',
            'texto': m.texto,
            'fecha': m.enviado_en,
            'tipo': m.get_tipo_display(),
            'exitoso': m.exitoso,
        })

    mensajes.sort(key=lambda m: m['fecha'])
    for m in mensajes:
        m['fecha'] = timezone.localtime(m['fecha']).strftime('%d/%m/%Y %H:%M')

    return JsonResponse({
        'paciente': paciente.nombre if paciente else None,
        'telefono': telefono,
        'mensajes': mensajes,
    })


@login_required
@require_POST
def whatsapp_marcar_atendido(request):
    """
    Marca como atendidos todos los mensajes pendientes de una conversación (un paciente
    o un número sin identificar). Body JSON: { "paciente_id": N } o { "wa_id": "528..." }.
    """
    data = json.loads(request.body)
    paciente_id = data.get('paciente_id')
    wa_id = data.get('wa_id')

    qs = MensajeWhatsAppEntrante.objects.filter(atendido=False)
    qs = qs.filter(paciente_id=paciente_id) if paciente_id else qs.filter(wa_id=wa_id, paciente__isnull=True)

    marcados = qs.update(atendido=True, atendido_por=request.user, atendido_en=timezone.now())
    return JsonResponse({'ok': True, 'marcados': marcados})


@login_required
@require_POST
def whatsapp_toggle_automatizacion(request):
    """Activa/desactiva el envío automático (cron) de recordatorios, confirmaciones y encuestas."""
    config = ConfiguracionWhatsApp.get_actual()
    config.automatizacion_activa = not config.automatizacion_activa
    config.save(update_fields=['automatizacion_activa'])
    return JsonResponse({'automatizacion_activa': config.automatizacion_activa})


@login_required
@require_POST
def whatsapp_enviar_lote(request):
    """
    Envío masivo desde el panel de recordatorios pendientes.
    Body JSON: { "cita_ids": [1, 2, 3], "tipo": "recordatorio_5d" }
    """
    data = json.loads(request.body)
    tipo = data.get('tipo')
    ids = data.get('cita_ids', [])
    if tipo not in MAP_WHATSAPP_MANUAL:
        return JsonResponse({'error': 'Tipo inválido'}, status=400)

    nombre_template, get_params = MAP_WHATSAPP_MANUAL[tipo]
    campo = CAMPO_MARCA_WHATSAPP.get(tipo)

    citas = (
        Cita.objects
        .select_related('paciente', 'terapeuta', 'servicio', 'consultorio')
        .filter(pk__in=ids)
        .exclude(paciente__telefono='')
    )

    resultados = []
    for cita in citas:
        datos = wa.construir_parametros_cita(cita)
        parametros = get_params(datos)
        try:
            resp = wa.enviar_template(cita.paciente.telefono, nombre_template, parametros)
            exitoso = 'messages' in resp
        except Exception as e:
            resp = {'error': str(e)}
            exitoso = False

        if exitoso and campo:
            setattr(cita, campo, timezone.now())
            cita.save(update_fields=[campo])

        MensajeWhatsApp.objects.create(
            cita=cita,
            paciente=cita.paciente,
            telefono=cita.paciente.telefono,
            tipo=tipo,
            origen='manual',
            texto=wa.renderizar_template(nombre_template, parametros),
            exitoso=exitoso,
            respuesta_api=resp,
            enviado_por=request.user,
        )
        resultados.append({'cita_id': cita.pk, 'ok': exitoso})

    return JsonResponse({'resultados': resultados})



@login_required
def demos_whatsapp(request):
    """
    Panel exclusivo de is_superuser: enviar plantillas de WhatsApp a números
    sueltos (prospectos, no ligados a un Paciente) para mostrar el producto
    antes de que un cliente lo contrate. Arranca con la campaña Dorothea.
    """
    if not request.user.is_superuser:
        return redirect('home')

    historial = MensajeWhatsAppDemo.objects.select_related('enviado_por')[:100]
    claves_plantillas = {p['clave'] for c in wa.DEMOS_CAMPANAS.values() for p in c['plantillas']}
    plantilla_textos = {k: v for k, v in wa.TEMPLATE_BODIES.items() if k in claves_plantillas}

    return render(request, 'clinica/demos_whatsapp.html', {
        'campanas': wa.DEMOS_CAMPANAS,
        'campos_labels': wa.CAMPOS_DEMO_LABELS,
        'plantilla_textos': plantilla_textos,
        'historial': historial,
    })


@login_required
@require_POST
def demos_whatsapp_enviar_lote(request):
    """
    Envía en una sola llamada varias combinaciones de (campaña, plantilla,
    teléfono) armadas en la "cola" del panel DEMOS — permite mandar varios
    números y varias plantillas/campañas a la vez.
    Body JSON: { "envios": [
        {"campana": "dorothea", "plantilla": "recordatorio_pago_3_dias",
         "telefono": "8441234567", "campos": {"nombre": "...", ...}},
        ...
    ] }
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    data = json.loads(request.body)
    envios = data.get('envios') or []

    resultados = []
    for envio in envios:
        campana_clave = envio.get('campana', '')
        plantilla_clave = envio.get('plantilla', '')
        telefono = (envio.get('telefono') or '').strip()
        campos = envio.get('campos') or {}

        campana = wa.DEMOS_CAMPANAS.get(campana_clave)
        plantilla = next(
            (p for p in campana['plantillas'] if p['clave'] == plantilla_clave), None
        ) if campana else None

        if not plantilla or not telefono:
            resultados.append({'ok': False, 'telefono': telefono, 'plantilla': plantilla_clave,
                                'error': 'Datos incompletos'})
            continue

        parametros = [campos.get(campo, '') for campo in plantilla['campos']]

        try:
            resp = wa.enviar_template(telefono, plantilla_clave, parametros)
            exitoso = 'messages' in resp
        except Exception as e:
            resp = {'error': str(e)}
            exitoso = False

        MensajeWhatsAppDemo.objects.create(
            campana=campana_clave,
            tipo=plantilla_clave,
            telefono=telefono,
            nombre_contacto=campos.get('nombre', ''),
            texto=wa.renderizar_template(plantilla_clave, parametros),
            exitoso=exitoso,
            respuesta_api=resp,
            enviado_por=request.user,
        )
        resultados.append({
            'ok': exitoso, 'telefono': telefono, 'campana': campana_clave,
            'plantilla': plantilla_clave, 'meta_response': resp,
        })

    return JsonResponse({'resultados': resultados})


def demo_pago_dorothea(request):
    """
    Página de "pago" simulada para el botón de las plantillas recordatorio_pago_*
    de la demo de Dorothea. No procesa ningún pago real, solo muestra los datos
    recibidos por querystring (?nombre=&monto=&referencia=). Pública (sin login)
    porque quien la abre es el prospecto, no el personal de INTRA.
    """
    return render(request, 'clinica/demo_pago_dorothea.html', {
        'nombre': request.GET.get('nombre', 'Alumno(a)'),
        'monto': request.GET.get('monto', '850'),
        'referencia': request.GET.get('referencia', 'DEMO-12345'),
    })
# Vista para dar de alta a un paciente 
@login_required
def toggle_alta_paciente(request, paciente_id):
    paciente = get_object_or_404(Paciente, id=paciente_id)
    if request.method == "POST":
        paciente.dado_de_alta = not paciente.dado_de_alta
        if paciente.dado_de_alta:
            paciente.fecha_alta = timezone.now()
            messages.success(
                request,
                "Paciente dado de alta correctamente."
            )
        else:
            paciente.fecha_alta = None
            messages.success(
                request,
                "Paciente reactivado correctamente."
            )
        paciente.save()
    return redirect(
        "detalle_paciente",
        )

# Vista para dar de alta el paciente
@login_required
@require_POST
def toggle_alta_paciente(request, paciente_id):
    if not request.user.is_superuser:
        messages.error(
            request,
            "No tienes permisos para realizar esta acción."
        )
        return redirect('detalle_paciente', paciente_id=paciente_id)
    paciente = get_object_or_404(
        Paciente,
        pk=paciente_id
    )
    paciente.dado_de_alta = not paciente.dado_de_alta
    if paciente.dado_de_alta:
        paciente.fecha_alta = timezone.now()
        messages.success(
            request,
            "Paciente dado de alta correctamente."
        )
    else:
        paciente.fecha_alta = None
        messages.success(
            request,
            "Seguimiento reactivado correctamente."
        )
    paciente.save()
    return redirect(
        'detalle_paciente',

        paciente_id=paciente.id
    )

from django.http import HttpResponse
# Vista para suspender a un paciente
@login_required
def suspender_paciente(request, id):
    paciente = get_object_or_404(Paciente, id=id)
    if request.method == "POST":
        # Si ya estaba suspendido, se reactiva
        if paciente.estado == "Suspendido":
            paciente.estado = "Activo"
            paciente.fecha_suspension = None
            paciente.motivo_suspension = None
            paciente.suspendido_por = None
            messages.success(
                request,
                "El tratamiento del paciente fue reanudado correctamente."
            )
        # Si estaba activo, se suspende
        else:
            motivo = request.POST.get("motivo")
            paciente.estado = "Suspendido"
            paciente.fecha_suspension = timezone.now()
            paciente.motivo_suspension = motivo
            paciente.suspendido_por = request.user
            messages.success(
                request,
                "Paciente suspendido correctamente."
            )
        paciente.save()
        return redirect(
            "detalle_paciente",
            paciente_id=paciente.id
        )
    return render(
        request,
        "pacientes/suspender_paciente.html",
        {"paciente": paciente}
    )