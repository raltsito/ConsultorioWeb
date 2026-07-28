"""Carga preguntas de SCID-II y SCL-90 desde el Excel PSI.

Uso:
    python manage.py cargar_scid2_scl90
    python manage.py cargar_scid2_scl90 --ruta "C:/ruta/al/archivo.xlsx"
"""
import os

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from clinica.models import Instrumento, PreguntaInstrumento

RUTA_DEFAULT = os.path.join(settings.BASE_DIR, 'docs_xslx_instrumentos', 'scid2_scl90.xlsx')


class Command(BaseCommand):
    help = 'Carga preguntas de SCID-II y SCL-90 desde el Excel PSI'

    def add_arguments(self, parser):
        parser.add_argument(
            '--ruta', default=RUTA_DEFAULT,
            help='Ruta al archivo .xlsx (default: Downloads/PSI)',
        )

    def handle(self, *args, **options):
        ruta = options['ruta']
        self.stdout.write(f'Abriendo: {ruta}')
        wb = openpyxl.load_workbook(ruta, data_only=True)
        self._cargar_scid2(wb)
        self._cargar_scl90(wb)
        self.stdout.write(self.style.SUCCESS('Listo.'))

    # ── helpers ──────────────────────────────────────────────────────────────

    def _texto(self, header):
        """'1. ¿Texto?|radio-57' → '1. ¿Texto?'"""
        if not header:
            return None
        return str(header).split('|')[0].strip() or None

    # ── SCID-II ──────────────────────────────────────────────────────────────

    def _cargar_scid2(self, wb):
        ws = wb['DSCID']
        headers = [cell.value for cell in ws[1]]
        # Columnas 6-124 (índices 5-123) = 119 preguntas
        preguntas_raw = headers[5:124]

        ins, created = Instrumento.objects.update_or_create(
            clave='scid2',
            defaults={
                'nombre': 'SCID-II',
                'descripcion': (
                    'Entrevista clínica estructurada para trastornos de personalidad (DSM). '
                    'Evalúa rasgos y patrones de funcionamiento en 12 trastornos de personalidad.'
                ),
                'instrucciones': (
                    'A continuación encontrará una serie de afirmaciones acerca de cómo actúa, '
                    'piensa o se siente. Para cada enunciado, indique Sí o No según cómo se ha '
                    'sentido generalmente. No existen respuestas buenas ni malas.'
                ),
                'activo': True,
            },
        )
        accion = 'Creado' if created else 'Actualizado'

        ins.preguntas.all().delete()

        opciones = [{'valor': '1', 'etiqueta': 'Sí'}, {'valor': '0', 'etiqueta': 'No'}]
        nuevas = [
            PreguntaInstrumento(
                instrumento=ins, orden=i,
                texto=self._texto(h),
                tipo_respuesta=PreguntaInstrumento.TIPO_SI_NO,
                opciones=opciones,
                requerida=True,
            )
            for i, h in enumerate(preguntas_raw, 1)
            if self._texto(h)
        ]
        PreguntaInstrumento.objects.bulk_create(nuevas)
        self.stdout.write(
            self.style.SUCCESS(f'  SCID-II [{accion}]: {len(nuevas)} preguntas cargadas.')
        )

    # ── SCL-90 ───────────────────────────────────────────────────────────────

    def _cargar_scl90(self, wb):
        ws = wb['DSCL']
        headers = [cell.value for cell in ws[1]]
        # Columnas 9-98 (índices 8-97) = 90 preguntas
        preguntas_raw = headers[8:98]

        ins, created = Instrumento.objects.update_or_create(
            clave='scl90',
            defaults={
                'nombre': 'SCL-90',
                'descripcion': (
                    'Lista de comprobación de 90 síntomas. Evalúa 9 dimensiones psicopatológicas '
                    '(somatización, depresión, ansiedad, etc.) e índices globales de distrés.'
                ),
                'instrucciones': (
                    'A continuación encontrará una lista de problemas y molestias que a veces '
                    'la gente tiene. Lea cada uno con cuidado e indique en qué medida ese '
                    'problema le ha molestado o angustiado durante la última semana, '
                    'incluyendo el día de hoy.'
                ),
                'activo': True,
            },
        )
        accion = 'Creado' if created else 'Actualizado'

        ins.preguntas.all().delete()

        opciones = [
            {'valor': '0', 'etiqueta': 'Nada'},
            {'valor': '1', 'etiqueta': 'Un poco'},
            {'valor': '2', 'etiqueta': 'Bastante'},
            {'valor': '3', 'etiqueta': 'Mucho'},
            {'valor': '4', 'etiqueta': 'Muchísimo'},
        ]
        nuevas = [
            PreguntaInstrumento(
                instrumento=ins, orden=i,
                texto=self._texto(h),
                tipo_respuesta=PreguntaInstrumento.TIPO_ESCALA,
                opciones=opciones,
                requerida=True,
            )
            for i, h in enumerate(preguntas_raw, 1)
            if self._texto(h)
        ]
        PreguntaInstrumento.objects.bulk_create(nuevas)
        self.stdout.write(
            self.style.SUCCESS(f'  SCL-90 [{accion}]: {len(nuevas)} preguntas cargadas.')
        )
